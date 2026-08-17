# -*- coding: utf-8 -*-
"""
Extrator de Tarefas - Microsoft Planner
Plano: Entregas Tecnicas
Filtro: Helbert + Matheus + Wesley + flag MEDICOES
Saida: Demandas_Medicoes_Completo.xlsx

Estrategia: intercepta respostas de rede do Planner (mais confiavel que IDB).
Fallback: leitura direta do IndexedDB.

Dependencias:
    pip install playwright openpyxl
    playwright install chromium

Uso:
    py extrator.py
"""

import json
import os
import re
import sys
import asyncio
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# playwright e importado DENTRO de main(): so a captura precisa dele, e assim as
# funcoes puras (filtro, parsing v1/v4, status) podem ser testadas sem o browser
# instalado — os testes rodam no CI, que nao tem playwright.

# Dumps de diagnostico SO com --debug (ou EXTRATOR_DEBUG=1). Eles contem dado
# de cliente (nome da empresa, numero de OS, GUID de responsavel) e cauda de URL
# do Graph — e este repositorio e PUBLICO. Ficam fora do git pelo .gitignore.
DEBUG = ('--debug' in sys.argv) or os.environ.get('EXTRATOR_DEBUG') == '1'

# Exportar o PLANO INTEIRO quando o payload nao traz dado de categoria (a v4 nao
# manda appliedCategories no bulk). Sem esta flag, tarefa sem categoria e
# DESCARTADA — ver filter_tasks().
ACEITAR_SEM_CATEGORIA = '--sem-filtro-categoria' in sys.argv

# CONFIGURACOES =========================================================
PLAN_ID       = "JOHzljvSKkmfSsQ7SekCnWUAA8cz"
TENANT_ID     = "953ea640-963d-4af5-844c-03f21ebc048f"

HELBERT_ID    = "d7c006ff-d86b-4cf6-9f67-bbaa019f036e"
MATHEUS_ID    = "ba8e6795-6926-4c1f-ae38-aef4df737070"
WESLEY_ID     = "a4256cbb-7af4-4f9e-9c22-d47ee21966b5"
MEDICOES_CAT  = "00000009000000000000000000000000"

PLANNER_URL   = f"https://planner.cloud.microsoft/webui/plan/{PLAN_ID}/view/board"
TASK_URL      = "https://planner.cloud.microsoft/webui/plan/{plan_id}/view/board/task/{task_id}?tid={tid}"

OUTPUT_FILE   = "Demandas_Medicoes_Completo.xlsx"
PROFILE_DIR   = "./playwright_profile"

ASSIGNEES     = {HELBERT_ID, MATHEUS_ID, WESLEY_ID}
ASSIGNEE_NAMES = {
    HELBERT_ID: "Helbert",
    MATHEUS_ID: "Matheus",
    WESLEY_ID:  "Wesley",
}

# Patterns de URL que contem dados de tarefas
TASK_URL_PATTERNS = [
    '/tasks',
    '/task',
    f'plans/{PLAN_ID}',
    '/planner/plans',
    '/buckets',
    'GetAllPlans',
    'GetTasksForPlan',
    'GetBucketsForPlan',
]

# Dominios onde a tarefa pode estar
DOMINIOS_MICROSOFT = [
    'microsoft.com', 'office.com', 'microsoftonline.com',
    'planner', 'taskapi', 'graph.microsoft',
]

# ... e o que NUNCA carrega tarefa: login (payload de autenticacao), telemetria,
# presenca e foto. Sem esta lista o interceptador baixava e parseava tudo isso.
URL_NUNCA_CAPTURAR = [
    'login.microsoftonline.com', 'login.microsoft.com', 'login.live.com',
    'events.data.microsoft.com', 'browser.pipe.aria', 'mobile.pipe.aria',
    '/telemetry', '/beacon', '/ocsp', '/presence', '/avatar', '/photo',
    '/userphoto', 'clientconfig', 'favicon',
]

# Corpo maior que isto nao e lista de tarefa — e bundle/telemetria. Evita
# json.loads em megabytes a toa.
MAX_BODY_BYTES = 5 * 1024 * 1024

# Campos que indicam um objeto de TAREFA REAL (com dados de medição)
TASK_FIELD_INDICATORS = {
    'dueDateTime', 'percentComplete', 'assignments', 'userAssignments',
    'appliedCategories', 'checklist', 'completedDateTime',
}

JS_EXTRACT_COMMENTS = """
() => {
    const container = document.querySelector('[class*="fui-Chat"]');
    if (!container) return null;

    const walker = document.createTreeWalker(
        container,
        NodeFilter.SHOW_TEXT,
        null
    );

    const texts = [];
    let node;
    while ((node = walker.nextNode())) {
        const t = node.textContent.trim();
        if (!t || t.length < 3) continue;
        if (/^\\d{1,2}\\/\\d{1,2}\\s+\\d{1,2}:\\d{2}$/.test(t)) continue;
        if (/^(Hoje|Ontem|Yesterday|Today)$/i.test(t)) continue;
        if (t.length <= 50 && /^[A-Za-zA-u\\s\\.]+$/.test(t) && !t.includes('  ')) continue;
        texts.push(t);
    }
    return texts.join('\\n---\\n');
}
"""

# IDB fallback (mantido como backup)
JS_READ_IDB = """
async () => {
    return new Promise((resolve) => {
        const dbs = indexedDB.databases ? indexedDB.databases() : Promise.resolve([]);
        dbs.then(list => {
            const plannerDb = list.find(d => d.name && (
                d.name.startsWith('PlannerV2_') ||
                d.name.startsWith('PlannerV3_') ||
                d.name.toLowerCase().includes('planner')
            ));
            if (!plannerDb) return resolve({ tasks: [], buckets: [], dbName: null, allDbs: list.map(d=>d.name) });
            const req = indexedDB.open(plannerDb.name);
            req.onsuccess = () => {
                const db = req.result;
                const allStores = Array.from(db.objectStoreNames);
                const readStore = (name) => new Promise(res => {
                    if (!allStores.includes(name)) return res([]);
                    try {
                        const tx = db.transaction(name, 'readonly');
                        const all = tx.objectStore(name).getAll();
                        all.onsuccess = () => res(all.result || []);
                        all.onerror = () => res([]);
                    } catch(e) { res([]); }
                });
                const tName = allStores.find(s => s==='task'||s==='tasks'||s.toLowerCase()==='task') || 'task';
                const bName = allStores.find(s => s==='bucket'||s==='buckets'||s.toLowerCase()==='bucket') || 'bucket';
                Promise.all([readStore(tName), readStore(bName)]).then(([tasks, buckets]) => {
                    resolve({ tasks, buckets, dbName: plannerDb.name, allStores, allDbs: list.map(d=>d.name) });
                });
            };
            req.onerror = () => resolve({ tasks: [], buckets: [], dbName: plannerDb.name, allDbs: list.map(d=>d.name), error: String(req.error) });
        }).catch(e => resolve({ tasks: [], buckets: [], dbName: null, error: String(e) }));
    });
}
"""


# FUNCOES AUXILIARES =====================================================
def parse_date(val) -> str:
    if not val:
        return ""
    if isinstance(val, dict):
        val = val.get("date") or val.get("dateTime") or ""
    if not val:
        return ""
    try:
        s = str(val).replace("Z", "").split("T")[0]
        parts = s.split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return str(val)


def extract_os(nome: str) -> str:
    m = re.match(r'^(\d{5,10})', nome.strip())
    return m.group(1) if m else ""


def extract_cnpj(texto: str) -> str:
    if not texto:
        return ""
    m = re.search(r'\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[\.\s\/]?\d{4}[\.\s\-]?\d{2}', texto)
    return m.group(0).strip() if m else ""


def get_status(task: dict) -> str:
    # v4 tem campo "status" direto: notStarted, inProgress, completed, deferred, waitingOnOthers
    status_v4 = task.get("status", "")
    if status_v4 == "completed" or task.get("completedDateTime"):
        return "Concluida"
    if status_v4 == "inProgress":
        return "Em andamento"
    # Traduzido: esta coluna e lida por gente (Gabriel/Luiz Fernando) e o
    # importador so reconhece 'conclu*' — enum cru em ingles nao dizia nada.
    if status_v4 == "deferred":
        return "Adiada"
    if status_v4 == "waitingOnOthers":
        return "Aguardando terceiro"
    pct = task.get("percentComplete", 0)
    if pct == 100:
        return "Concluida"
    if pct == 50:
        return "Em andamento"
    if pct == 0:
        return "Nao iniciada"
    return f"{pct}%"


def build_checklist(task: dict):
    checklist = task.get("checklist", {})
    if not checklist:
        return "", ""
    items = list(checklist.values()) if isinstance(checklist, dict) else checklist
    linhas, done = [], 0
    for item in items:
        title = item.get("title", "")
        checked = item.get("isChecked", False)
        linhas.append(f"{'OK' if checked else '[ ]'} {title}")
        if checked:
            done += 1
    return "\n".join(linhas), f"{done}/{len(items)}"


def get_assignee_ids(task: dict) -> set:
    """Extrai IDs de responsáveis do task, suportando v1 e v4."""
    uids = set()
    # v1: assignments = {userId: {...}}
    a1 = task.get("assignments", {})
    if isinstance(a1, dict):
        uids.update(a1.keys())
    # v4: userAssignments = [{user: {id: "uuid-com-dashes"}, id: "uuid-sem-dashes", ...}]
    a2 = task.get("userAssignments", [])
    if isinstance(a2, list):
        for item in a2:
            if not isinstance(item, dict):
                continue
            # Prioridade: user.id (com dashes, formato padrão UUID)
            uid = (item.get("user") or {}).get("id", "")
            if not uid:
                # fallback: item.id pode ser sem dashes — normaliza inserindo dashes
                raw = item.get("id", "")
                if len(raw) == 32:
                    uid = f"{raw[0:8]}-{raw[8:12]}-{raw[12:16]}-{raw[16:20]}-{raw[20:]}"
                else:
                    uid = raw
            if uid:
                uids.add(uid)
    return uids


def build_assignees(task: dict) -> str:
    uids = get_assignee_ids(task)
    if not uids:
        return ""
    # sorted() porque get_assignee_ids devolve set: sem isso a ordem dos nomes
    # muda entre execucoes e dois exports do mesmo dia acusam diferenca falsa
    # (o importador ainda grava esse texto como `contato` da empresa).
    return ", ".join(sorted(ASSIGNEE_NAMES.get(uid, uid[:8] + "...") for uid in uids))


def build_labels(task: dict) -> str:
    # v1 format
    cats = task.get("appliedCategories", {})
    if isinstance(cats, dict) and cats:
        return ", ".join(k for k, v in cats.items() if v)
    # v4 format: categories as list
    cats_v4 = task.get("categories", [])
    if isinstance(cats_v4, list) and cats_v4:
        return ", ".join(str(c) for c in cats_v4)
    pv = task.get("planView", {})
    cat_ids = pv.get("appliedCategoryIds", []) if pv else []
    if cat_ids:
        return ", ".join(cat_ids)
    return ""


def get_task_title(task: dict) -> str:
    """Retorna o título do task (v1: title, v4: displayName)."""
    return task.get("title") or task.get("displayName") or task.get("taskName") or ""


def merge_tasks(raw_list: list) -> dict:
    """Merge task data by id — later entries overwrite earlier (more complete data wins)."""
    merged = {}
    for t in raw_list:
        tid = t.get("id") or t.get("taskId") or t.get("@odata.id", "")
        if not tid:
            continue
        if tid in merged:
            # Merge: update with non-null values
            for k, v in t.items():
                if v is not None and v != "" and v != {} and v != []:
                    merged[tid][k] = v
        else:
            merged[tid] = dict(t)
    return merged


def filter_tasks(task_map: dict, bucket_map: dict, aceitar_sem_categoria=None):
    print(f"   Total de tarefas capturadas: {len(task_map)}")
    plan_ids = set(t.get("planId", "") for t in task_map.values())
    print(f"   Plan IDs encontrados: {plan_ids}")

    if aceitar_sem_categoria is None:
        aceitar_sem_categoria = ACEITAR_SEM_CATEGORIA

    filtered = []
    sem_plano = sem_resp = sem_cat = sem_dado_cat = 0

    # Salva amostra para debug de estrutura (so com --debug: contem dado de cliente)
    if DEBUG:
        sample_tasks = list(task_map.values())[:3]
        try:
            with open("debug_task_sample.json", "w", encoding="utf-8") as _f:
                json.dump(sample_tasks, _f, ensure_ascii=False, indent=2)
            print(f"   [DEBUG] Amostra de {len(sample_tasks)} tasks salva em debug_task_sample.json")
        except Exception as _e:
            print(f"   [DEBUG] Nao foi possivel salvar amostra: {_e}")

    for t in task_map.values():
        if t.get("planId") != PLAN_ID:
            sem_plano += 1
            continue
        uids = get_assignee_ids(t)
        if not any(uid in ASSIGNEES for uid in uids):
            sem_resp += 1
            continue
        # Verifica categoria MEDICOES — suporta v1 e v4
        pv = t.get("planView", {})
        cat_ids = pv.get("appliedCategoryIds", []) if pv else []
        applied = t.get("appliedCategories", {})
        cats_v4 = t.get("categories", [])
        if isinstance(cats_v4, list):
            cat_ids = cat_ids + [str(c) for c in cats_v4]
        has_any_cat_data = bool(cat_ids or applied or cats_v4)
        if has_any_cat_data:
            has_cat = (
                MEDICOES_CAT in cat_ids
                or (isinstance(applied, dict) and applied.get(MEDICOES_CAT, False))
            )
            if not has_cat:
                sem_cat += 1
                continue
        elif not aceitar_sem_categoria:
            # Tarefa do plano SEM nenhum dado de categoria no payload (a v4 nao
            # manda appliedCategories no bulk). Aceitar aqui DESLIGAVA o filtro que
            # define o escopo do extrator: o plano tem ~8000 tarefas e so ~143 tem
            # a label MEDICOES — o export virava o plano inteiro (treinamento,
            # PPR/PCA, ergonomia) e o importador criava demanda de medicao pra tudo.
            sem_dado_cat += 1
            continue
        filtered.append(t)

    print(f"   Excluidas: {sem_plano} sem plano | {sem_resp} sem responsavel | {sem_cat} sem cat MEDICOES")
    if sem_dado_cat:
        print(f"   [ATENCAO] {sem_dado_cat} tarefas do plano vieram SEM dado de categoria no")
        print(f"             payload e foram DESCARTADAS — nao da para saber se sao Medicoes.")
        print(f"             A label vive em appliedCategories ({MEDICOES_CAT}), que a API v4")
        print(f"             nao manda no bulk. Para exportar o plano INTEIRO mesmo assim:")
        print(f"             py extrator.py --sem-filtro-categoria")
    if aceitar_sem_categoria:
        print(f"   [ATENCAO] --sem-filtro-categoria ligado: entrou tarefa do plano SEM a label")
        print(f"             MEDICOES. O xlsx NAO e mais so de medicoes — nao importar direto.")
    print(f"   Filtradas: {len(filtered)} tarefas")

    if not filtered and task_map:
        plano = [t for t in task_map.values() if t.get("planId") == PLAN_ID]
        if plano:
            t0 = plano[0]
            print(f"\n   AMOSTRA completa (primeiras chaves):")
            print(f"     chaves: {list(t0.keys())[:20]}")
            print(f"     titulo: {get_task_title(t0)[:60]}")
            print(f"     userAssignments: {t0.get('userAssignments', 'N/A')}")
            print(f"     assignments: {t0.get('assignments', 'N/A')}")
            print(f"     categories: {t0.get('categories', t0.get('appliedCategories','N/A'))}")
            print(f"     => Rode com --debug para gravar debug_task_sample.json")

    return filtered, bucket_map


# INTERCEPTACAO DE REDE ==================================================
def should_capture(url: str) -> bool:
    """Captura JSON de dominio Microsoft, MENOS o que nunca carrega tarefa.

    O filtro largo existe de proposito: os endpoints da v4 nao sao conhecidos, e
    exigir a lista TASK_URL_PATTERNS deixaria a captura cega. Mas 'qualquer coisa
    da Microsoft' incluia o endpoint de LOGIN e a telemetria — corpo de resposta
    baixado e parseado a toa, incluindo payload de autenticacao.
    """
    url_lower = url.lower()
    if any(b in url_lower for b in URL_NUNCA_CAPTURAR):
        return False
    return any(d in url_lower for d in DOMINIOS_MICROSOFT)


def is_real_task(item: dict) -> bool:
    """Verifica se um objeto é uma tarefa real (não bucket/plano interno)."""
    keys = set(item.keys())
    # Tarefa real tem pelo menos um desses campos
    if keys & TASK_FIELD_INDICATORS:
        return True
    # Na v4, tarefa tem 'displayName' MAS version deve conter 'Task' (não só Bucket)
    version = str(item.get("version", ""))
    if "Bucket" in version and "Task" not in version:
        return False  # é um bucket interno da v4
    # Tarefa v4 sem prazo, sem responsável e sem etiqueta não tem NENHUM campo de
    # TASK_FIELD_INDICATORS — só o `version`. Sem esta linha ela não era tarefa nem
    # bucket e sumia sem entrar em contador nenhum (o extrator relatava 0 tarefas).
    if "Task" in version:
        return True
    return False


def is_bucket(item: dict) -> bool:
    """Identifica buckets (colunas do quadro)."""
    version = item.get("version", "")
    if "Bucket" in version and "Task" not in version:
        return True
    # Formato v1: tem name + orderHint mas não percentComplete
    return ("orderHint" in item and "name" in item
            and "percentComplete" not in item and "bucketId" not in item)


def extract_items_from_json(data, url: str = "") -> tuple:
    """Extrai tasks e buckets de um payload JSON. Retorna (tasks, buckets)."""
    tasks, buckets = [], []

    def process_item(item):
        if not isinstance(item, dict):
            return
        if is_real_task(item):
            tasks.append(item)
        elif is_bucket(item):
            buckets.append(item)
        # v4 bucket com displayName = bucket (coluna)
        elif "displayName" in item and "planId" in item:
            version = item.get("version", "")
            if "Bucket" in version:
                # Tratar como bucket da v4
                buckets.append({
                    "id": item.get("id", ""),
                    "name": item.get("displayName", ""),
                    "planId": item.get("planId", ""),
                })

    if isinstance(data, dict):
        # OData collection
        for key in ("value", "tasks", "task", "items"):
            val = data.get(key, [])
            if isinstance(val, list):
                for item in val:
                    process_item(item)
        for key in ("buckets", "bucket"):
            val = data.get(key, [])
            if isinstance(val, list):
                for item in val:
                    if isinstance(item, dict):
                        buckets.append(item)
        # Objeto direto
        process_item(data)
    elif isinstance(data, list):
        for item in data:
            process_item(item)

    return tasks, buckets


# COMENTARIOS ============================================================
async def collect_comments(page: "Page", tasks_with_chat: list) -> dict:
    comments = {}
    total = len(tasks_with_chat)
    for i, task in enumerate(tasks_with_chat, 1):
        task_id = task.get("id") or task.get("taskId", "")
        if not task_id:
            continue
        url = TASK_URL.format(plan_id=PLAN_ID, task_id=task_id, tid=TENANT_ID)
        print(f"  [{i}/{total}] Tarefa {task_id[:12]}...")
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            try:
                await page.wait_for_selector('[class*="fui-Chat"]', timeout=25000)
                await asyncio.sleep(3)
            except Exception:
                print(f"    AVISO: Chat nao carregou")
                comments[task_id] = ""
                continue
            result = await page.evaluate(JS_EXTRACT_COMMENTS)
            comments[task_id] = result or ""
            count = len([c for c in (result or "").split("---") if c.strip()])
            print(f"    {count} comentario(s)")
        except Exception as e:
            print(f"    ERRO: {e}")
            comments[task_id] = ""
    return comments


# EXCEL ==================================================================
def generate_excel(tasks: list, bucket_map: dict, comments: dict, output: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Demandas Medicoes"

    headers = [
        "Nome da Tarefa", "OS", "Data de Criacao", "Data de Prazo",
        "Data de Conclusao", "Responsavel(is)", "Status", "Progresso (%)",
        "Checklist", "Checklist Progresso", "Grupo (Bucket)",
        "Etiquetas", "Descricao", "CNPJ", "Tem Comentarios", "Comentarios"
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    col_widths = [50, 12, 14, 14, 14, 25, 15, 12, 40, 16, 25, 20, 60, 20, 16, 80]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    for row_idx, task in enumerate(tasks, 2):
        task_id = task.get("id") or task.get("taskId", "")
        nome = get_task_title(task)
        # notes: v1 = {value: "..."} | v4 = "texto plano" ou None
        notes_raw = task.get("notes") or task.get("description") or task.get("notePreviewText") or ""
        if isinstance(notes_raw, dict):
            descricao = notes_raw.get("value") or notes_raw.get("text") or ""
        else:
            descricao = str(notes_raw) if notes_raw else ""
        checklist_txt, checklist_prog = build_checklist(task)
        comentarios = comments.get(task_id, "")

        row_data = [
            nome,
            extract_os(nome) or task.get("numeroOS", ""),
            parse_date(task.get("createdDateTime") or (task.get("creationInfo") or {}).get("createdDateTime")),
            parse_date(task.get("dueDateTime") or task.get("scheduledDateTime") or task.get("deadline")),
            parse_date(task.get("completedDateTime")),
            build_assignees(task),
            get_status(task),
            task.get("percentComplete", 0),
            checklist_txt,
            checklist_prog,
            bucket_map.get(task.get("bucketId", ""), ""),
            build_labels(task),
            descricao,
            extract_cnpj(descricao),
            "Sim" if comentarios.strip() else "Nao",
            comentarios,
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(output)
    print(f"\nOK: Excel salvo: {output} ({len(tasks)} linhas)")


# MAIN ==================================================================
async def main():
    print("=" * 60)
    print("  EXTRATOR PLANNER - Entregas Tecnicas / MEDICOES")
    print("  Estrategia: interceptacao de rede")
    print("=" * 60)

    # Buffers para dados interceptados
    tasks_raw:   list = []
    buckets_raw: list = []
    captured_urls: list = []

    profile_path = Path(PROFILE_DIR).resolve()
    profile_path.mkdir(parents=True, exist_ok=True)

    # Import tardio: so a captura precisa do playwright (ver topo do arquivo).
    from playwright.async_api import async_playwright

    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=str(profile_path),
            headless=False,
            args=["--start-maximized"],
            no_viewport=True,
        )

        page = context.pages[0] if context.pages else await context.new_page()

        # Registra interceptador ANTES de navegar
        all_json_urls: list = []  # todos os URLs com JSON capturado

        async def on_response(resp):
            url = resp.url
            if not should_capture(url):
                return
            try:
                ct = resp.headers.get("content-type", "")
                if "json" not in ct:
                    return
                body = await resp.body()
                if not body or len(body) < 50 or len(body) > MAX_BODY_BYTES:
                    return
                data = json.loads(body)
                t_new, b_new = extract_items_from_json(data, url)
                # Guarda URL com JSON para debug (mesmo sem tarefas)
                if DEBUG and isinstance(data, dict) and ("value" in data or "id" in data):
                    all_json_urls.append(url[-100:])
                if t_new or b_new:
                    tasks_raw.extend(t_new)
                    buckets_raw.extend(b_new)
                    captured_urls.append(url)
                    print(f"   [NET] +{len(t_new)} tarefas, +{len(b_new)} buckets | {url[-80:]}")
            except Exception:
                pass

        page.on("response", lambda r: asyncio.ensure_future(on_response(r)))

        print(f"\n1. Abrindo o Planner...")
        try:
            await page.goto(PLANNER_URL, wait_until="commit", timeout=30000)
        except Exception:
            pass  # Redirect para login eh normal

        print("\n" + "=" * 60)
        print("  JANELA ABERTA.")
        print("")
        print("  Se pedir LOGIN: faca login na conta Microsoft.")
        print("  Aguarde o Planner mostrar as tarefas no quadro.")
        print("  IMPORTANTE: clique no plano 'Entregas Tecnicas'")
        print("  e aguarde TODOS os cartoes aparecerem.")
        print("")
        print("  Os dados serao capturados automaticamente da rede.")
        print("  Quando as tarefas estiverem visiveis, volte aqui.")
        print("=" * 60)

        loop = asyncio.get_event_loop()
        await loop.run_in_executor(None, input, "\n  >> ENTER quando as tarefas estiverem visiveis: ")

        # Aguarda requisicoes pendentes terminarem
        print("   Aguardando ultimas requisicoes de rede (5s)...")
        await asyncio.sleep(5)

        print(f"\n2. Dados capturados via rede:")
        print(f"   URLs interceptadas: {len(captured_urls)}")
        print(f"   Tarefas brutas: {len(tasks_raw)}")
        print(f"   Buckets brutos: {len(buckets_raw)}")
        # Salva todos os URLs JSON para diagnóstico (so com --debug)
        if DEBUG:
            try:
                with open("debug_all_urls.json", "w", encoding="utf-8") as _f:
                    json.dump(all_json_urls, _f, ensure_ascii=False, indent=2)
                print(f"   [DEBUG] {len(all_json_urls)} URLs JSON salvos em debug_all_urls.json")
            except Exception:
                pass

        # Fallback para IDB se nenhuma tarefa capturada
        if not tasks_raw:
            print("\n   AVISO: Nenhuma tarefa capturada via rede. Tentando IDB...")
            try:
                idb = await page.evaluate(JS_READ_IDB)
                tasks_raw.extend(idb.get("tasks", []))
                buckets_raw.extend(idb.get("buckets", []))
                print(f"   IDB: {len(idb.get('tasks',[]))} tarefas, banco={idb.get('dbName')}")
                print(f"   IDB stores: {idb.get('allStores', [])}")
            except Exception as e:
                print(f"   ERRO IDB: {e}")

        # Monta mapa de buckets
        bucket_map: dict = {}
        for b in buckets_raw:
            bid = b.get("id") or b.get("bucketId", "")
            bname = b.get("name", "")
            if bid and bname:
                bucket_map[bid] = bname

        # Deduplica e merge tarefas
        task_map = merge_tasks(tasks_raw)

        print(f"\n3. Filtrando tarefas...")
        filtered, bucket_map = filter_tasks(task_map, bucket_map)

        if not filtered:
            print("\n   ERRO: Nenhuma tarefa encontrada apos filtro.")
            print("   Dicas:")
            print("   - Certifique-se de que o plano 'Entregas Tecnicas' carregou")
            print("   - Role a pagina para baixo para carregar todos os cartoes")
            print("   - Aguarde o Planner sincronizar completamente")
            print(f"\n   Total de tasks (sem filtro): {len(task_map)}")
            if task_map:
                sample = next(iter(task_map.values()))
                print(f"   Amostra: {sample.get('title','')[:50]}, planId={sample.get('planId','')}")
            await context.close()
            return

        tasks_with_chat = [t for t in filtered if t.get("hasActiveUpdates") or t.get("hasChat")]
        print(f"\n4. Tarefas com comentarios: {len(tasks_with_chat)}")

        comments = {}
        if tasks_with_chat:
            print("   Coletando comentarios...")
            comments = await collect_comments(page, tasks_with_chat)

        print(f"\n5. Gerando Excel...")
        generate_excel(filtered, bucket_map, comments, OUTPUT_FILE)

        await context.close()
        print("\nConcluido!")


if __name__ == "__main__":
    asyncio.run(main())
