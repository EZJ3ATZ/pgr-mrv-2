import os, re, shutil, zipfile, io, tempfile, base64
from datetime import datetime
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory
from flask_login import login_required, current_user
import xml.etree.ElementTree as ET

try:
    from pdfminer.high_level import extract_text as pdf_extract
    PDF_OK = True
except ImportError:
    PDF_OK = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max (fotos)

def _secret_key():
    """Chave de sessão. Nunca usa a chave publica do repo em producao.
    Ordem: SECRET_KEY (env) -> derivada do DATABASE_URL (secreto, estavel entre
    restarts, unico por deploy) -> fallback de dev (so quando nao ha Postgres)."""
    k = os.environ.get('SECRET_KEY')
    if k:
        return k
    db_url = os.environ.get('DATABASE_URL')
    if db_url:
        import hashlib
        print('[app] SECRET_KEY ausente — derivando de DATABASE_URL. '
              'Defina SECRET_KEY no Railway para desacoplar a chave do banco.')
        return hashlib.sha256(('sst-session::' + db_url).encode()).hexdigest()
    return 'dev-key-troque-em-producao'

app.secret_key = _secret_key()

# ── Hardening: cookie de sessão + headers de segurança ────────────────
# Achado do blackbox 10/07 (CWE-693 / OWASP A05): produção sem nenhum
# header de segurança; HSTS ausente sobre o login permite SSL-strip.
_EM_PRODUCAO = bool(os.environ.get('RAILWAY_ENVIRONMENT'))

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_EM_PRODUCAO,   # local roda em http
    REMEMBER_COOKIE_HTTPONLY=True,
    REMEMBER_COOKIE_SAMESITE='Lax',
    REMEMBER_COOKIE_SECURE=_EM_PRODUCAO,
)

# CSP pragmática: cobre o que os templates realmente usam (fonts Google,
# chart.js no jsdelivr, spline-viewer no unpkg, logo em ocupacional.com.br).
# 'unsafe-inline'/'unsafe-eval' necessários: JS inline nos templates + WASM
# do spline. Ainda bloqueia script de origem não listada, clickjacking
# (frame-ancestors), form-action e <base> externos.
_CSP = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline' 'unsafe-eval' "
    "https://cdn.jsdelivr.net https://unpkg.com; "
    "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
    "font-src 'self' data: https://fonts.gstatic.com; "
    "img-src 'self' data: blob: https:; "
    "connect-src 'self' https:; "
    "worker-src 'self' blob:; "
    "frame-ancestors 'self'; "
    "base-uri 'self'; "
    "form-action 'self'; "
    "object-src 'none'"
)

@app.after_request
def _security_headers(resp):
    h = resp.headers
    h.setdefault('X-Content-Type-Options', 'nosniff')
    h.setdefault('X-Frame-Options', 'SAMEORIGIN')
    h.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    h.setdefault('Permissions-Policy',
                 'camera=(), microphone=(), geolocation=(), payment=()')
    h.setdefault('Content-Security-Policy', _CSP)
    # HSTS só quando a resposta sai por https (Railway termina TLS no proxy)
    if request.is_secure or request.headers.get('X-Forwarded-Proto') == 'https':
        h.setdefault('Strict-Transport-Security',
                     'max-age=31536000; includeSubDomains')
    return resp

# ── Observabilidade: logging estruturado + Sentry ─────────────────────
try:
    from controle.monitoring import setup_logging, init_sentry
    setup_logging()
    init_sentry(app)
except Exception as _mon_err:
    import logging
    logging.basicConfig(level=logging.INFO)
    logging.getLogger(__name__).warning(f'[monitoring] não iniciado: {_mon_err}')

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
TPL_DIR    = os.path.join(BASE_DIR, 'tpl')
MODEL_DIR  = os.path.join(BASE_DIR, 'modelo_unpacked')

# ── Modulo Controle de Medicoes e Amostradores (isolado via Blueprint) ─
try:
    from controle import controle_bp, auth_bp, login_manager, init_db as _controle_init_db
    from controle.db import get_db, row_to_dict, registrar_evento
    login_manager.init_app(app)
    app.register_blueprint(auth_bp)
    app.register_blueprint(controle_bp)
    _controle_init_db()
except Exception as _e:
    import traceback
    print(f'[controle] erro ao carregar modulo: {_e}')
    traceback.print_exc()



# ── Log de tempo de resposta de cada interação ─────────────────────────
# Mede TODA requisição (app + blueprints) sem tocar em nenhuma view: duração,
# quanto foi banco e quantas consultas o pedido disparou. Leitura só admin, em
# Saúde do Sistema. Desligar com PERF_LOG=0 se algum dia atrapalhar.
try:
    from controle.perf import init_app as _perf_init
    _perf_init(app)
except Exception as _pe:
    print(f'[perf] não iniciado: {_pe}')


# ── Service Worker servido na RAIZ ────────────────────────────────────
# Em /static/sw.js o escopo seria só /static/ — não controlaria /campo nem
# /mobile (offline nunca funcionava nessas telas). Servindo em /sw.js o
# escopo vira "/" e o SW passa a interceptar todas as páginas.
@app.route('/sw.js')
def _root_service_worker():
    resp = app.send_static_file('sw.js')
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp

# ── Rede de segurança global: handlers de erro 404/500 ────────────────
# Sem isto, qualquer erro não-tratado ou URL inválida devolve HTML padrão
# do Flask — e o frontend (ctrlFetch) espera JSON, quebrando a tela.
import logging as _logging
_err_log = _logging.getLogger('app.errors')

def _quer_json():
    """True se a requisição é de uma rota de API (deve responder JSON)."""
    try:
        p = request.path or ''
        if p.startswith('/controle') or p.startswith('/api'):
            return True
        if request.is_json:
            return True
        accept = request.headers.get('Accept', '')
        if 'application/json' in accept:
            return True
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return True
    except Exception:
        pass
    return False

# ── Trava global: role=visualizador é somente leitura ─────────────────
# Vale para TODOS os blueprints e rotas do app.
# /auth/ fica fora para o visualizador conseguir logar e trocar a própria senha.
@app.before_request
def _bloquear_escrita_visualizador():
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    if request.path.startswith('/auth/'):
        return None
    try:
        from flask_login import current_user as _cu
        if _cu.is_authenticated and getattr(_cu, 'role', '') == 'visualizador':
            return jsonify({'erro': 'Seu perfil é somente leitura — ação não permitida.',
                            'status': 403}), 403
    except Exception:
        pass
    return None

@app.errorhandler(404)
def _handle_404(e):
    if _quer_json():
        return jsonify({'erro': 'Recurso não encontrado', 'status': 404,
                        'path': request.path}), 404
    return e  # navegação normal: deixa o Flask renderizar a página padrão

@app.errorhandler(500)
def _handle_500(e):
    _err_log.error(f'[500] {request.method} {request.path}', exc_info=True)
    if _quer_json():
        return jsonify({'erro': 'Erro interno do servidor. Tente novamente.',
                        'status': 500, 'path': request.path}), 500
    return e

@app.errorhandler(Exception)
def _handle_uncaught(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e  # 404/redirect/etc. — Flask trata
    _err_log.error(f'[EXC] {request.method} {request.path}: {e}', exc_info=True)
    if _quer_json():
        return jsonify({'erro': 'Erro interno do servidor. Tente novamente.',
                        'status': 500, 'detalhe': str(e), 'path': request.path}), 500
    return e

# ── Scheduler: sync automático Microsoft Planner ──────────────────────
_scheduler_started = False
def _start_planner_scheduler():
    global _scheduler_started
    if os.environ.get('DISABLE_SCHEDULER') == '1':
        # Testes e scripts importam app.py sem querer sync em background.
        print('[scheduler] DISABLE_SCHEDULER=1 — sync automatico desligado')
        return
    if _scheduler_started:
        return
    _scheduler_started = True
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.interval import IntervalTrigger
        from apscheduler.triggers.date import DateTrigger
        from controle.planner_sync import sync_planner
        from controle.graph import graph_ok

        SYNC_INTERVAL_MINUTES = int(os.environ.get('PLANNER_SYNC_INTERVAL', '15'))
        PLANNER_GROUP_ID = '4c80214b-6801-414a-9fc7-27feff0b3de6'

        def _sync_job():
            if not graph_ok():
                return
            try:
                stats = sync_planner(label_filter='Medições', group_filter=PLANNER_GROUP_ID)
                print(f'[scheduler] Planner sync: {stats.get("criadas",0)} criadas, '
                      f'{stats.get("atualizadas",0)} atualizadas, '
                      f'{stats.get("erros",[]).__len__()} erros')
            except Exception as e:
                print(f'[scheduler] Planner sync erro: {e}')

        from datetime import datetime as _dt, timedelta as _td
        scheduler = BackgroundScheduler(daemon=True)

        # Boot-sync: roda 1x, 60s após o startup — restaura dados do Planner após redeploy
        scheduler.add_job(
            _sync_job,
            trigger=DateTrigger(run_date=_dt.now() + _td(seconds=60)),
            id='planner_boot_sync',
            name='Planner Boot Sync (1x)',
            replace_existing=True,
        )

        # ── Alerta do log (só para o Matheus — ALERTA_PARA) ───────────────
        # "quebrou" a cada 15 min; digest 1x/dia. As travas anti-ruído (mudança
        # de estado, dedup, teto diário, ALERTA_EMAIL=0) estão em controle/alerta.py.
        def _alerta_job(digest=False):
            try:
                from controle.alerta import verificar
                r = verificar(forcar_digest=digest)
                if r.get('quebrou') or r.get('digest'):
                    print(f'[alerta] {r.get("quebrou")} digest={r.get("digest")} '
                          f'suprimidos={r.get("suprimidos")}')
            except Exception as e:
                print(f'[alerta] erro: {e}')

        scheduler.add_job(
            _alerta_job,
            trigger=IntervalTrigger(minutes=15,
                                    start_date=_dt.now() + _td(minutes=5)),
            id='alerta_quebrou', name='Alerta — quebrou', replace_existing=True,
            max_instances=1,
        )
        # 10h UTC = 07h de Brasília. O container roda em UTC — a mesma pegadinha
        # do pg_cron do CRM, que dispara em GMT.
        from apscheduler.triggers.cron import CronTrigger as _Cron
        scheduler.add_job(
            lambda: _alerta_job(digest=True),
            trigger=_Cron(hour=10, minute=0),
            id='alerta_digest', name='Alerta — digest diário', replace_existing=True,
            max_instances=1,
        )

        scheduler.add_job(
            _sync_job,
            trigger=IntervalTrigger(
                minutes=SYNC_INTERVAL_MINUTES,
                start_date=_dt.now() + _td(minutes=SYNC_INTERVAL_MINUTES),
            ),
            id='planner_sync',
            name='Microsoft Planner Sync',
            replace_existing=True,
            max_instances=1,
        )
        # Consistência: roda 1x/dia às 06:00 UTC
        def _consistencia_job():
            try:
                from controle.consistencia import run_consistencia_geral
                r = run_consistencia_geral()
                print(f'[scheduler] Consistência: {r.get("divergencias_novas",0)} divergências novas')
            except Exception as e:
                print(f'[scheduler] Consistência erro: {e}')

        from apscheduler.triggers.cron import CronTrigger
        scheduler.add_job(
            _consistencia_job,
            trigger=CronTrigger(hour=6, minute=0),
            id='consistencia_diaria',
            name='Consistência Operacional Diária',
            replace_existing=True,
            max_instances=1,
        )

        # Lab: varredura dos e-mails do lab (envio→data_envio_lab, RA→resultado).
        # LEVE (só corpo+inbox, sem baixar anexos) a cada 3h; PESADA (com anexos
        # das cadeias/laudos) 1x/dia às 05:30 UTC.
        def _lab_job(parse_anexos):
            if not graph_ok():
                return
            try:
                from controle.lab_inbox import sincronizar_lab
                r = sincronizar_lab(apply=True, parse_anexos=parse_anexos)
                print(f'[scheduler] Lab sync (anexos={parse_anexos}): '
                      f'{r.get("aplicadas",0)} status, {r.get("envio_auto_datados",0)} envios, '
                      f'{r.get("resultado_auto_datados",0)} resultados, {r.get("mailboxes_lidas",0)} caixas')
            except Exception as e:
                print(f'[scheduler] Lab sync erro: {e}')

        scheduler.add_job(
            lambda: _lab_job(False),
            trigger=IntervalTrigger(hours=3, start_date=_dt.now() + _td(minutes=8)),
            id='lab_sync_leve', name='Lab Sync Leve (3h)',
            replace_existing=True, max_instances=1,
        )
        scheduler.add_job(
            lambda: _lab_job(True),
            trigger=CronTrigger(hour=5, minute=30),
            id='lab_sync_pesada', name='Lab Sync Pesada (diária, com anexos)',
            replace_existing=True, max_instances=1,
        )

        # Backfill de RAs: varre TODO o histórico do lab (via $search), lê o PDF do
        # laudo e conclui o amostrador parado no lab com RA já emitido. É o que fecha
        # o ciclo sozinho (antes só existia no botão manual). 1x/dia às 05:45 UTC,
        # logo após o lab sync pesado.
        def _ra_backfill_job():
            if not graph_ok():
                return
            try:
                from controle.lab_inbox import backfill_ras
                r = backfill_ras(apply=True)
                print(f'[scheduler] RA backfill: {r.get("concluiriam",0)} concluídos, '
                      f'{r.get("casaram",0)} casados, {r.get("medicoes_baixadas",0)} medições')
            except Exception as e:
                print(f'[scheduler] RA backfill erro: {e}')

        scheduler.add_job(
            _ra_backfill_job,
            trigger=CronTrigger(hour=5, minute=45),
            id='ra_backfill_diario', name='Backfill RAs (diário)',
            replace_existing=True, max_instances=1,
        )

        scheduler.start()
        print(f'[scheduler] boot-sync em 60s + sync a cada {SYNC_INTERVAL_MINUTES} minutos + consistência diária 06h + lab leve 3h/pesada 05h30 + backfill RAs 05h45 iniciado')
    except ImportError:
        print('[scheduler] APScheduler nao instalado — sync automatico desabilitado')
    except Exception as e:
        print(f'[scheduler] erro ao iniciar scheduler: {e}')

# Iniciar scheduler após o app estar pronto (evita duplo-start no debug)
import atexit
with app.app_context():
    _start_planner_scheduler()

MESES_PT = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
            7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

def mes_ano():
    d = datetime.now()
    return f"{MESES_PT[d.month]} / {d.year}"

# ── GHE ─────────────────────────────────────────────────────────────
# ════════════════════════════════════════════════════════════════════
# GHE_AGENTES
# Fonte confirmada com data: GHE Oregon - Residencial Oregon (20/02/2024)
# GHEs novos sem data confirmada (Mata das Borboletas): listados em GHE_SEM_DATA
# ════════════════════════════════════════════════════════════════════

# GHEs cujas medições não têm data confirmada — campo data aparece como ???
GHE_SEM_DATA = {
    "PAISAGISMO",
    "SERVICOS_GERAIS",
    "ASSIST_TEC_ELETRICA",
    "ASSIST_TEC_ESPEC",
}

GHE_AGENTES = {
    # Ruído 87,49 | PNOS-Resp 0,90 — Oregon 20/02/2024
    "ACABAMENTO":[
        ('ruido','87,49 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,90 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 83,49 Moderado | PNOS-Total 2,66
    "GESSO_REJUNTE":[
        ('ruido','76,06 dB(A)','Baixo',False,False),
        ('quant','Poeira não Fibrogênica (PNOS-Total)','Químico','10','5','2,66 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 65,25 Baixo | Monotonia | Postura Sentada
    "ADMINISTRATIVO":[
        ('ruido','74,88 dB(A)','Baixo',False,False),
        ('ergon','Monotonia','Risco Baixo'),
        ('ergon','Postura Sentada','Risco Baixo'),
    ],
    # Ruído 73,84 Baixo | Posturas | Queda
    "ALMOXARIFADO":[
        ('ruido','83,57 dB(A)','Moderado',True,False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 72,83 Baixo | PNOS 0,02 | Posturas | Queda
    "APOIO_PRODUCAO":[
        ('ruido','81,77 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,15 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 83,85 Moderado | PNOS 0,76 | Posturas | Perfurocortantes | Queda
    "ARMACAO":[
        ('ruido','88,98 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,19 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Objetos Perfurocortantes','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 84,47 Moderado | Poeira Madeira 0,40 | PNOS 0,23 | Posturas | Perfurocortantes | Queda — Parque Canoas 2025
    "CARPINTARIA":[
        ('ruido','84,47 dB(A)','Moderado',True,False),
        ('quant','Poeira de Madeira','Químico','1','0,5','0,40 mg/m³','Risco Baixo',False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,23 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Objetos Perfurocortantes','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Central Betoneira — mantido do Oregon (não aparece no Mata das Borboletas)
    "CENTRAL_BETONEIRA":[
        ('ruido','85,06 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,62 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
    ],
    # Ruído 82,09 Moderado | PNOS 0,38 | Posturas | Queda
    "ELETRICA":[
        ('ruido','81,46 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,38 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 80,07 Moderado | PNOS 0,33 | Posturas | Queda
    "ESTRUTURA_ALVENARIA":[
        ('ruido','81,57 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,33 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 93,71 Alto | PNOS 0,39 | Esforço | Posturas | Queda | Altura NR35
    "ESTRUTURA_PAREDE":[
        ('ruido','96,63 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,28 mg/m³','Risco Baixo',False),
        ('ergon','Esforço Físico Intenso','Risco Moderado'),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
        ('acid','Trabalho em Altura - NR35','Risco Moderado'),
    ],
    # Ruído 82,79 Moderado | MEK 13mg | PNOS 0,83 | Posturas | Queda
    "HIDRAULICA":[
        ('ruido','81,90 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,28 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 81,09 Moderado | Vibração AREN 0,81 | Vibração VDVR 16,6 | PNOS 1,71 Moderado | Postura Sentada
    "MAQUINAS_GERAL":[
        ('ruido','87,79 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,74 mg/m³','Risco Baixo',False),
        ('ergon','Postura Sentada','Risco Baixo'),
    ],
    # Ruído 88,48 Alto | Vibração AREN 0,86/VDVR 18,3 Moderado | PNOS 0,78 | Postura Sentada — Parque Canoas 2025
    "MAQUINAS_PEQUENO_PORTE":[
        ('ruido','88,48 dB(A)','Alto',True,True),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,78 mg/m³','Risco Baixo',False),
        ('ergon','Postura Sentada','Risco Baixo'),
    ],
    # Ruído 78,48 Baixo | PNOS 0,17 | Posturas | Queda
    "MAQUINAS_ESTAC":[
        ('ruido','81,51 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,07 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 70,43 Baixo | PNOS 1,87 Moderado | Esforço | Posturas | Queda
    "OPERACIONAL":[
        ('ruido','80,29 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,19 mg/m³','Risco Baixo',False),
        ('ergon','Esforço Físico Intenso','Risco Moderado'),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Paisagismo — Ruído 64,79 Baixo | Posturas | Queda — Parque Canoas 2025 (sem PNOS medido)
    "PAISAGISMO":[
        ('ruido','64,79 dB(A)','Baixo',False,False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 77,60 Baixo | PNOS 0,47 | Posturas | Queda
    "PINTURA":[
        ('ruido','82,54 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,47 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 72,31 Baixo | PNOS 0,15 | Monotonia
    "PORTARIA":[
        ('ruido','68,66 dB(A)','Baixo',False,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,10 mg/m³','Risco Baixo',False),
        ('ergon','Monotonia','Risco Baixo'),
    ],
    # Polivalente — mantido do Oregon
    "POLIVALENTE":[
        ('ruido','84,52 dB(A)','Moderado',True,False),
        ('quant','Poeira de Madeira','Químico','1','0,5','0,33 mg/m³','Risco Baixo',False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,05 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Objetos Perfurocortantes','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Ruído 84,12 Moderado | PNOS 0,17 | Posturas | Perfurocortantes | Queda
    "SERRALHERIA":[
        ('ruido','82,64 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,19 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Objetos Perfurocortantes','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Serviços Gerais — NOVO: sem ruído; químicos (limpeza) + Posturas
    "SERVICOS_GERAIS":[
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,12 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
    ],
    # Ruído 81,22 Moderado | PNOS 0,44 | Posturas | Queda
    "SUPERVISAO":[
        ('ruido','81,65 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,44 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Queda de Objetos','Risco Moderado'),
    ],
    # Assistência Técnica Elétrica — Ruído 84,52 | PNOS 0,05 | Posturas | Eletricidade — Parque Canoas 2025
    "ASSIST_TEC_ELETRICA":[
        ('ruido','84,52 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,05 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
        ('acid','Eletricidade','Risco Baixo'),
    ],
    # Assistência Técnica Especializada — Ruído 84,52 | PNOS 0,05 | Posturas — Parque Canoas 2025
    "ASSIST_TEC_ESPEC":[
        ('ruido','84,52 dB(A)','Moderado',True,False),
        ('quant','Poeira não Fibrogênica (PNOS-Respirável)','Químico','2.640','1.320','0,05 mg/m³','Risco Baixo',False),
        ('ergon','Posturas Incomodas','Risco Baixo'),
    ],
}

CARGOS_SUGESTOES = sorted([
    "Ajudante Armador","Ajudante Eletricista","Ajudante Geral","Ajudante Pratico",
    "Almoxarife","Almoxarife Pleno","Analista Administrativo","Apontador",
    "Armador","Armador Pleno","Assistente Administrativo",
    "Assistente Tecnico Edificacoes","Assistente Tecnico Seguranca Trabalho",
    "Auxiliar Administrativo","Auxiliar Almoxarife","Auxiliar Armador",
    "Auxiliar Bombeiro Hidraulico","Auxiliar Carpinteiro",
    "Auxiliar de Limpeza","Auxiliar de Servicos Gerais",
    "Auxiliar Eletricista","Auxiliar Encanador","Auxiliar Engenharia",
    "Auxiliar Ferreiro","Auxiliar Limpeza","Auxiliar Montador",
    "Auxiliar Montador Formas Metalicas","Auxiliar Obras","Auxiliar Pedreiro",
    "Auxiliar Pintor","Auxiliar Producao","Auxiliar Seguranca Trabalho",
    "Auxiliar Tecnico Seguranca Trabalho","Azulejista",
    "Bombeiro Hidraulico","Bombeiro Pleno","Cabo Turma",
    "Carpinteiro","Carpinteiro Forma","Carpinteiro Pleno","Carpinteiro Polivalente",
    "Carpinteiro Serrador","Contra Mestre",
    "Eletricista","Eletricista Instalador Predial","Eletricista Pleno",
    "Eletricista Pos Entrega","Encanador",
    "Encarregado","Encarregado Acabamento","Encarregado Almoxarife",
    "Encarregado Armador","Encarregado Carpintaria","Encarregado de Obras",
    "Encarregado Eletrica","Encarregado Forma","Encarregado Geral",
    "Encarregado Geral Instalacoes","Encarregado Geral Obras",
    "Encarregado Hidraulica","Encarregado Instalacoes","Encarregado Obras Forma",
    "Encarregado Obras Instalacoes","Encarregado Turma",
    "Engenheiro","Engenheiro Junior","Engenheiro Pleno","Engenheiro Senior",
    "Estagiario","Faxineiro","Ferramenteiro","Ferreiro",
    "Gesseiro","Gesseiro Pleno","Guariteiro","Jardineiro","Ladrilheiro","Marceneiro",
    "Meio Oficial","Meio Oficial Armador","Meio Oficial Azulejista",
    "Meio Oficial Bombeiro","Meio Oficial Carpinteiro","Meio Oficial Eletricista",
    "Meio Oficial Encanador","Meio Oficial Ferreiro","Meio Oficial Gesseiro",
    "Meio Oficial Montador","Meio Oficial Montador Formas Metalicas",
    "Meio Oficial Pedreiro","Meio Oficial Pintor","Meio Oficial Pos Entrega",
    "Mestre Geral Obras","Mestre Obras",
    "Montador","Montador Andaimes","Montador de Forma Metalica",
    "Montador Esquadrias","Montador Formas Metalicas Pleno",
    "Oficial","Oficial Pleno","Oficial Polivalente",
    "Operador Betoneira","Operador Cremalheira","Operador Elevador Carga",
    "Operador Equipamentos","Operador Grua","Operador Guincho",
    "Operador Maquinas Geral","Operador Maquinas Leves","Operador Maquinas Pesadas",
    "Pedreiro","Pedreiro Acabamento","Pedreiro I","Pedreiro Pleno",
    "Pintor","Pintor Pleno","Porteiro","Profissional Lider",
    "Profissional Pos Entrega","Profissional Pos Entrega Polivalente",
    "Rejuntador","Serralheiro","Servente","Sinaleiro","Soldador",
    "Supervisor Instalacoes","Tecnico Ambiental","Tecnico Edificacoes",
    "Tecnico Instalacoes","Tecnico Seguranca Trabalho I","Tecnico Seguranca Trabalho II",
    "Topografo","Vigia","Vigia Noturno",
], key=str.lower)

def get_ghe(cargo):
    """Retorna o GHE do cargo (Mata das Borboletas) ou None se não reconhecido."""
    c = cargo.upper()
    # Acabamento (azulejista, ladrilheiro, pedreiro acabamento)
    if any(x in c for x in ["AZULEJ","LADRILH","PEDREIRO ACABAMENTO"]): return "ACABAMENTO"
    # Gesso/Rejunte
    if any(x in c for x in ["GESSEIRO","REJUNT"]): return "GESSO_REJUNTE"
    # Administrativo
    if any(x in c for x in ["ADMIN","ANALISTA","ENGENHEIRO","TOPOGRAFO","APONTADOR","APRENDIZ"]): return "ADMINISTRATIVO"
    # Almoxarifado (inclui ferramenteiro — mas Encarregado Almoxarife vai pra Supervisão)
    if "ENCARREGADO ALMOXARIFE" in c: return "SUPERVISAO"
    if any(x in c for x in ["ALMOXARIFE","FERRAMENTEIRO"]): return "ALMOXARIFADO"
    # Apoio Produção (técnicos, estagiários, segurança do trabalho)
    if any(x in c for x in ["TECNICO EDIF","TECNICO SEGUR","TECNICO AMBI","TECNICO INSTAL",
                              "ESTAGIARIO","ASSISTENTE TECNICO","AUXILIAR SEGUR",
                              "AUXILIAR TECNICO SEGUR"]): return "APOIO_PRODUCAO"
    # Armação (inclui ferreiro, ajudante armador)
    if any(x in c for x in ["ARMADOR","AUXILIAR ARMAD","AJUDANTE ARMAD","FERREIRO","ARMADOR PLENO"]): return "ARMACAO"
    # Carpintaria (inclui marceneiro, carpinteiro polivalente)
    if any(x in c for x in ["CARPINT","SERRAD","MARCENEIRO"]): return "CARPINTARIA"
    # Central Betoneira
    if "BETONEIRA" in c: return "CENTRAL_BETONEIRA"
    # Pós-entrega elétrica especificamente
    if "ELETRICISTA POS ENTREGA" in c: return "ASSIST_TEC_ELETRICA"
    # Profissional pós-entrega (polivalente ou especializado)
    if "POS ENTREGA" in c: return "ASSIST_TEC_ESPEC"
    # Elétrica (eletricista — não encarregado, não pós-entrega)
    if "ELETRIC" in c and "ENCARREGADO" not in c: return "ELETRICA"
    # Hidráulica (inclui encanador, bombeiro, auxiliar)
    if any(x in c for x in ["BOMBEIRO","ENCANADOR","HIDRAUL","AUXILIAR ENCANADOR",
                              "AUXILIAR BOMBEIRO","BOMBEIRO PLENO"]): return "HIDRAULICA"
    # Máquinas leves (pequeno porte) — deve vir ANTES de MAQUINAS_GERAL
    if any(x in c for x in ["MAQUINAS LEVE","MAQUINAS LEVES","MAQUINAS PEQUENO"]): return "MAQUINAS_PEQUENO_PORTE"
    # Máquinas grandes/pesadas
    if any(x in c for x in ["OPERADOR MAQUIN","MAQUINAS GERAL","MAQUINAS PESAD",
                              "OPERADOR EQUIP"]): return "MAQUINAS_GERAL"
    # Máquinas estacionárias (cremalheira, grua, guincho, sinaleiro, elevador)
    if any(x in c for x in ["CREMALHEIRA","GRUA","SINALEIRO","GUINCHO",
                              "ELEVADOR CARGA","ELEVADOR FORM"]): return "MAQUINAS_ESTAC"
    # Paisagismo
    if "JARDINEIRO" in c: return "PAISAGISMO"
    # Polivalente
    if "POLIVALENTE" in c: return "POLIVALENTE"
    # Pintura (inclui auxiliar pintor, pintor pleno)
    if any(x in c for x in ["PINTOR","AUXILIAR PINT"]): return "PINTURA"
    # Portaria (inclui guariteiro)
    if any(x in c for x in ["PORTEIRO","VIGIA","GUARITEIRO"]): return "PORTARIA"
    # Serralheria (inclui soldador)
    if any(x in c for x in ["SERRALHEIRO","SOLDADOR"]): return "SERRALHERIA"
    # Serviços Gerais (limpeza, faxina)
    if any(x in c for x in ["FAXINEIRO","AUXILIAR LIMPEZA","AUXILIAR SERVICOS GERAIS"]): return "SERVICOS_GERAIS"
    # Supervisão (encarregados, mestres, cabo turma, supervisor)
    if any(x in c for x in ["ENCARREGADO","MESTRE","PROFISSIONAL LIDER",
                              "SUPERVISOR","CABO TURMA"]): return "SUPERVISAO"
    # Estrutura Parede (montadores de formas metálicas)
    if any(x in c for x in ["MONTADOR FORMA","MONTADOR FORMAS","AUXILIAR MONTADOR",
                              "MEIO OFICIAL MONTADOR"]): return "ESTRUTURA_PAREDE"
    if c.strip() in ["MONTADOR", "MONTADOR DE FORMA METALICA"]: return "ESTRUTURA_PAREDE"
    # Estrutura Alvenaria (pedreiro, meio oficial, montador andaimes/esquadrias, oficial)
    if any(x in c for x in ["PEDREIRO","MEIO OFICIAL PEDREIRO","MEIO OFICIAL",
                              "AUXILIAR PEDREIRO","MONTADOR ANDAIME","MONTADOR ESQUADRIA",
                              "OFICIAL PLENO","PEDREIRO PLENO","PEDREIRO I"]): return "ESTRUTURA_ALVENARIA"
    if c.strip() == "OFICIAL": return "ESTRUTURA_ALVENARIA"
    # Operacional (servente, ajudante, auxiliar obras/produção)
    if any(x in c for x in ["SERVENTE","AJUDANTE GERAL","AJUDANTE PRATICO",
                              "AUXILIAR DE LIMPEZA","AUXILIAR DE SERVICO",
                              "AUXILIAR OBRAS","AUXILIAR PRODUCAO"]): return "OPERACIONAL"
    if "AJUDANTE" in c: return "OPERACIONAL"
    # Cargo não reconhecido
    return None

DESCRICOES = {
    "PEDREIRO":"Executar serviços de acabamento e reparos de blocos e superfícies concretadas, assentamento de tijolos, reboco e arremates de estruturas construídas, preparação de argamassa de diversos tipos, colocação de telhas, lajes pré-moldadas, pisos, azulejos, ferragens, manilhas, bancadas e peças sanitárias, de acordo com orientações e solicitações recebidas do superior imediato.",
    "MEIO OFICIAL PEDREIRO":"Auxiliar os oficiais na realização de obras de edificações de paredes, pisos, construções em alvenarias, concretagem, cimentados, revestimentos entre outros trabalhos da construção civil.",
    "MEIO OFICIAL":"Auxiliar os oficiais na realização de obras de edificações de paredes, pisos, construções em alvenarias, concretagem, cimentados e revestimentos.",
    "SERVENTE":"Realizar limpeza e organização do canteiro de obras, transporte de materiais e equipamentos, auxiliar as demais frentes de trabalho conforme solicitação do superior imediato.",
    "AUXILIAR DE LIMPEZA":"Realizar a limpeza e organização das instalações do canteiro de obras e áreas sociais, garantindo boas condições de higiene e segurança.",
    "AUXILIAR DE SERVICOS GERAIS":"Realizar atividades diversas de apoio ao canteiro de obras, incluindo limpeza, organização e suporte às atividades dos oficiais.",
    "AJUDANTE PRATICO":"Auxiliar nas atividades do canteiro de obras, transporte de materiais, limpeza e suporte às atividades dos oficiais.",
    "REJUNTADOR":"Realizar serviços de rejuntamento em pisos, paredes, azulejos e outros elementos de construção civil.",
    "PINTOR":"Realizar a pintura de superfícies internas e externas de edificações, preparar superfícies, aplicar tintas, vernizes e outros revestimentos conforme especificações técnicas.",
    "MEIO OFICIAL PINTOR":"Auxiliar os pintores na preparação e pintura de superfícies, mistura de materiais e aplicação de revestimentos.",
    "ARMADOR":"Realizar corte, dobragem e montagem de armações de aço para estruturas de concreto armado, seguindo projetos e especificações técnicas.",
    "AUXILIAR ARMADOR":"Auxiliar os armadores no corte, dobragem e montagem de armações de aço para estruturas de concreto armado.",
    "CARPINTEIRO":"Executar serviços de carpintaria em obras de construção civil, incluindo montagem de formas para concreto, andaimes e outras estruturas de madeira.",
    "MEIO OFICIAL CARPINTEIRO":"Auxiliar os carpinteiros na montagem de formas para concreto e outras estruturas de madeira.",
    "ELETRICISTA":"Realizar a instalação, manutenção e reparo de sistemas elétricos em obras de construção civil, seguindo as normas técnicas de segurança vigentes.",
    "MEIO OFICIAL ELETRICISTA":"Auxiliar os eletricistas nas atividades de instalação, manutenção e reparo de sistemas elétricos.",
    "GESSEIRO":"Executar serviços de revestimento com gesso, incluindo aplicação de reboco, massa corrida e outros acabamentos em paredes e tetos.",
    "AZULEJISTA":"Assentar azulejos, cerâmicas, pastilhas e outros revestimentos em pisos e paredes, seguindo projeto e especificações técnicas.",
    "BOMBEIRO HIDRAULICO":"Realizar instalação, manutenção e reparo de sistemas hidráulicos e de esgoto em obras de construção civil.",
    "MEIO OFICIAL BOMBEIRO":"Auxiliar os bombeiros hidráulicos nas instalações e manutenções de sistemas hidráulicos e sanitários.",
    "MONTADOR":"Montar e desmontar formas metálicas e estruturas para concretagem de paredes e lajes.",
    "MONTADOR DE FORMA METALICA":"Montar e desmontar formas metálicas para concretagem de paredes e lajes em construção civil.",
    "ENCARREGADO":"Supervisionar e coordenar as equipes de trabalho no canteiro de obras, garantindo prazos, qualidade e segurança.",
    "ENCARREGADO DE OBRAS":"Supervisionar e coordenar as equipes de trabalho no canteiro de obras, garantindo prazos, qualidade e segurança.",
    "ENCARREGADO ELETRICA":"Supervisionar e coordenar a equipe de elétrica no canteiro de obras, garantindo qualidade e segurança nas instalações.",
    "PORTEIRO":"Controlar o acesso de pessoas e veículos ao canteiro de obras, garantindo a segurança das instalações.",
    "VIGIA":"Realizar a vigilância e proteção do canteiro de obras durante os períodos determinados.",
    "TOPOGRAFO":"Realizar levantamentos topográficos, controle de nível e demarcação de pontos para execução de obras.",
    "AUXILIAR ADMINISTRATIVO":"Executar atividades administrativas de apoio às operações do canteiro de obras.",
    "ENGENHEIRO":"Planejar, coordenar e supervisionar as atividades de engenharia no canteiro de obras, garantindo qualidade técnica e segurança.",
}

def get_desc(cargo):
    for k, v in DESCRICOES.items():
        if cargo.upper() == k:
            return v
    return f"Executar as atividades inerentes ao cargo de {cargo} no canteiro de obras, conforme orientações do superior imediato e normas de segurança."

# ── Templates ────────────────────────────────────────────────────────
# ATENÇÃO: valores com espaço trailing replicam o formato exato do XML
ORIG = {
    "ruido_db":   "81,57 dB(A) ",          # tem espaço trailing no XML
    "pnos_agent": "Poeira não Fibrogênica (PNOS-Respirável)",
    "pnos_medi":  "0,10 mg/m³ ",           # tem espaço trailing no XML
    "pnos_lt":    "2.640",
    "pnos_na":    "1.320",
    "ergon1":     "Posturas",
    "ergon2":     "Incomodas",
    "acid1":      "Queda",
    "acid2":      "de",
    "acid3":      "Objetos",
    "cargo_desc": "Executar serviços de acabamento e reparos de blocos e superfícies concretadas, assentamento de tijolos, reboco e arremates de estruturas construídas, preparação de argamassa de diversos tipos, colocação de telhas, lajes pré-moldadas, pisos, azulejos, ferragens, manilhas, bancadas e peças sanitárias, de acordo com orientações e solicitações recebidas do superior imediato.",
    "ruido_fund": "Acima do nível de ação, conforme NR-09 da Portaria 3214/78 do M.T.E.",
    "ruido_key":  "Foi identificada a exposição ao agente ruído, sendo necessário.",
    "ruido_rec":  "Recomendamos o uso de protetor auditivo e a realização dos\u00a0exames médicos, audiometria, face o que estabelece a NR-07 (PCMSO) da Portaria 3214 do M.T.E., diante do nível de ruído ter ultrapassado o\u00a0limite de tolerância (LT)\u00a0de\u00a085 dB(A).",
    "pnos_key":   "Foi identificado a exposição às poeiras (PNOS) no canteiro de obra, sendo necessário:",
    "pnos_fund":  "Concentração abaixo do Nível de Ação. ", # tem espaço trailing no XML
    "data":       "Março / 2026",
}

def load_tpl(name):
    with open(os.path.join(TPL_DIR, name + '.xml'), 'r', encoding='utf-8') as f:
        return f.read()

def xs(s): return s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
def xclean(s): return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', s or '')
def spacer(): return '\n    <w:p><w:pPr><w:spacing w:after="0" w:line="240" w:lineRule="auto"/></w:pPr></w:p>\n'

def _replace_run_text(xml, old_text, new_text):
    """Substitui texto dentro de um run XML, respeitando espaços trailing."""
    return xml.replace(f'>{old_text}<', f'>{new_text}<')

def adapt_ruido(db_val, nivel, acima_acao, acima_lt):
    t = load_tpl('ruido')
    # Substituir o valor de dB — o novo valor também precisa do espaço trailing
    t = _replace_run_text(t, ORIG["ruido_db"], db_val + ' ')
    if nivel == 'Baixo': t = t.replace('>Moderado<', '>Baixo<', 1)
    elif nivel == 'Alto': t = t.replace('>Moderado<', '>Alto<', 1)
    if acima_lt:
        t = t.replace(ORIG["ruido_fund"],
                      'Acima do Limite de Tolerância (LT) de 85 dB(A), conforme NR-15 Anexo 1 da Portaria 3214/78 do M.T.E.')
        t = t.replace(ORIG["ruido_rec"],
                      'Uso obrigatório de protetor auditivo (CA válido). Audiometria conforme NR-07 (PCMSO). Medidas de controle de engenharia na fonte geradora.')
    elif not acima_acao:
        t = t.replace(ORIG["ruido_fund"],
                      'Abaixo do nível de ação, conforme NR-09 da Portaria 3214/78 do M.T.E.')
        t = t.replace(ORIG["ruido_key"],
                      'O valor de ruído avaliado está abaixo do nível de ação. Manter monitoramento periódico.')
        t = t.replace(ORIG["ruido_rec"],
                      'Manter boas práticas de higiene ocupacional. Monitorar periodicamente os níveis de ruído.')
    else:
        # Acima do nível de ação porém ABAIXO do LT — o boilerplate do template
        # ("ter ultrapassado o LT de 85 dB(A)") estaria errado aqui
        t = t.replace('ter ultrapassado o limite de tolerância (LT) de 85 dB(A).',
                      'não ter ultrapassado o limite de tolerância (LT) de 85 dB(A), '
                      'porém ter atingido o nível de ação.')
    return t

def adapt_quant(nome_ag, grupo, lt_val, na_val, medicao, nivel_risco, acima_na):
    t = load_tpl('pnos')
    t = t.replace(ORIG["pnos_agent"], nome_ag)
    t = _replace_run_text(t, ORIG["pnos_lt"], lt_val)
    t = _replace_run_text(t, ORIG["pnos_na"], na_val)
    # Medição também tem espaço trailing
    t = _replace_run_text(t, ORIG["pnos_medi"], medicao + ' ')
    n = nivel_risco.replace('Risco ', '')
    if n == 'Baixo': t = t.replace('>Moderado<', '>Baixo<', 1)
    elif n == 'Alto': t = t.replace('>Moderado<', '>Alto<', 1)
    if 'Madeira' in nome_ag:
        t = t.replace(ORIG["pnos_key"],
                      'Foi identificada a exposição à poeira de madeira no canteiro de obra, sendo necessário:')
    elif 'Total' in nome_ag:
        t = t.replace(ORIG["pnos_key"],
                      'Foi identificada a exposição às poeiras (PNOS-Total) no canteiro de obra, sendo necessário:')
    if acima_na:
        t = t.replace(ORIG["pnos_fund"], 'Concentração acima do Nível de Ação. Adotar medidas de controle. ')
    return t

def adapt_ergon(nome_ag, nivel_risco):
    t = load_tpl('ergonomico')
    n = nivel_risco.replace('Risco ', '')
    if nome_ag != 'Posturas Incomodas':
        t = _replace_run_text(t, ORIG["ergon1"], xs(nome_ag))
        t = _replace_run_text(t, ORIG["ergon2"], '')
    if n == 'Baixo': t = t.replace('>Moderado<', '>Baixo<', 1)
    elif n == 'Alto': t = t.replace('>Moderado<', '>Alto<', 1)
    return t

def adapt_acid(nome_ag, nivel_risco):
    t = load_tpl('acidente')
    n = nivel_risco.replace('Risco ', '')
    if nome_ag == 'Objetos Perfurocortantes':
        t = _replace_run_text(t, ORIG["acid1"], 'Objetos Perfurocortantes')
        t = _replace_run_text(t, ORIG["acid2"], '')
        t = _replace_run_text(t, ORIG["acid3"], '')
    elif nome_ag == 'Trabalho em Altura - NR35':
        t = _replace_run_text(t, ORIG["acid1"], 'Trabalho em Altura - NR35')
        t = _replace_run_text(t, ORIG["acid2"], '')
        t = _replace_run_text(t, ORIG["acid3"], '')
    elif nome_ag == 'Eletricidade':
        t = _replace_run_text(t, ORIG["acid1"], 'Eletricidade')
        t = _replace_run_text(t, ORIG["acid2"], '')
        t = _replace_run_text(t, ORIG["acid3"], '')
    if n == 'Baixo': t = t.replace('>Moderado<', '>Baixo<', 1)
    elif n == 'Alto': t = t.replace('>Moderado<', '>Alto<', 1)
    return t

def build_cargo_section(cargo, cidade, uf):
    """
    Constrói a seção de um cargo com seus riscos.
    - GHE_SEM_DATA: datas aparecem como ??? (medição sem data confirmada)
    - GHE reconhecido com data: usa 20/02/2024 (Oregon)
    """
    ghe = get_ghe(cargo)
    ghe_desconhecido = ghe is None
    if ghe is None:
        ghe = 'OPERACIONAL'  # fallback — endpoint já bloqueou antes de chegar aqui

    # Verificar se este GHE tem data de medição conhecida
    sem_data = ghe in GHE_SEM_DATA

    agentes = GHE_AGENTES.get(ghe, GHE_AGENTES['OPERACIONAL'])

    sc = load_tpl('setor_cargo')
    sc = sc.replace('Setor: Ribeirão das Neves - MG</w:t>', f'Setor: {cidade} - {uf}</w:t>')
    sc = sc.replace('RIBEIRÃO DAS NEVES - MG</w:t>', f'{cidade.upper()} - {uf.upper()}</w:t>')
    sc = sc.replace('>Cargo: Pedreiro<', f'>Cargo: {xs(cargo)}<')
    sc = sc.replace('>Cargo: Pedreiro</w:t>', f'>Cargo: {xs(cargo)}</w:t>')
    sc = sc.replace(ORIG["cargo_desc"], get_desc(cargo))

    titulo_texto = f'Especificação dos Riscos - Cargo: {xs(cargo)} '
    titulo = load_tpl('titulo_riscos').replace(
        'Especificação dos Riscos - Cargo: Pedreiro ', titulo_texto)

    risk = ''
    for ag in agentes:
        if ag[0] == 'ruido':
            _, db, nivel, aa, alt = ag
            tbl = adapt_ruido(db, nivel, aa, alt)
            if sem_data:
                # Data no XML está como <w:t>20/02/2024</w:t>
                tbl = tbl.replace('>20/02/2024<', '>??/??/???? - DATA PENDENTE<')
            risk += tbl + spacer()
        elif ag[0] == 'quant':
            _, nome_ag, grupo, lt, na, med, nr, acima = ag
            tbl = adapt_quant(nome_ag, grupo, lt, na, med, nr, acima)
            if sem_data:
                tbl = tbl.replace('>20/02/2024<', '>??/??/????<')
            risk += tbl + spacer()
        elif ag[0] == 'ergon':
            _, nome_ag, nr = ag
            risk += adapt_ergon(nome_ag, nr) + spacer()
        elif ag[0] == 'acid':
            _, nome_ag, nr = ag
            risk += adapt_acid(nome_ag, nr) + spacer()

    return sc + titulo + risk

def gerar_docx_bytes(nome, cnpj, rua, numero, complemento, cep, bairro, cidade, uf, cargos,
                     cnae='', descricao_cnae='', grau_risco=''):
    """Gera o PGR. `cnae`/`descricao_cnae`/`grau_risco` vêm do CADASTRO da empresa
    (`db.dados_cadastro_empresa`) — antes de 28/07/2026 não eram substituídos e o
    documento saía com os da empresa de referência do template para todo cliente.
    Campo sem valor no cadastro vira '???', a convenção que o PGR já usa para
    medição sem data confirmada (GHE_SEM_DATA) — some no meio da capa é pior."""
    data_atual = mes_ano()
    partes = [p for p in [rua, (f'Nº {numero}' if numero else ''), complemento] if p.strip()]
    endereco = ' '.join(partes) or 'A definir'
    part1 = load_tpl('part1')
    part3 = load_tpl('part3')

    # Faltando no cadastro → '???' (não o valor do template, que é de outra
    # empresa, nem vazio, que passa batido no meio da tabela da capa).
    _pgr_cnae = xs(str(cnae or '').strip()) or '???'
    _pgr_desc = xs(str(descricao_cnae or '').strip()) or '???'
    _pgr_grau = str(grau_risco or '').strip()
    # O template escreve o grau com 2 dígitos ('03'); mantém a tipografia quando
    # o cadastro tem só '3', para o documento não ficar com dois estilos.
    if _pgr_grau.isdigit() and len(_pgr_grau) == 1:
        _pgr_grau = _pgr_grau.zfill(2)
    _pgr_grau = xs(_pgr_grau) or '???'

    def subst(t):
        t = t.replace('63.370.132 MARCIO DA SILVA', xs(nome))
        t = t.replace('63.370.132/0001-25', xs(cnpj))
        t = t.replace('Al Das Sibipurunas Nº 1137', xs(endereco))
        t = t.replace('33.830-360', xs(cep) or 'A definir')
        t = t.replace('>Ribeirão das Neves<', f'>{xs(cidade)}<')
        t = t.replace('Vale das Acácias', xs(bairro) or 'A definir')
        t = t.replace('RIBEIRÃO DAS NEVES - MG</w:t>', f'{xs(cidade).upper()} - {xs(uf).upper()}</w:t>')
        t = t.replace('Setor: Ribeirão das Neves - MG</w:t>', f'Setor: {xs(cidade)} - {xs(uf)}</w:t>')
        t = t.replace(f'>{ORIG["data"]}<', f'>{data_atual}<')
        # ── CNAE / descrição / grau de risco, do cadastro da empresa ──────
        # Ordem importa: a string mais longa primeiro. Em part1 (capa) o valor do
        # grau está em parágrafo separado do rótulo e '>03<' é ÚNICO no arquivo;
        # em part3 ele vem no MESMO run, dentro do nome do treinamento
        # ("Treinamento CIPA - NR-05 - Grau de Risco 03"), que também tem de
        # acompanhar a empresa — o dimensionamento da CIPA depende do grau.
        t = t.replace('43.99-1-03', _pgr_cnae)
        t = t.replace('Obras de alvenaria', _pgr_desc)
        t = t.replace('Grau de Risco 03', f'Grau de Risco {_pgr_grau}')
        t = t.replace('>03<', f'>{_pgr_grau}<')
        return t
    p1 = subst(part1)
    p3 = subst(part3)
    # Tabela "Setor / Cargo / Nº de Funcionários": a linha do template (cargo
    # "Pedreiro") é o MOLDE — clona uma linha por cargo selecionado
    _ci = p1.find('>Pedreiro<')
    if _ci != -1:
        _trs = max(p1.rfind('<w:tr ', 0, _ci), p1.rfind('<w:tr>', 0, _ci))
        _tre = p1.find('</w:tr>', _ci) + len('</w:tr>')
        _molde = p1[_trs:_tre]
        p1 = (p1[:_trs] +
              ''.join(_molde.replace('>Pedreiro<', f'>{xs(c)}<') for c in cargos) +
              p1[_tre:])
    # Índice dinâmico
    tpl_idx = load_tpl('indice_cargo')
    idx_novo = ''
    for i, cargo in enumerate(cargos):
        para = tpl_idx.replace('>Cargo: Pedreiro<', f'>Cargo: {xs(cargo)}<')
        new_id = '%08X' % (0x762D6EE3 + i + 1)
        para = para.replace('762D6EE3', new_id)
        idx_novo += para + '\n'
    if tpl_idx in p1:
        p1 = p1.replace(tpl_idx, idx_novo, 1)
    new_cargos = ''.join(build_cargo_section(c, cidade, uf) for c in cargos)
    new_xml = p1 + '\n' + new_cargos + '\n    ' + p3
    # Reatribuir IDs duplicados — Word moderno rejeita w:id e w14:paraId repetidos
    _id_counter = [1]
    def _new_wid(m):
        _id_counter[0] += 1
        return f'w:id="{_id_counter[0]}"'
    new_xml = re.sub(r'w:id="\d+"', _new_wid, new_xml)
    _pid_seen = set()
    def _new_paraid(m):
        import random
        v = m.group(1)
        while v in _pid_seen:
            v = '%08X' % random.randint(1, 0x7FFFFFFE)
        _pid_seen.add(v)
        return f'w14:paraId="{v}"'
    new_xml = re.sub(r'w14:paraId="([^"]+)"', _new_paraid, new_xml)
    ET.fromstring(new_xml)  # valida
    # Empacotar em memória
    work_dir = tempfile.mkdtemp()
    try:
        shutil.copytree(MODEL_DIR, os.path.join(work_dir, 'doc'))
        with open(os.path.join(work_dir, 'doc', 'word', 'document.xml'), 'w', encoding='utf-8') as f:
            f.write(new_xml)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, files in os.walk(os.path.join(work_dir, 'doc')):
                for file in files:
                    abs_path = os.path.join(root, file)
                    rel_path = os.path.relpath(abs_path, os.path.join(work_dir, 'doc'))
                    zf.write(abs_path, rel_path)
        buf.seek(0)
        return buf.getvalue()
    finally:
        shutil.rmtree(work_dir, ignore_errors=True)

# ── Extração de PDF ───────────────────────────────────────────────
_LABEL_PAT = re.compile(
    r'^(CNPJ|INSCRI|ENDERE|N[UÚ]MERO|COMPLEMENTO|CEP|BAIRRO|MUNIC[IÍ]P|UF\b|DATA|FISCAL|'
    r'CARGOS|CONTATO|ESTA FICHA|AP[OÓ]S|NOVA EMPREI|Kelly|suporte|engenharia)',
    re.I
)
_ADDR_PAT = re.compile(r'^(RUA\b|R\s|AV\b|AVENIDA\b|AL\s|ALAMEDA\b|PC[AÇ]|TRAV|VIA\s)', re.I)
_DATE_PAT = re.compile(r'^\d{2}[/\-]\d{2}[/\-]\d{4}$|^\d{2}\s+\w+[,\s]+\d{4}')
_NUM_PAT  = re.compile(r'^\d{2}[\.\d/\-\s]+$')

def _is_label(s):
    return bool(_LABEL_PAT.match(s)) or (s.endswith(':') and len(s) < 40)

def _nome_valido(s):
    if not s or len(s) <= 3: return False
    if _is_label(s): return False
    if _ADDR_PAT.match(s): return False
    if _DATE_PAT.match(s): return False
    if _NUM_PAT.match(s): return False
    return True

def extrair_pdf(file_bytes):
    dados = {"nome":"","cnpj":"","rua":"","numero":"","complemento":"","cep":"","bairro":"","cidade":"","uf":"MG","cargos":[]}
    if not PDF_OK:
        return dados
    try:
        buf = io.BytesIO(file_bytes)
        texto = pdf_extract(buf)
    except:
        return dados
    full = ' '.join(l.strip() for l in texto.split('\n') if l.strip())
    lines = [l.strip() for l in texto.split('\n') if l.strip()]

    # CNPJ — primeiro match no texto
    cnpj_m = re.search(r'\d{2}[\.\s]?\d{3}[\.\s]?\d{3}[/\s]?\d{4}[-\s]?\d{2}', full)
    if cnpj_m:
        dados['cnpj'] = re.sub(r'\s', '', cnpj_m.group())

    # Estratégia 1: linha após "RAZÃO SOCIAL:" (funciona quando pdfminer extrai par label→valor)
    for i, l in enumerate(lines):
        if not re.search(r'RAZ[ÃA]O SOCIAL|NOME EMPRES', l, re.I):
            continue
        # inline: "RAZÃO SOCIAL: NOME AQUI"
        inline = re.split(r'RAZ[ÃA]O SOCIAL[\s:]*|NOME EMPRES[A-Z]*[\s:]*', l, flags=re.I)
        c = inline[-1].strip() if len(inline) > 1 else ''
        if _nome_valido(c):
            dados['nome'] = c
            break
        # próxima linha que não é label
        for j in range(i + 1, min(i + 3, len(lines))):
            c = lines[j].strip()
            # Remove data do início da linha se concatenada (ex: "08/04/2026NomeDaEmpresa")
            c = re.sub(r'^(\d{2}[/\-]\d{2}[/\-]\d{4}|\d{2}\s+\w+[,\s]+\d{4})\s*', '', c).strip()
            if _nome_valido(c):
                dados['nome'] = c
                break
        break

    # Estratégia 2: linha(s) imediatamente antes do CNPJ (funciona quando pdfminer extrai por bloco)
    if not dados['nome'] and cnpj_m:
        cnpj_raw = re.sub(r'[\s\.]', '', cnpj_m.group())
        for j, ln in enumerate(lines):
            ln_norm = re.sub(r'[\s\.]', '', ln)
            # CNPJ está nesta linha (exato ou embutido numa linha longa)
            if cnpj_raw == ln_norm or cnpj_raw in re.sub(r'[\s\.]', '', ln):
                if cnpj_raw == ln_norm:
                    # Linha própria — busca nas linhas anteriores
                    for k in range(j - 1, max(-1, j - 10), -1):
                        if _nome_valido(lines[k]):
                            dados['nome'] = lines[k]
                            break
                else:
                    # Linha longa com tudo concatenado — extrai texto entre data e CNPJ completo
                    idx_cnpj = ln.find(cnpj_m.group())
                    if idx_cnpj < 0:
                        idx_cnpj = ln.find(cnpj_m.group().split('/')[0])
                    if idx_cnpj > 0:
                        trecho = ln[:idx_cnpj].strip()
                        # Remove data do início (ex: "07 abril, 2026" ou "08/04/2026")
                        trecho = re.sub(r'^(\d{2}[/\-]\d{2}[/\-]\d{4}|\d{2}\s+\w+[,\s]+\d{4})\s*', '', trecho).strip()
                        if _nome_valido(trecho):
                            dados['nome'] = trecho
                break

    # CEP
    cep_m = re.search(r'\d{5}-?\d{3}', full)
    if cep_m:
        dados['cep'] = cep_m.group()

    # UF
    uf_m = re.search(r'\b(MG|SP|RJ|ES|GO|BA|PR|SC|RS|DF|MT|MS|AM|PA|CE|PE|MA|RN|PB|AL|SE|PI|TO|RO|AC|RR|AP)\b', full)
    if uf_m:
        dados['uf'] = uf_m.group()

    # Endereço — ignora linhas que são só o label "ENDEREÇO:" sem valor
    for i, l in enumerate(lines):
        if re.match(r'^(RUA|AV\b|AVENIDA|LOGRADOURO|R\s)', l, re.I) and len(l) > 5:
            dados['rua'] = l
            break
        if re.match(r'^ENDERE', l, re.I):
            for j in range(i + 1, min(i + 5, len(lines))):
                candidate = lines[j].strip()
                if re.match(r'^(RUA|AV\b|AVENIDA|R\s)', candidate, re.I) and len(candidate) > 5:
                    dados['rua'] = candidate
                    break

    # Cargos
    for cargo in ["Pedreiro","Servente","Pintor","Meio Oficial Pintor","Armador","Eletricista",
                  "Carpinteiro","Gesseiro","Rejuntador","Azulejista","Bombeiro Hidraulico",
                  "Encarregado","Montador","Meio Oficial Pedreiro","Meio Oficial Eletricista",
                  "Auxiliar de Limpeza","Porteiro","Vigia","Topografo","Auxiliar Administrativo","Engenheiro"]:
        if re.search(re.escape(cargo), full, re.I):
            dados['cargos'].append(cargo)
    return dados

# ── Rotas ─────────────────────────────────────────────────────────

# (rota /sw.js duplicada removida — a de cima, _root_service_worker, é a que vence)

@app.route('/login')
def login_redirect():
    from flask import redirect, url_for
    return redirect(url_for('auth.login'))

@app.route('/favicon.ico')
def favicon():
    """Ícone da aba do navegador (usa o ícone do PWA). Cobre todas as páginas."""
    return send_from_directory('static', 'icon-192.png', mimetype='image/png')


# Marcador de build — atualizar a cada push para conferir qual versão está no ar
BUILD_MARK = '2026-07-10-r47-aba-usuarios-admin'


@app.route('/healthz')
def healthz():
    return jsonify({'ok': True, 'build': BUILD_MARK})


@app.route('/')
def index():
    if not current_user.is_authenticated:
        return render_template('landing.html')
    usuarios_pendentes = 0
    if getattr(current_user, 'role', '') == 'admin':
        try:
            with get_db() as conn:
                row = conn.execute('SELECT COUNT(*) AS c FROM usuarios WHERE ativo=0').fetchone()
            usuarios_pendentes = row['c'] if row else 0
        except Exception:
            pass
    return render_template('index.html', cargos_sugestoes=CARGOS_SUGESTOES,
                           usuarios_pendentes=usuarios_pendentes)


@app.route('/extrair', methods=['POST'])
@login_required
def extrair():
    if 'pdf' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    f = request.files['pdf']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'erro': 'Envie um arquivo PDF'}), 400
    dados = extrair_pdf(f.read())
    return jsonify(dados)

@app.route('/gerar', methods=['POST'])
@login_required
def gerar():
    data = request.json
    nome   = xclean(data.get('nome','').strip())
    cnpj   = xclean(data.get('cnpj','').strip())
    rua    = xclean(data.get('rua','').strip())
    numero = xclean(data.get('numero','').strip())
    compl  = xclean(data.get('complemento','').strip())
    cep    = xclean(data.get('cep','').strip())
    bairro = xclean(data.get('bairro','').strip())
    cidade = xclean(data.get('cidade','').strip()) or 'Belo Horizonte'
    uf     = xclean(data.get('uf','MG').strip().upper())
    cargos = [xclean(c.strip()) for c in data.get('cargos',[]) if c.strip()]

    if not nome: return jsonify({'erro': 'Informe a Razão Social'}), 400
    if not cargos: return jsonify({'erro': 'Adicione pelo menos um cargo'}), 400

    # ── Verificar problemas ANTES de gerar ──────────────────────────
    alertas = []

    # 1. Cargos não reconhecidos no GHE
    desconhecidos = [c for c in cargos if get_ghe(c) is None]
    if desconhecidos:
        return jsonify({
            'aviso': True,
            'erro': f'Cargo(s) não reconhecido(s) no GHE MRV: {", ".join(desconhecidos)}. '
                    f'Verifique a grafia ou solicite o cadastro deste cargo.'
        }), 422

    # 2. Cargos com GHE sem data de medição confirmada
    sem_data = [c for c in cargos if get_ghe(c) in GHE_SEM_DATA]
    if sem_data:
        ghes_afetados = list(dict.fromkeys(get_ghe(c) for c in sem_data))
        alertas.append(f'Data de medição não confirmada para: {", ".join(sem_data)} '
                       f'(GHE: {", ".join(ghes_afetados)}). '
                       f'O campo de data aparecerá como ??? no PGR gerado.')

    # 3. Campos obrigatórios ausentes
    faltando = []
    if not cnpj: faltando.append('CNPJ')
    if not cep:  faltando.append('CEP')
    if not cidade or cidade == 'Belo Horizonte' and not data.get('cidade','').strip():
        faltando.append('Cidade')
    if faltando:
        alertas.append(f'Campos não preenchidos: {", ".join(faltando)}. '
                       f'O PGR será gerado com esses campos em branco.')

    # 4. CNAE / descrição / grau de risco vêm do CADASTRO (decisão 28/07/2026).
    # O form do PGR não pede esses campos; antes disto o documento saía com os da
    # empresa de referência do template para todo cliente.
    from controle.db import dados_cadastro_empresa
    _cad = dados_cadastro_empresa(nome, cnpj)
    _sem = [lbl for lbl, val in (('CNAE', _cad['cnae']),
                                 ('Descrição do CNAE', _cad['descricao_cnae']),
                                 ('Grau de Risco', _cad['grau_risco'])) if not val]
    if _sem:
        _onde = ('a empresa não está cadastrada' if _cad['id'] is None
                 else 'o cadastro da empresa está sem esses dados')
        alertas.append(
            f'{", ".join(_sem)}: {_onde}. Esses campos sairão como "???" no PGR — '
            f'preencha o cadastro da empresa e gere de novo. '
            f'(CNAE e descrição se preenchem sozinhos ao gerar um laudo de calor, '
            f'químico ou ruído; o grau de risco é preenchido à mão.)')

    try:
        docx_bytes = gerar_docx_bytes(nome,cnpj,rua,numero,compl,cep,bairro,cidade,uf,cargos,
                                     cnae=_cad['cnae'], descricao_cnae=_cad['descricao_cnae'],
                                     grau_risco=_cad['grau_risco'])
        nome_safe = re.sub(r'[/\\:*?"<>|]','_',nome)
        filename = f"PGR - {nome_safe} - {mes_ano().replace(' / ','_')}.docx"

        usuario = current_user.nome if current_user.is_authenticated else 'anônimo'
        registrar_evento('pgr_gerado', f'PGR: {nome} ({len(cargos)} cargos)',
                         usuario=usuario, ip=request.remote_addr)

        # Se há alertas, retornar JSON com aviso + base64 do arquivo
        if alertas:
            import base64
            return jsonify({
                'aviso_gerado': True,
                'alertas': alertas,
                'filename': filename,
                'docx_b64': base64.b64encode(docx_bytes).decode()
            })

        return send_file(
            io.BytesIO(docx_bytes),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

@app.route('/ghe/<cargo>')
def ghe_info(cargo):
    ghe = get_ghe(cargo)
    if ghe is None:
        return jsonify({'ghe': 'NÃO_RECONHECIDO', 'agentes': 0, 'aviso': True})
    agentes = GHE_AGENTES.get(ghe, [])
    return jsonify({'ghe': ghe, 'agentes': len(agentes)})

# ── Laudo de Calor ────────────────────────────────────────────────────

# Quadro 1 do Anexo 3 da NR-15 (redação da Portaria SEPRT 1.359/2019):
# taxa de metabolismo média M (W) → limite de exposição IBUTG_MAX (ºC).
# Tabela oficial completa — vai até 606 W (antes truncava em 346 e todo
# trabalho pesado ganhava limite 27,5, mais permissivo que a norma).
_NR15_QUADRO1 = [
    (100,33.7),(102,33.6),(104,33.5),(106,33.4),(108,33.3),(110,33.2),
    (112,33.1),(115,33.0),(117,32.9),(119,32.8),(122,32.7),(124,32.6),
    (127,32.5),(129,32.4),(132,32.3),(135,32.2),(137,32.1),(140,32.0),
    (143,31.9),(146,31.8),(149,31.7),(152,31.6),(155,31.5),(158,31.4),
    (161,31.3),(165,31.2),(168,31.1),(171,31.0),(175,30.9),(178,30.8),
    (182,30.7),(186,30.6),(189,30.5),(193,30.4),(197,30.3),(201,30.2),
    (205,30.1),(209,30.0),(214,29.9),(218,29.8),(222,29.7),(227,29.6),
    (231,29.5),(236,29.4),(241,29.3),(246,29.2),(251,29.1),(256,29.0),
    (261,28.9),(266,28.8),(272,28.7),(277,28.6),(283,28.5),(289,28.4),
    (294,28.3),(300,28.2),(306,28.1),(313,28.0),(319,27.9),(325,27.8),
    (332,27.7),(339,27.6),(346,27.5),(353,27.4),(360,27.3),(367,27.2),
    (374,27.1),(382,27.0),(390,26.9),(398,26.8),(406,26.7),(414,26.6),
    (422,26.5),(431,26.4),(440,26.3),(448,26.2),(458,26.1),(467,26.0),
    (476,25.9),(486,25.8),(496,25.7),(506,25.6),(516,25.5),(526,25.4),
    (537,25.3),(548,25.2),(559,25.1),(570,25.0),(582,24.9),(594,24.8),
    (606,24.7)]

def get_limite_nr15(m_medio):
    T = _NR15_QUADRO1
    if m_medio <= T[0][0]:  return T[0][1]
    if m_medio >= T[-1][0]: return T[-1][1]
    for i in range(len(T)-1):
        m1,i1 = T[i]; m2,i2 = T[i+1]
        if m1 <= m_medio <= m2:
            return round(i1 + (i2-i1)*(m_medio-m1)/(m2-m1), 1)
    return 30.0

def _ibutg_ponto(p):
    """IBUTG do ponto + fórmula exibida no laudo.
    TBS preenchido (>0) = céu aberto com carga solar direta →
    IBUTG = 0,7·tbn + 0,1·tbs + 0,2·tg (NHO 06 / Anexo 3 da NR-15).
    Sem TBS = ambiente interno/sem carga solar → 0,7·tbn + 0,3·tg.
    Mesma convenção da planilha de campo (TBS digitado ativa o externo)."""
    tbn = float(p.get('tbn') or 0)
    tg  = float(p.get('tg') or 0)
    try:
        tbs = float(p.get('tbs') or 0)
    except (TypeError, ValueError):
        tbs = 0.0
    if tbs > 0:
        ibutg   = round(0.7*tbn + 0.1*tbs + 0.2*tg, 1)
        formula = f'IBUTG = (0,7 x {_fx(tbn)}) + (0,1 x {_fx(tbs)}) + (0,2 x {_fx(tg)})'
    else:
        ibutg   = round(0.7*tbn + 0.3*tg, 1)
        formula = f'IBUTG = (0,7 x {_fx(tbn)}) + (0,3 x {_fx(tg)})'
    return ibutg, formula

def _fx(v):
    try: return str(round(float(v),2)).replace('.',',')
    except: return str(v)

def _xe(s):
    import html as _h
    return _h.escape(str(s))

def _rr(bxml, label, val, nth=1):
    """Replace o valor da célula SEGUINTE à célula do label (nth ocorrência).
    Antes escrevia sempre na ÚLTIMA célula da linha — em linhas com dois pares
    label/valor (ex.: 'Data da coleta' + 'Data da análise') o primeiro valor
    caía na célula errada e era sobrescrito pelo segundo."""
    pos = -1
    for _ in range(nth):
        nxt = bxml.find(f'>{label}</', pos + 1)
        if nxt == -1: return bxml
        pos = nxt
    tr_s = bxml.rfind('<w:tr ', 0, pos)
    tr_e = bxml.find('</w:tr>', pos) + 7
    row  = bxml[tr_s:tr_e]
    tcs  = [m.start() for m in re.finditer('<w:tc>', row)]
    if len(tcs) < 2: return bxml
    lbl_off = pos - tr_s
    nxt_tcs = [t for t in tcs if t > lbl_off]
    if nxt_tcs:
        lts = nxt_tcs[0]
        lte = row.find('</w:tc>', lts) + 7
    else:
        lts = tcs[-1]; lte = row.rfind('</w:tc>') + 7
    vc  = row[lts:lte]
    def g(pat, default=''):
        m = re.search(pat, vc, re.DOTALL); return m.group(0) if m else default
    tcp = g(r'<w:tcPr>.*?</w:tcPr>')
    ppr = g(r'<w:pPr>.*?</w:pPr>')
    rpr = g(r'<w:rPr>.*?</w:rPr>')
    pm  = re.search(r'<w:p ([^>]*?)>', vc); pa = pm.group(1) if pm else ''
    nc  = f'<w:tc>{tcp}<w:p {pa}>{ppr}<w:r>{rpr}<w:t xml:space="preserve">{_xe(val)}</w:t></w:r></w:p></w:tc>'
    return bxml[:tr_s] + row[:lts] + nc + row[lte:] + bxml[tr_e:]

def _ri(bxml, label, val):
    """Replace value runs after inline label in same paragraph."""
    pos = bxml.find(f'>{label}</')
    if pos == -1: return bxml
    re_e = bxml.find('</w:r>', pos) + 6
    pe   = bxml.find('</w:p>', re_e)
    m    = re.search(r'<w:rPr>.*?</w:rPr>', bxml[re_e:pe], re.DOTALL)
    rpr  = m.group(0) if m else ''
    nr   = f'<w:r>{rpr}<w:t xml:space="preserve">{_xe(val)}</w:t></w:r>'
    return bxml[:re_e] + nr + bxml[pe:]

def _rp(bxml, label, val):
    """Replace o conteúdo do PARÁGRAFO seguinte ao parágrafo do label.
    Usado quando o template tem o label e o valor em parágrafos separados
    da mesma célula (ex.: 'Vazão Média (L/min): ' seguido de '0,19550')."""
    pos = bxml.find(f'>{label}</')
    if pos == -1: return bxml
    p_end = bxml.find('</w:p>', pos) + 6
    np_m  = re.search(r'<w:p[ >]', bxml[p_end:])
    if not np_m: return bxml
    np_s  = p_end + np_m.start()
    np_e  = bxml.find('</w:p>', np_s) + 6
    para  = bxml[np_s:np_e]
    r_m   = re.search(r'<w:r[ >]', para)
    if not r_m: return bxml
    r_s   = r_m.start()
    m     = re.search(r'<w:rPr>.*?</w:rPr>', para[r_s:], re.DOTALL)
    rpr   = m.group(0) if m else ''
    return (bxml[:np_s] + para[:r_s] +
            f'<w:r>{rpr}<w:t xml:space="preserve">{_xe(val)}</w:t></w:r></w:p>' +
            bxml[np_e:])

# ══════════════════════════════════════════════════════════════════
# Quimico — constants & helpers
# ══════════════════════════════════════════════════════════════════
import json as _json

CERTS_DIR = os.path.join(BASE_DIR, 'static', 'certs')
_GUIA_PATH = os.path.join(BASE_DIR, 'guia_metodos.json')

_PUMP_SN = {
    'bdx':     ['20230702024', '20230702029', '20230702030', '20141201119'],
    'airlite': ['A060502', 'A061553', 'A061585', 'A062462', 'A63555'],
    'turam':   ['2420120549', '2420120550', '2420120551'],
    'inlite':  ['25040902602B', '25040903102B', '25040907102B'],
}
_PUMP_CERT_PAGES = {
    'bdx':     {'20230702024': 2, '20230702029': 2, '20230702030': 2, '20141201119': 2,
                '38356': 2, '38357': 2, '38358': 2, '38359': 2},
    'airlite': {'A060502': 2, 'A061553': 4, 'A061585': 2, 'A062462': 2, 'A63555': 2},
    'turam':   {'2420120549': 1, '2420120550': 1, '2420120551': 1},
    'inlite':  {'25040902602B': 2, '25040903102B': 2, '25040907102B': 2},
}
_PUMP_NAMES = {
    'bdx': 'BDX II – GILLIAN', 'airlite': 'AIRLITE – SKC',
    'turam': 'FORMIS – TURAM',  'inlite':  'INLITE VENTUSPRO',
}
_CALIB_CERT_PAGES = {'defender510m': 2, 'tsi4143f': 2}
_CALIB_NAMES = {'defender510m': 'DEFENDER 510M', 'tsi4143f': 'TSI 4143F'}

def _guia_norm(x):
    import unicodedata as _ud
    return _ud.normalize('NFD', str(x)).encode('ascii', 'ignore').decode('ascii').upper().strip()

def _guia_entry(agente_str, guia):
    """Acha a entrada do guia_metodos p/ um agente livre (nome, nome (sinônimo)
    ou nome (CAS)). Casa por CAS, por nome exato (string toda / dentro de
    parênteses / base sem parênteses) e, por fim, por conteúdo parcial."""
    if not agente_str or not guia:
        return None
    s = str(agente_str).strip()
    for cas in re.findall(r'\b(\d{2,7}-\d{2}-\d)\b', s):
        if cas in guia and guia[cas]:
            return guia[cas][0]
    cands = {_guia_norm(s)}
    for inside in re.findall(r'\(([^)]*)\)', s):
        cands.add(_guia_norm(inside))
    cands.add(_guia_norm(re.sub(r'\s*\([^)]*\)\s*', ' ', s)))
    cands.discard('')
    for entries in guia.values():
        for e in entries:
            if _guia_norm(e.get('nome', '')) in cands:
                return e
    ns = _guia_norm(s)
    for entries in guia.values():
        for e in entries:
            n = _guia_norm(e.get('nome', ''))
            if n and (n in ns or ns in n):
                return e
    return None

# Calibradores de nivel sonoro (ruido) — Chrompack SmartCal, frota do grupo (Ocupacional/Assiste).
# Quando ha 2 certificados p/ mesma serie, vale a calibracao mais recente (ja consolidado abaixo).
# Calibracao acustica IEC 60942 e anual → validade = data_calib + 1 ano.
_CALIB_RUIDO_MARCA = 'CHROMPACK'
_CALIB_RUIDO_MODELO = 'SMARTCAL'
_CALIB_RUIDO = {
    '1562': {'serie': 'CAL0000001562', 'cert': '142.574',    'data_calib': '2023-02-14', 'marca': 'CHROMPACK'},
    '1575': {'serie': 'CAL0000001575', 'cert': '182.920', 'data_calib': '2026-05-21', 'marca': 'CHROMPACK'},
    '2150': {'serie': 'CAL0000002150', 'cert': '172.833',    'data_calib': '2025-08-18', 'marca': 'CHROMPACK'},
    '1614': {'serie': 'CAL0000001614', 'cert': '181.238',    'data_calib': '2026-04-08', 'marca': 'CHROMPACK'},
    '0284': {'serie': 'CAL0000000284', 'cert': '181.239',    'data_calib': '2026-04-08', 'marca': 'CHROMPACK'},
    # Calibrador Inlite (CalPro) — atende SÓ dosimetros Inlite. Serie conforme certificado (25035711).
    '25035711': {'serie': '25035711', 'cert': '42.179-2025', 'data_calib': '2025-08-28', 'marca': 'INLITE', 'modelo': 'CalPro'},
}

# Dosimetros de ruido — frota do grupo (Chrompack SmartdB + Inlite DoseMax V2).
# Chrompack dosim. usam calibrador Chrompack; Inlite usam o calibrador Inlite.
# OBS: serie 1153 nao incluida — o PDF "1153" contem cert da serie 1154 (duplicado). Verificar com fornecedor.
_DOSIM_RUIDO = {
    'chrompack': {
        # 2025 (cal 19/08/2025) — modelo SmartdB LITE
        '1149': {'serie': '0000001149', 'cert': '172.887', 'data_calib': '2025-08-19'},
        '1150': {'serie': '0000001150', 'cert': '172.889', 'data_calib': '2025-08-19'},
        '1151': {'serie': '0000001151', 'cert': '172.890', 'data_calib': '2025-08-19'},
        '1152': {'serie': '0000001152', 'cert': '172.888', 'data_calib': '2025-08-19'},
        '1153': {'serie': '0000001153', 'cert': '172.906', 'data_calib': '2025-08-19'},
        '1154': {'serie': '0000001154', 'cert': '172.886', 'data_calib': '2025-08-19'},
        '5309': {'serie': '0000005309', 'cert': '183.247', 'data_calib': '2026-06-01'},
        # 2026 (cal 09/04/2026) — modelo SmartdB
        '1085': {'serie': '0000001085', 'cert': '181.280', 'data_calib': '2026-04-09'},
        '1086': {'serie': '0000001086', 'cert': '181.272', 'data_calib': '2026-04-09'},
        '1088': {'serie': '0000001088', 'cert': '181.273', 'data_calib': '2026-04-09'},
        '5062': {'serie': '0000005062', 'cert': '181.281', 'data_calib': '2026-04-09'},
        '5063': {'serie': '0000005063', 'cert': '181.277', 'data_calib': '2026-04-09'},
        '5064': {'serie': '0000005064', 'cert': '181.275', 'data_calib': '2026-04-09'},
        '5065': {'serie': '0000005065', 'cert': '181.279', 'data_calib': '2026-04-09'},
        '5066': {'serie': '0000005066', 'cert': '181.271', 'data_calib': '2026-04-09'},
        '5308': {'serie': '0000005308', 'cert': '181.274', 'data_calib': '2026-04-09'},
        '5311': {'serie': '0000005311', 'cert': '181.278', 'data_calib': '2026-04-09'},
        # Lote recalibrado em 01/06/2026 (certificados conferidos nos PDFs).
        '1084': {'serie': '0000001084', 'cert': '183.249', 'data_calib': '2026-06-01'},
        '5307': {'serie': '0000005307', 'cert': '183.250', 'data_calib': '2026-06-01'},
        '5306': {'serie': '0000005306', 'cert': '183.248', 'data_calib': '2026-06-01'},
        '5310': {'serie': '0000005310', 'cert': '183.254', 'data_calib': '2026-06-01'},
        '5061': {'serie': '0000005061', 'cert': '183.252', 'data_calib': '2026-06-01'},
        '0320': {'serie': '0000000320', 'cert': '183.251', 'data_calib': '2026-06-01'},  # SmartdB LITE
        '1087': {'serie': '0000001087', 'cert': '183.253', 'data_calib': '2026-06-01'},
    },
    'inlite': {
        # DoseMax V2 (cal 28/08/2025)
        '2507050490AA': {'serie': '2507050490AA', 'cert': '42.191-2025', 'data_calib': '2025-08-28'},
        '2507050520AA': {'serie': '2507050520AA', 'cert': '42.189-2025', 'data_calib': '2025-08-28'},
        '2507050710AA': {'serie': '2507050710AA', 'cert': '42.180-2025', 'data_calib': '2025-08-28'},
        '2507050750AA': {'serie': '2507050750AA', 'cert': '42.190-2025', 'data_calib': '2025-08-28'},
        '2507050810AA': {'serie': '2507050810AA', 'cert': '42.192-2025', 'data_calib': '2025-08-28'},
    },
}
_DOSIM_RUIDO_MODELO = {'chrompack': 'SmartdB', 'inlite': 'DoseMax V2'}


def _q_img_para(rid, iid, cx=6858000, cy=9693700):
    """Inline image paragraph — full-page sized by default."""
    ah = f'{(iid * 0x1F3A + 0x3456) & 0xFFFFFFFE:08X}'
    return (
        f'<w:p w14:paraId="{ah}" w14:textId="77777777">'
        '<w:pPr><w:jc w:val="center"/><w:spacing w:before="0" w:after="0"/></w:pPr>'
        f'<w:r><w:drawing><wp:inline distT="0" distB="0" distL="0" distR="0">'
        f'<wp:extent cx="{cx}" cy="{cy}"/><wp:effectExtent l="0" t="0" r="0" b="0"/>'
        f'<wp:docPr id="{iid}" name="qimg{iid}"/>'
        '<wp:cNvGraphicFramePr><a:graphicFrameLocks '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/>'
        '</wp:cNvGraphicFramePr>'
        '<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        '<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        '<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:nvPicPr><pic:cNvPr id="{iid}" name="qimg{iid}.jpg"/><pic:cNvPicPr/></pic:nvPicPr>'
        f'<pic:blipFill><a:blip r:embed="{rid}"/>'
        '<a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
        f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
        '</pic:pic></a:graphicData></a:graphic>'
        '</wp:inline></w:drawing></w:r></w:p>'
    )


def _q_add_file(path, extra_rels, extra_media, ctr):
    ctr[0] += 1
    rid = f'rId{ctr[0]}'; iid = ctr[0]
    fname = f'media/qf_{ctr[0]}.jpg'
    extra_rels.append(
        f'<Relationship Id="{rid}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        f'Target="{fname}"/>')
    with open(path, 'rb') as f:
        extra_media[f'word/{fname}'] = f.read()
    return rid, iid


def _q_add_b64(b64str, extra_rels, extra_media, ctr):
    ctr[0] += 1
    rid = f'rId{ctr[0]}'; iid = ctr[0]
    if ',' in b64str:
        hdr, data = b64str.split(',', 1)
        ext = 'jpeg' if ('jpeg' in hdr or 'jpg' in hdr) else 'png'
    else:
        data, ext = b64str, 'jpeg'
    fname = f'media/qu_{ctr[0]}.{ext}'
    extra_rels.append(
        f'<Relationship Id="{rid}" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
        f'Target="{fname}"/>')
    extra_media[f'word/{fname}'] = base64.b64decode(data)
    return rid, iid


def _img_wh(data):
    """(largura, altura) em px de um PNG/JPEG; (0, 0) se não reconhecer."""
    import struct
    try:
        if data[:8] == b'\x89PNG\r\n\x1a\n':
            w, h = struct.unpack('>II', data[16:24]); return w, h
        if data[:2] == b'\xff\xd8':            # JPEG: acha o SOF
            i = 2
            while i < len(data) - 9:
                if data[i] != 0xFF:
                    i += 1; continue
                if data[i + 1] in (0xC0, 0xC1, 0xC2, 0xC3):
                    h, w = struct.unpack('>HH', data[i + 5:i + 9]); return w, h
                i += 2 + struct.unpack('>H', data[i + 2:i + 4])[0]
    except Exception:
        pass
    return 0, 0


def _q_fit_extent(data, txt_w=5400040, txt_h=7811770):
    """(cx, cy) em EMU p/ a imagem preencher a largura útil A4 preservando a
    proporção real (cap na altura útil). Evita imagem estourando a margem."""
    w, h = _img_wh(data)
    if w and h:
        cx = txt_w; cy = int(cx * h / w)
        if cy > txt_h:
            cy = txt_h; cx = int(cy * w / h)
        return cx, cy
    return txt_w, int(txt_w * 1.414)   # fallback A4 retrato


def _q_sec_head(xml, text):
    """Return (tbl_start, tbl_end) of the table that contains this section heading
    (second occurrence — first is TOC)."""
    p1 = xml.find(text)
    pos = xml.find(text, p1 + 1) if p1 != -1 else p1
    if pos == -1:
        return -1, -1
    para = max(xml.rfind('<w:p ', 0, pos), xml.rfind('<w:p>', 0, pos))
    ts   = xml.rfind('<w:tbl>', 0, para)
    te   = xml.find('</w:tbl>', pos) + 8
    return ts, te


def _qf_num(v):
    """Primeiro número de um campo de limite/concentração ('20 ppm' → 20.0)."""
    try:
        return float(str(v).split()[0].replace(',', '.'))
    except Exception:
        return None


def _fmt_num(x):
    """4 casas, sem zeros à direita, decimal com vírgula (padrão do laudo)."""
    return (f'{x:.4f}'.rstrip('0').rstrip('.')).replace('.', ',')


def _jornada_horas(txt):
    """Lê a carga horária do campo livre de jornada.

    Devolve `(horas, base)` com base em `'dia'` ou `'semana'`, ou `(None, None)`
    quando não dá para ler com segurança. O campo nasce livre na tela e chega em
    formatos bem diferentes: `09:00H` (o padrão), `08:00 / 08 horas`, `12x36`,
    `44h semanais`, `8h/dia`.

    Na dúvida devolve None de propósito. Ler errado é pior que não ler: o número
    vira Limite de Tolerância em documento assinado, e quem chama volta para a
    jornada padrão de 44 h, que é o que o gerador sempre fez.
    """
    t = str(txt or '').strip().lower()
    if not t:
        return (None, None)
    t = t.replace(',', '.')

    # Escala de turno: 12x36, 24x48. O 1º número são as horas do dia — mas só
    # quando o 2º é descanso em HORAS. 6x1 e 5x2 são escalas de DIAS e não
    # dizem nada sobre a carga horária.
    m = re.search(r'\b(\d{1,2})\s*[x×]\s*(\d{1,3})\b', t)
    if m:
        h, desc = float(m.group(1)), float(m.group(2))
        if desc < 12 or h <= 0 or h > 24:
            return (None, None)
        if h <= 16:
            return (h, 'dia')
        # Plantão longo (24x48): o modelo diário degenera perto de 24 h — em
        # 24 h o limite corrigido iria a zero. A escala dá o ciclo inteiro, então
        # a carga semanal sai da regra de três: 24 h a cada 72 h = 56 h/semana.
        return (h * 168.0 / (h + desc), 'semana')

    # Faixa de horário ("08:00 às 17:00") não é duração: não dá para descontar
    # intervalo sem inventar. Fica sem leitura.
    if len(re.findall(r'\b\d{1,2}\s*:\s*\d{2}', t)) > 1:
        return (None, None)

    semanal = bool(re.search(r'seman|/\s*sem\b|\bsem\b|por semana', t))
    diaria = bool(re.search(r'di[áa]ri|/\s*dia\b|\bdia\b|por dia|\bdiurn', t))

    # HH:MM sozinho é duração da jornada do dia (o padrão da tela é "09:00H").
    m = re.search(r'\b(\d{1,2})\s*:\s*(\d{2})', t)
    if m:
        h = float(m.group(1)) + float(m.group(2)) / 60.0
        if semanal and 0 < h < 168:
            return (h, 'semana')
        return ((h, 'dia') if 0 < h < 24 else (None, None))

    m = re.search(r'(\d{1,3}(?:\.\d+)?)\s*(?:h\b|hs\b|hr|hora)', t) or \
        re.search(r'^(\d{1,3}(?:\.\d+)?)$', t)
    if not m:
        return (None, None)
    h = float(m.group(1))
    if semanal and not diaria:
        return ((h, 'semana') if 0 < h < 168 else (None, None))
    if diaria and not semanal:
        return ((h, 'dia') if 0 < h < 24 else (None, None))
    # Sem palavra que diga a base, o tamanho do número decide.
    if 0 < h <= 16:
        return (h, 'dia')
    if 20 <= h < 168:
        return (h, 'semana')
    return (None, None)


# Jornada usada quando o campo não é legível: 44 h semanais, que dá 0,88 — o
# número fixo que o gerador aplicava em toda avaliação antes de 27/08/2026.
BS_FATOR_PADRAO = 0.88


def _fator_brief_scala(jornada_txt):
    """Fator de redução de Brief & Scala para a jornada REAL da avaliação.

    Pedido do Bernardo em 26/08/2026: "incluir a carga horária semanal ou diária
    e realizar o cálculo". Antes disso o fator era 0,88 chumbado no código e o
    laudo declarava "jornada de 44 horas semanais" para qualquer empresa.

        semanal:  FR = (40/Hsr) × ((168 − Hsr)/128)
        diário:   FR = (8/Hd)   × ((24  − Hd)/16)

    Devolve `(fator, rotulo, leu)`. O `rotulo` vai impresso no laudo porque o
    documento é assinado e tem de declarar de qual jornada saiu o limite.

    Três regras que são de uso, não da fórmula:
      • **teto em 1,0** — o modelo corrige jornada ESTENDIDA. Numa jornada curta
        a conta passa de 1 e afrouxaria o limite acima do próprio TLV da ACGIH.
        Acima de 1 o fator vira 1 e o limite fica o da ACGIH, sem correção.
      • **piso de 0,88 quando só se conhece a jornada DIÁRIA.** Saber que o
        turno é de 8 h não diz quantos dias tem a semana, e a base da casa é a
        semana brasileira de 44 h. Sem esse piso, `08:00` (o valor mais comum do
        campo) daria FR = 1 e o laudo sairia MENOS restritivo do que era antes
        desta mudança — afrouxar limite em documento assinado, calado. Jornada
        semanal declarada não tem piso: aí a informação é sobre a semana.
      • **fallback 0,88** — jornada ilegível repete o comportamento antigo em
        vez de arriscar um número inventado.
    """
    h, base = _jornada_horas(jornada_txt)
    if h is None:
        return (BS_FATOR_PADRAO, 'jornada de 44 horas semanais (padrão)', False)

    if base == 'semana':
        fr = (40.0 / h) * ((168.0 - h) / 128.0)
        rot = 'jornada de %s horas semanais' % _fmt_h(h)
        if fr >= 1.0:
            return (1.0, rot + ', sem redução aplicável', True)
        return (fr, rot, True)

    fr = (8.0 / h) * ((24.0 - h) / 16.0)
    if fr >= BS_FATOR_PADRAO:
        # Turno normal: quem manda é a semana de 44 h, como sempre foi.
        return (BS_FATOR_PADRAO,
                'jornada de 44 horas semanais (turno de %s horas diárias)' % _fmt_h(h),
                True)
    return (fr, 'jornada de %s horas diárias' % _fmt_h(h), True)


def _fmt_h(v):
    """8.0 -> '8'; 8.8 -> '8,8'. Jornada em texto de laudo, sem zero à toa."""
    return ('%.1f' % v).rstrip('0').rstrip('.').replace('.', ',')


def _classificar_quimico(ev):
    """Veredicto por limite de UMA avaliação química. FONTE ÚNICA da seção VI
    (conclusão por avaliação) e da IX (quadro resumo).

    Existe porque as duas seções faziam a própria conta e podiam discordar: a IX
    comparava com UM limite só (`ltNR15`, ou `ltTWA` como fallback) e estampava
    "18,5 (REGULAR)" numa avaliação que a VI concluía IRREGULAR pela ACGIH —
    o LT-TWA de 20 ppm corrigido pela Brief & Scala cai para 17,6 e a concentração
    passava. Quem batia o olho só no resumo lia REGULAR numa exposição acima do
    limite corrigido. Achado em 28/07/2026 e corrigido a pedido do Matheus.

    Devolve dict:
      conc     float | None   — concentração
      nd       bool           — "<" ou N.D.: abaixo do limite de detecção
      nr15     (limite, ok)   | None
      acgih    (lt_corrigido, unidade, ok) | None   — LT-TWA × Brief & Scala
      bs       {fator, rotulo, leu_jornada} | None  — de onde saiu a correção
      stel     (limite, ok)   | None                — só quando duração <= 15 min
      ok_geral bool | None    — todos os limites aplicáveis atendidos
    """
    conc_txt = str(ev.get('concentracao', '') or '').strip()
    nd = ('<' in conc_txt) or conc_txt.upper() in ('N.D.', 'ND', 'NÃO DETECTADO', '')
    cv = _qf_num(conc_txt)

    r = {'conc': cv, 'nd': nd, 'nr15': None, 'acgih': None, 'bs': None,
         'stel': None, 'ok_geral': None}
    if nd:
        # Não detectado = abaixo do LD e portanto de qualquer LT (#9 Bernardo).
        r['ok_geral'] = True
        return r

    ltn = _qf_num(ev.get('ltNR15', ''))
    if ltn is not None and cv is not None:
        r['nr15'] = (ltn, cv < ltn)

    ltw = _qf_num(ev.get('ltTWA', ''))
    if ltw is not None and cv is not None:
        _um = re.search(r'(mg/m³|mg/m3|µg/m³|μg/m³|ppm|mg|f/cc)',
                        str(ev.get('ltTWA', '')), re.I)
        # Brief & Scala pela jornada REAL da avaliação (antes: 0,88 fixo).
        _fr, _rot, _leu = _fator_brief_scala(ev.get('jornada', ''))
        ltc = ltw * _fr
        r['bs'] = {'fator': _fr, 'rotulo': _rot, 'leu_jornada': _leu}
        r['acgih'] = (ltc, (_um.group(1).lower() if _um else 'ppm'), cv < ltc)

    lts = _qf_num(ev.get('ltSTEL', ''))
    dur = _qf_num(ev.get('tempoColeta', ''))
    if lts is not None and cv is not None and dur is not None and dur <= 15:
        r['stel'] = (lts, cv < lts)           # STEL só se aplica a <= 15 min

    veredictos = [v[-1] for v in (r['nr15'], r['acgih'], r['stel']) if v is not None]
    r['ok_geral'] = all(veredictos) if veredictos else None
    return r


def _veredicto_quimico(ev):
    """Veredicto de UMA avaliação no formato que a tela consome.

    Mesma `_classificar_quimico` do laudo — a aba de resultados do laboratório
    NÃO refaz a conta em JavaScript. Antes ela comparava com um limite só
    (`ltNR15 || ltTWA`) e sem a correção Brief & Scala, então mostrava REGULAR
    em avaliação que o laudo assinava como IRREGULAR: com LT-TWA 20 ppm (17,6
    corrigido) uma concentração de 18,5 aparecia verde na tela e reprovava no
    documento. É o mesmo furo que a seção IX tinha e que foi corrigido em
    28/07/2026 — a tela ficou para trás.

    nivel: nd | ok | atencao | ruim | sem_limite
      atencao = NR-15 atendida mas ACGIH/STEL acima (legalmente conforme,
      tecnicamente exposto) — mesma distinção que o quadro IX pinta de amarelo.
    """
    cl = _classificar_quimico(ev)
    if cl['nd']:
        return {'txt': 'NÃO DETECTADO', 'nivel': 'nd', 'detalhe': 'abaixo do limite de detecção'}

    partes = []
    if cl['nr15'] is not None:
        partes.append('NR-15: ' + ('REGULAR' if cl['nr15'][1] else 'IRREGULAR'))
    if cl['acgih'] is not None:
        partes.append('ACGIH: ' + ('REGULAR' if cl['acgih'][2] else 'IRREGULAR')
                      + f' (LT {_fmt_num(cl["acgih"][0])} {cl["acgih"][1]} corrigido)')
    if cl['stel'] is not None:
        partes.append('STEL: ' + ('REGULAR' if cl['stel'][1] else 'IRREGULAR'))

    if cl['ok_geral'] is None:
        return {'txt': 'sem limite', 'nivel': 'sem_limite',
                'detalhe': 'nenhum limite informado para este agente'}
    if cl['ok_geral']:
        return {'txt': 'REGULAR', 'nivel': 'ok', 'detalhe': ' · '.join(partes)}

    nr15_ok = cl['nr15'][1] if cl['nr15'] is not None else None
    return {'txt': 'IRREGULAR',
            'nivel': 'ruim' if nr15_ok is False else 'atencao',
            'detalhe': ' · '.join(partes)}


@app.route('/quimico/classificar', methods=['POST'])
@login_required
def quimico_classificar():
    """Veredicto das avaliações químicas para a aba de resultados do lab.

    A tela manda a lista e recebe o MESMO veredicto que sairá no laudo.
    """
    d = request.get_json(silent=True) or {}
    avals = d.get('avaliacoes') or []
    if not isinstance(avals, list):
        return jsonify({'erro': 'avaliacoes deve ser uma lista'}), 400
    if len(avals) > 500:
        return jsonify({'erro': 'no máximo 500 avaliações por vez'}), 400
    out = []
    for ev in avals:
        if not isinstance(ev, dict):
            out.append({'txt': '—', 'nivel': 'sem_limite', 'detalhe': ''})
            continue
        try:
            out.append(_veredicto_quimico(ev))
        except Exception:
            # Uma avaliação malformada não pode derrubar a grade inteira.
            out.append({'txt': '—', 'nivel': 'sem_limite', 'detalhe': 'não foi possível avaliar'})
    return jsonify({'resultados': out})


def _build_ix_xml(evals):
    """Build section IX Quadro Resumo table from evaluations."""
    # RESULTADO ganhou largura (1860 → 2460) porque agora cabe mais de um
    # veredicto na célula; os outros 3 cederam 200 cada e o total não muda.
    cws   = [2300, 2300, 2300, 2460]
    total = sum(cws)
    bdr   = ('<w:top w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
             '<w:left w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
             '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="000000"/>'
             '<w:right w:val="single" w:sz="4" w:space="0" w:color="000000"/>')

    def _tc(txt, w, bold=False, fill='FFFFFF'):
        b = '<w:b/><w:bCs/>' if bold else ''
        return (f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/>'
                f'<w:tcBorders>{bdr}</w:tcBorders>'
                f'<w:shd w:val="clear" w:color="auto" w:fill="{fill}"/></w:tcPr>'
                f'<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                f'<w:r><w:rPr><w:sz w:val="18"/><w:szCs w:val="18"/>{b}</w:rPr>'
                f'<w:t>{_xe(txt)}</w:t></w:r></w:p></w:tc>')

    def _tr(*cells):
        return '<w:tr>' + ''.join(cells) + '</w:tr>'

    grid = ''.join(f'<w:gridCol w:w="{w}"/>' for w in cws)
    hdr  = _tr(_tc('SETOR', cws[0], bold=True, fill='BDD7EE'),
               _tc('CARGO / LOCAL', cws[1], bold=True, fill='BDD7EE'),
               _tc('AGENTE AVALIADO', cws[2], bold=True, fill='BDD7EE'),
               _tc('RESULTADO', cws[3], bold=True, fill='BDD7EE'))
    rows = ''
    for ev in evals:
        conc = ev.get('concentracao', 'N.D.')
        cs   = str(conc).strip()
        cl   = _classificar_quimico(ev)
        # Um veredicto POR LIMITE, nomeando a norma — o resumo tem que contar a
        # mesma história da conclusão da seção VI (#8 Bernardo).
        if cl['nd']:
            fill = 'C6EFCE'
            res  = (cs + ' (NÃO DETECTADO)') if (cs and cs.upper() not in ('N.D.', 'ND')) else 'N.D. (NÃO DETECTADO)'
        else:
            partes = []
            if cl['nr15'] is not None:
                partes.append('NR-15: ' + ('REGULAR' if cl['nr15'][1] else 'IRREGULAR'))
            if cl['acgih'] is not None:
                partes.append('ACGIH: ' + ('REGULAR' if cl['acgih'][2] else 'IRREGULAR'))
            if cl['stel'] is not None:
                partes.append('STEL: ' + ('REGULAR' if cl['stel'][1] else 'IRREGULAR'))
            res = f'{conc} ({" · ".join(partes)})' if partes else (conc or 'N.D.')

            # Cor: verde só quando TODOS os limites são atendidos. Quando a NR-15
            # passa mas a ACGIH não, vai AMARELO em vez de vermelho — legalmente
            # conforme e tecnicamente acima não é a mesma coisa que estourar a
            # NR-15, e pintar tudo de vermelho apagaria essa diferença.
            _nr15_ok = cl['nr15'][1] if cl['nr15'] is not None else None
            if cl['ok_geral'] is None:
                fill = 'FFFFFF'
            elif cl['ok_geral']:
                fill = 'C6EFCE'
            elif _nr15_ok is False:
                fill = 'FFC7CE'
            else:
                fill = 'FFEB9C'
        # CARGO p/ amostragem pessoal; LOCAL/ambiente (campo trabalhador) p/ amostragem
        # de área/fixa, em vez de "NÃO INFORMADO" (Matheus: usar o ambiente da planilha de campo).
        _cl = (ev.get('cargo', '') or '').strip()
        if (not _cl) or _cl.upper() == 'NÃO INFORMADO':
            _cl = (ev.get('trabalhador', '') or '').strip() or _cl or 'Não informado'
        rows += _tr(_tc(ev.get('setor', ''), cws[0]),
                    _tc(_cl, cws[1]),
                    _tc(ev.get('agente', ''), cws[2]),
                    _tc(res, cws[3], fill=fill))
    return (f'<w:tbl><w:tblPr><w:tblW w:w="{total}" w:type="dxa"/>'
            f'<w:tblBorders>{bdr}</w:tblBorders></w:tblPr>'
            f'<w:tblGrid>{grid}</w:tblGrid>'
            + hdr + rows + '</w:tbl>'
            '<w:p><w:pPr><w:spacing w:after="0"/></w:pPr></w:p>')


# ── Quadros de referência da NR-15, Anexo 3 — texto padrão do laudo de calor ──
# Redação VIGENTE (Portaria SEPRT 1.359/2019). Quadro 1 é renderizado direto
# de _NR15_QUADRO1 (mesma tabela que calcula o limite — uma fonte só).
# (texto, valor, categoria?) — categoria = linha-título (negrito, sem valor)
_NR15_Q2_ROWS = [
    ('SENTADO', '', True),
    ('Em repouso', '100', False),
    ('Trabalho leve com as mãos', '126', False),
    ('Trabalho moderado com as mãos', '153', False),
    ('Trabalho pesado com as mãos', '171', False),
    ('Trabalho leve com um braço', '162', False),
    ('Trabalho moderado com um braço', '198', False),
    ('Trabalho pesado com um braço', '234', False),
    ('Trabalho leve com dois braços', '216', False),
    ('Trabalho moderado com dois braços', '252', False),
    ('Trabalho pesado com dois braços', '288', False),
    ('Trabalho leve com braços e pernas', '324', False),
    ('Trabalho moderado com braços e pernas', '441', False),
    ('Trabalho pesado com braços e pernas', '603', False),
    ('EM PÉ, AGACHADO OU AJOELHADO', '', True),
    ('Em repouso', '126', False),
    ('Trabalho leve com as mãos', '153', False),
    ('Trabalho moderado com as mãos', '180', False),
    ('Trabalho pesado com as mãos', '198', False),
    ('Trabalho leve com um braço', '189', False),
    ('Trabalho moderado com um braço', '225', False),
    ('Trabalho pesado com um braço', '261', False),
    ('Trabalho leve com dois braços', '243', False),
    ('Trabalho moderado com dois braços', '279', False),
    ('Trabalho pesado com dois braços', '315', False),
    ('Trabalho leve com o corpo', '351', False),
    ('Trabalho moderado com o corpo', '468', False),
    ('Trabalho pesado com o corpo', '630', False),
    ('EM PÉ, EM MOVIMENTO', '', True),
    ('Andando no plano, sem carga — 2 km/h', '198', False),
    ('Andando no plano, sem carga — 3 km/h', '252', False),
    ('Andando no plano, sem carga — 4 km/h', '297', False),
    ('Andando no plano, sem carga — 5 km/h', '360', False),
    ('Andando no plano, com carga — 10 kg, 4 km/h', '333', False),
    ('Andando no plano, com carga — 30 kg, 4 km/h', '450', False),
    ('Correndo no plano — 9 km/h', '787', False),
    ('Correndo no plano — 12 km/h', '873', False),
    ('Correndo no plano — 15 km/h', '990', False),
    ('Subindo rampa, sem carga — 5º de inclinação, 4 km/h', '324', False),
    ('Subindo rampa, sem carga — 15º de inclinação, 3 km/h', '378', False),
    ('Subindo rampa, sem carga — 25º de inclinação, 3 km/h', '540', False),
    ('Subindo rampa, com carga de 20 kg — 15º de inclinação, 4 km/h', '486', False),
    ('Subindo rampa, com carga de 20 kg — 25º de inclinação, 4 km/h', '738', False),
    ('Descendo rampa (5 km/h), sem carga — 5º de inclinação', '243', False),
    ('Descendo rampa (5 km/h), sem carga — 15º de inclinação', '252', False),
    ('Descendo rampa (5 km/h), sem carga — 25º de inclinação', '324', False),
    ('Subindo escada (80 degraus/min, degrau de 0,17 m) — sem carga', '522', False),
    ('Subindo escada (80 degraus/min, degrau de 0,17 m) — com carga de 20 kg', '648', False),
    ('Descendo escada (80 degraus/min, degrau de 0,17 m) — sem carga', '279', False),
    ('Descendo escada (80 degraus/min, degrau de 0,17 m) — com carga de 20 kg', '400', False),
    ('OUTRAS ATIVIDADES', '', True),
    ('Trabalho moderado de braços (ex.: varrer, trabalho em almoxarifado)', '320', False),
    ('Trabalho moderado de levantar ou empurrar', '349', False),
    ('Trabalho de empurrar carrinhos de mão, no mesmo plano, com carga', '391', False),
    ('Trabalho de carregar pesos ou com movimentos vigorosos com os braços (ex.: trabalho com foice)', '495', False),
    ('Trabalho pesado de levantar, empurrar ou arrastar pesos (ex.: remoção com pá, abertura de valas)', '524', False),
]
_QBORDER = ('<w:tcBorders>'
            '<w:top w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:left w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:right w:val="single" w:sz="4" w:space="0" w:color="auto"/></w:tcBorders>')

def _calor_cell(text, w, bold=False, fill=None, align='left'):
    rpr = '<w:rPr><w:sz w:val="16"/>' + ('<w:b/>' if bold else '') + '</w:rPr>'
    shd = f'<w:shd w:val="clear" w:color="auto" w:fill="{fill}"/>' if fill else ''
    return (f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/>{shd}{_QBORDER}'
            '<w:vAlign w:val="center"/></w:tcPr>'
            f'<w:p><w:pPr><w:spacing w:after="0"/><w:jc w:val="{align}"/></w:pPr>'
            f'<w:r>{rpr}<w:t xml:space="preserve">{_xe(str(text))}</w:t></w:r></w:p></w:tc>')

def _calor_table(grid, rows_xml):
    g = ''.join(f'<w:gridCol w:w="{w}"/>' for w in grid)
    return ('<w:tbl><w:tblPr><w:tblW w:w="0" w:type="auto"/><w:jc w:val="center"/>'
            '<w:tblBorders>'
            '<w:top w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:left w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:right w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            '</w:tblBorders></w:tblPr>'
            f'<w:tblGrid>{g}</w:tblGrid>{rows_xml}</w:tbl>')

def _calor_heading(text, size=22):
    return ('<w:p><w:pPr><w:spacing w:before="160" w:after="80"/><w:jc w:val="center"/></w:pPr>'
            f'<w:r><w:rPr><w:b/><w:sz w:val="{size}"/></w:rPr>'
            f'<w:t xml:space="preserve">{_xe(text)}</w:t></w:r></w:p>')

def _calor_para(text, size=16, italic=True):
    it = '<w:i/>' if italic else ''
    return ('<w:p><w:pPr><w:spacing w:after="80"/><w:jc w:val="both"/></w:pPr>'
            f'<w:r><w:rPr>{it}<w:sz w:val="{size}"/></w:rPr>'
            f'<w:t xml:space="preserve">{_xe(text)}</w:t></w:r></w:p>')

def _build_quadros_xml():
    """Quadros 1 e 2 do Anexo 3 da NR-15 (redação da Portaria SEPRT 1.359/2019)
    — inseridos na metodologia do laudo."""
    HDR, CAT = 'D9D9D9', 'F2F2F2'
    # Quadro 1: M (W) → IBUTG máx (ºC), em 3 pares de colunas como no texto oficial
    g1 = [1550, 1550, 1550, 1550, 1550, 1550]
    ncol = 3
    per_col = -(-len(_NR15_QUADRO1) // ncol)   # teto da divisão
    r1 = '<w:tr>' + ''.join(
        _calor_cell('M (W)', g1[0], bold=True, fill=HDR, align='center')
        + _calor_cell('IBUTG MÁX (ºC)', g1[1], bold=True, fill=HDR, align='center')
        for _ in range(ncol)) + '</w:tr>'
    for i in range(per_col):
        cells = ''
        for c in range(ncol):
            idx = c * per_col + i
            if idx < len(_NR15_QUADRO1):
                m, lim = _NR15_QUADRO1[idx]
                cells += (_calor_cell(str(m), g1[0], align='center')
                          + _calor_cell(_fx(lim), g1[1], align='center'))
            else:
                cells += _calor_cell('', g1[0]) + _calor_cell('', g1[1])
        r1 += f'<w:tr>{cells}</w:tr>'
    t1 = _calor_table(g1, r1)

    g2 = [7600, 1500]
    r2 = ('<w:tr>' + _calor_cell('Atividade', g2[0], bold=True, fill=HDR)
          + _calor_cell('Taxa metabólica (W)', g2[1], bold=True, fill=HDR, align='center') + '</w:tr>')
    for atv, val, cat in _NR15_Q2_ROWS:
        if cat:
            r2 += ('<w:tr>' + _calor_cell(atv, g2[0], bold=True, fill=CAT)
                   + _calor_cell('', g2[1], fill=CAT) + '</w:tr>')
        else:
            r2 += ('<w:tr>' + _calor_cell(atv, g2[0])
                   + _calor_cell(val, g2[1], align='center') + '</w:tr>')
    t2 = _calor_table(g2, r2)

    pb = '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'
    return (_calor_heading('QUADROS DE REFERÊNCIA — LIMITES DE EXPOSIÇÃO OCUPACIONAL AO CALOR '
                           '(NR-15, ANEXO 3 — Portaria SEPRT Nº 1.359/2019)')
            + _calor_para('O IBUTG é determinado conforme a NHO 06 (2ª edição — 2017) da '
                          'Fundacentro, metodologia adotada pelo Anexo 3 da NR-15: '
                          'IBUTG = 0,7·tbn + 0,3·tg para ambientes internos ou externos sem '
                          'carga solar direta; IBUTG = 0,7·tbn + 0,1·tbs + 0,2·tg para '
                          'ambientes externos com carga solar direta (tbn = temperatura de '
                          'bulbo úmido natural; tg = temperatura de globo; tbs = temperatura '
                          'de bulbo seco).')
            + _calor_heading('Quadro Nº 1 — Limite de exposição ocupacional ao calor '
                             '(taxa de metabolismo M × IBUTG máximo)', 18)
            + t1
            + _calor_heading('Quadro Nº 2 — Taxa metabólica por tipo de atividade', 18)
            + t2
            + _calor_para('A taxa de metabolismo média (M) é ponderada pelo tempo, conforme o '
                          'Quadro Nº 2, e determina o limite de exposição (IBUTG máximo) '
                          'aplicável a cada avaliação pelo Quadro Nº 1.')
            + pb)

def _build_histograma_xml(b64, add_image):
    """Histograma anexado pelo técnico — página própria antes do certificado."""
    if not b64:
        return ''
    try:
        rid, iid = add_image(b64)
    except Exception:
        return ''
    cx, cy = 5760000, 3600000  # fallback ~16:10
    try:
        from PIL import Image as _PILImg
        _raw = b64.split(',', 1)[1] if ',' in b64 else b64
        _im = _PILImg.open(io.BytesIO(base64.b64decode(_raw)))
        iw, ih = _im.size
        if iw and ih:
            MAXW, MAXH = 5760000, 8200000
            cx, cy = MAXW, int(MAXW * ih / iw)
            if cy > MAXH:
                cy, cx = MAXH, int(MAXH * iw / ih)
    except Exception:
        pass
    img = (f'<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r><w:drawing>'
           f'<wp:inline distT="0" distB="0" distL="0" distR="0" wp14:anchorId="7A000001" wp14:editId="7A00ABCD">'
           f'<wp:extent cx="{cx}" cy="{cy}"/><wp:effectExtent l="0" t="0" r="0" b="0"/>'
           f'<wp:docPr id="{iid}" name="Histograma {iid}"/>'
           f'<wp:cNvGraphicFramePr><a:graphicFrameLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/></wp:cNvGraphicFramePr>'
           f'<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
           f'<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
           f'<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
           f'<pic:nvPicPr><pic:cNvPr id="{iid}" name="histograma_{iid}.png"/><pic:cNvPicPr/></pic:nvPicPr>'
           f'<pic:blipFill><a:blip r:embed="{rid}"/><a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
           f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
           f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
           f'</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r></w:p>')
    pb = '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'
    return pb + _calor_heading('HISTOGRAMA DA AVALIAÇÃO') + img


def gerar_calor_bytes(d):
    emp  = d.get('empresa',{})
    aval = d.get('avaliacao',{})
    sets = d.get('setores',[])

    tpl = os.path.join(BASE_DIR,'template_calor.docx')
    with open(tpl,'rb') as f: raw = f.read()
    zin = zipfile.ZipFile(io.BytesIO(raw))
    xml = zin.read('word/document.xml').decode('utf-8')
    rels_xml = zin.read('word/_rels/document.xml.rels').decode('utf-8')

    # ── Logo da empresa ──────────────────────────────────────────────
    # image1.png (rId8) = logo da empresa na capa; substitui ou limpa
    BLANK_PNG = base64.b64decode(
        'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAAC0lEQVQI12NgAAIABQ'
        'AABjkB6QAAAABJRU5ErkJggg==')
    logo_b64 = emp.get('logo','')
    if logo_b64:
        _hdr, _dat = logo_b64.split(',',1) if ',' in logo_b64 else ('', logo_b64)
        logo_bytes = base64.b64decode(_dat)
    else:
        logo_bytes = BLANK_PNG

    # ── Company replacements (Dahana template) ───────────────────────
    razao = emp.get('razaoSocial','')
    xml = xml.replace('COMERCIAL DAHANA LTDA', _xe(razao))
    xml = xml.replace('Av. General Olímpio Mourão Filho, 717', _xe(emp.get('endereco','')))
    xml = xml.replace('00.070.509/0034-79', _xe(emp.get('cnpj','')))
    xml = xml.replace('31710-690', _xe(emp.get('cep','')))
    xml = xml.replace('>Planalto<', f'>{_xe(emp.get("bairro",""))}<')
    xml = xml.replace('>Belo Horizonte<', f'>{_xe(emp.get("cidade","Belo Horizonte"))}<')
    xml = xml.replace('>MG<', f'>{_xe(emp.get("uf","MG"))}<')
    xml = xml.replace('47.11-3-02', _xe(emp.get('cnae','')))
    xml = xml.replace('Comércio varejista de mercadorias em geral, com predominância de produtos alimentícios - supermercados', _xe(emp.get('descricaoCnae','')))
    # Grau de risco: o template traz o da empresa de referência (2, comércio
    # varejista) e nada substituía — todo laudo de calor saía com 2, qualquer que
    # fosse a empresa. Achado em 28/07/2026 montando os golden tests. Mesmo
    # `_rp` do laudo químico (label e valor em parágrafos separados na capa).
    _grau = str(emp.get('grauRisco','')).strip()
    if _grau:
        xml = _rp(xml, 'Grau de Risco', _grau)
    xml = xml.replace('>Thais Taveira<', f'>{_xe(emp.get("contato",""))}<')
    xml = xml.replace('(31)3359-3389', _xe(emp.get('telefone','')))
    # 2º telefone do template (Dahana tem dois) — remove se empresa só forneceu um
    xml = xml.replace('(31)98743-8342', _xe(emp.get('telefone2','') or ''))
    xml = xml.replace('thais.conde@supernosso.com.br', _xe(emp.get('email','')))
    xml = xml.replace('BELO HORIZONTE, MAIO DE 2026.', _xe(aval.get('cidadeCarta','BELO HORIZONTE, MAIO DE 2026')) + '.')
    equip_old = 'Net.Temp – Chrompack Smart TEMP | S/N: IBU0000000209 | Calibração: 24/03/2026 | Certificado Nº 180.646'
    xml = xml.replace(equip_old, _xe(aval.get('equipamento', equip_old)))

    # ── Sector block template ────────────────────────────────────────
    calor2_poses = [m.start() for m in re.finditer('w:val="CALOR2"', xml)]
    if not calor2_poses: raise ValueError("CALOR2 style not found in template")
    sec_start = xml.rfind('<w:p ', 0, calor2_poses[0])
    cert_idx  = xml.find('CERTIFICADO DE CALIBRA', sec_start)
    tbl_end   = xml.rfind('</w:tbl>', sec_start, cert_idx) + len('</w:tbl>')

    # Use first sector as template
    if len(calor2_poses) > 1:
        sec_tpl = xml[sec_start : xml.rfind('<w:p ', 0, calor2_poses[1])]
    else:
        sec_tpl = xml[sec_start:tbl_end]

    tbl_open      = sec_tpl.find('<w:tbl>')
    calor2_para   = sec_tpl[:tbl_open]
    tbl_in_sec    = sec_tpl[tbl_open:]
    first_tr_pos  = tbl_in_sec.find('<w:tr')
    tbl_prefix    = tbl_in_sec[:first_tr_pos]  # <w:tbl> + tblPr + tblGrid
    tpl_rows      = re.findall(r'<w:tr[ >].*?</w:tr>', tbl_in_sec, re.DOTALL)

    row_hdr        = tpl_rows[0]   # header
    row_data_tpl   = tpl_rows[2]   # "Fritadeira" – clean data row
    row_ibutg_tpl  = tpl_rows[4]   # IBUTG médio
    row_act1_tpl   = tpl_rows[5]   # Atividade 01 (com labels "Tipo" e "Taxa M")
    row_actn_tpl   = tpl_rows[6]   # Atividade 02 (continuation)
    row_mmed_tpl   = tpl_rows[8]   # M média
    row_hor_tpl    = tpl_rows[9]   # Horário / vestimenta
    row_conc_hdr   = tpl_rows[10]  # CONCLUSÃO header
    row_conc_tpl   = tpl_rows[11]  # Conclusion text

    # ── Photo support ────────────────────────────────────────────────
    extra_rels  = []
    extra_media = {}
    _img_ctr    = [31]
    _iid_ctr    = [20]

    def _add_image(b64_str):
        _img_ctr[0] += 1; _iid_ctr[0] += 1
        rid = f'rId{_img_ctr[0]}'; iid = _iid_ctr[0]
        if b64_str.startswith('data:'):
            hdr, data = b64_str.split(',', 1)
            ext = 'jpeg' if ('jpeg' in hdr or 'jpg' in hdr) else 'png'
        else:
            data, ext = b64_str, 'jpeg'
        fname = f'media/calor_foto_{_img_ctr[0]}.{ext}'
        extra_rels.append(
            f'<Relationship Id="{rid}" '
            f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" '
            f'Target="{fname}"/>')
        extra_media[f'word/{fname}'] = base64.b64decode(data)
        return rid, iid

    def _photo_cell(rid, iid, caption, w, span, pid):
        CX, CY = 2251710, 1689000
        ah = f'{pid:08X}'; eh = f'{(pid^0xABCD):08X}'
        gs = f'<w:gridSpan w:val="{span}"/>' if span > 1 else ''
        return (
            f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/>{gs}'
            f'<w:tcBorders><w:top w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:left w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:bottom w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:right w:val="single" w:sz="4" w:space="0" w:color="auto"/></w:tcBorders>'
            f'<w:shd w:val="clear" w:color="auto" w:fill="FFFFFF" w:themeFill="background1"/>'
            f'<w:vAlign w:val="center"/></w:tcPr>'
            f'<w:p w14:paraId="{ah}" w14:textId="77777777">'
            f'<w:pPr><w:pStyle w:val="CORPODETEXTO"/><w:ind w:firstLine="0"/><w:jc w:val="center"/></w:pPr>'
            f'<w:r><w:drawing>'
            f'<wp:inline distT="0" distB="0" distL="0" distR="0" wp14:anchorId="{ah}" wp14:editId="{eh}">'
            f'<wp:extent cx="{CX}" cy="{CY}"/><wp:effectExtent l="0" t="0" r="0" b="0"/>'
            f'<wp:docPr id="{iid}" name="Foto {iid}"/>'
            f'<wp:cNvGraphicFramePr><a:graphicFrameLocks xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" noChangeAspect="1"/></wp:cNvGraphicFramePr>'
            f'<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            f'<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
            f'<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
            f'<pic:nvPicPr><pic:cNvPr id="{iid}" name="foto_{iid}.jpeg"/><pic:cNvPicPr/></pic:nvPicPr>'
            f'<pic:blipFill><a:blip r:embed="{rid}">'
            f'<a:extLst><a:ext uri="{{28A0092B-C50C-407E-A947-70E740481C1C}}">'
            f'<a14:useLocalDpi xmlns:a14="http://schemas.microsoft.com/office/drawing/2010/main" val="0"/>'
            f'</a:ext></a:extLst></a:blip>'
            f'<a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
            f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{CX}" cy="{CY}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
            f'</pic:pic></a:graphicData></a:graphic></wp:inline></w:drawing></w:r>'
            f'<w:r><w:t>{_xe(caption)}</w:t></w:r></w:p></w:tc>'
        )

    def _empty_cell(caption, w, span, pid):
        ah = f'{pid:08X}'
        gs = f'<w:gridSpan w:val="{span}"/>' if span > 1 else ''
        return (
            f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/>{gs}'
            f'<w:tcBorders><w:top w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:left w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:bottom w:val="single" w:sz="4" w:space="0" w:color="auto"/>'
            f'<w:right w:val="single" w:sz="4" w:space="0" w:color="auto"/></w:tcBorders>'
            f'<w:shd w:val="clear" w:color="auto" w:fill="FFFFFF" w:themeFill="background1"/>'
            f'<w:vAlign w:val="center"/></w:tcPr>'
            f'<w:p w14:paraId="{ah}" w14:textId="77777777">'
            f'<w:pPr><w:pStyle w:val="CORPODETEXTO"/><w:ind w:firstLine="0"/><w:jc w:val="center"/></w:pPr>'
            f'<w:r><w:t>{_xe(caption)}</w:t></w:r></w:p></w:tc>'
        )

    # Cell widths & gridSpans for photo rows (10-column table, total 11055 DXA)
    _photo_specs = {
        1: [(11055, 10)],
        2: [(3968, 5), (7087, 5)],
        3: [(3683, 4), (3686, 4), (3686, 2)],
    }

    def _photo_rows(si, pontos):
        # Each ponto has up to 3 fotos → one photo row per ponto (3 cells)
        rows = []
        spc = _photo_specs[3]  # always 3 cells: (3683,4), (3686,4), (3686,2)
        for pi, p in enumerate(pontos):
            fotos = p.get('fotos') or []
            # Pad/trim to exactly 3 slots
            fotos = (list(fotos) + [None, None, None])[:3]
            if not any(fotos): continue  # skip row if no photos at all
            cap = p.get('local', '')
            cells = []
            for fi, b64 in enumerate(fotos):
                cw, gs = spc[fi]
                pid = (si * 0x100000 + pi * 0x1000 + fi + 0x500000) & 0xFFFFFFF
                if b64:
                    rid, iid = _add_image(b64)
                    cells.append(_photo_cell(rid, iid, cap, cw, gs, pid))
                else:
                    cells.append(_empty_cell('', cw, gs, pid))
            tr_pid = (si * 0x100000 + pi * 0x1000 + 0x600000) & 0xFFFFFFF
            rows.append(
                f'<w:tr w14:paraId="{tr_pid:08X}" w14:textId="77777777">'
                f'<w:trPr><w:cantSplit/><w:trHeight w:val="3500"/></w:trPr>'
                + ''.join(cells) + '</w:tr>')
        return rows

    # ── Row generators ───────────────────────────────────────────────
    def _uid(base_hex, si, offset):
        return f'{(int(base_hex, 16) + si*0x10000 + offset) & 0xFFFFFFFF:08X}'

    def _bump_ids(r, si, offset=0):
        return re.sub(r'(w14:paraId=")([0-9A-Fa-f]{8})',
                      lambda m: m.group(1)+_uid(m.group(2), si, offset), r)

    def make_data_row(pi, si, p):
        tbn = float(p.get('tbn') or 0); tg = float(p.get('tg') or 0)
        try:
            tbs = float(p.get('tbs') or 0)
        except (TypeError, ValueError):
            tbs = 0.0
        ibutg, formula = _ibutg_ponto(p)
        r = _bump_ids(row_data_tpl, si, (pi+1)*0x100)
        r = r.replace('>Padaria – Fritadeira<', f'>{_xe(p.get("local",f"Ponto {pi+1}"))}<')
        r = r.replace('>15<', f'>{int(float(p.get("tempo",60)))}<')
        r = r.replace('>21,0<', f'>{_fx(tbn)}<')
        r = r.replace('>25,3<', f'>{_fx(tbs)}<')
        r = r.replace('>26,3<', f'>{_fx(tg)}<')
        r = r.replace('IBUTG = (0,7 x 21,0) + (0,3 x 26,3)', formula)
        r = r.replace('>22,6<', f'>{_fx(ibutg)}<')
        return r

    def make_act_row(pi, si, p):
        ativ = p.get('atividade','Trabalho Moderado – De p\xe9, com os bra\xe7os e tronco')
        M = round(float(p.get('M',198)))
        if pi == 0:
            r = _bump_ids(row_act1_tpl, si, 0x8000)
            r = r.replace(' Trabalho Moderado – De p\xe9, com os bra\xe7os e tronco',
                          f' {_xe(ativ)}')
            r = r.replace('>198 W<', f'>{M} W<')
        else:
            r = _bump_ids(row_actn_tpl, si, (pi+1)*0x8000)
            r = r.replace('>Atividade 02:<', f'>Atividade {pi+1:02d}:<')
            r = r.replace(' Trabalho Moderado – De p\xe9, com os bra\xe7os e tronco',
                          f' {_xe(ativ)}')
            r = r.replace('>198<', f'>{M}<')
        return r

    # ── Sector assembly ──────────────────────────────────────────────
    sector_blocks = []
    for si, setor in enumerate(sets):
        nome_s    = setor.get('nome', f'SETOR {si+1}')
        horario   = setor.get('horario','')
        vestimenta = setor.get('vestimenta','Uniforme de Trabalho (0)')
        pontos    = setor.get('pontos',[])

        total_t = sum(float(p.get('tempo',60)) for p in pontos) or 1
        ibutg_m = round(sum(_ibutg_ponto(p)[0]*float(p.get('tempo',60)) for p in pontos)/total_t, 1)
        m_med   = sum(float(p.get('M',198))*float(p.get('tempo',60)) for p in pontos)/total_t
        limite  = get_limite_nr15(m_med)
        ok      = ibutg_m <= limite

        c1 = (f'O limite de tolerância para exposição ao calor, segundo o Quadro Nº 1, do Anexo Nº 3, '
              f'na NR-15, para uma taxa de metabolismo média de {round(m_med)} W é de {_fx(limite)} IBUTG.')
        c2 = (f'O IBUTG médio encontrado foi de {_fx(ibutg_m)} ºC, ' +
              ('não ultrapassando o limite de tolerância.' if ok else 'ultrapassando o limite de tolerância.'))

        # CALOR2 title paragraph
        cp = _bump_ids(calor2_para, si)
        cp = re.sub(r'<w:t>[^<]*Avalia[^<]*</w:t>',
                    f'<w:t>Avaliação {si+1:02d} – Departamento: {_xe(nome_s)}</w:t>', cp)

        # Build rows
        data_rows = ''.join(make_data_row(pi, si, p) for pi, p in enumerate(pontos))

        r_ibutg = _bump_ids(row_ibutg_tpl, si, 0x1000)
        r_ibutg = re.sub(r'IBUTG \(M[eé]dio\) = [0-9,]+ ºC',
                         f'IBUTG (Médio) = {_fx(ibutg_m)} ºC', r_ibutg)

        act_rows = ''.join(make_act_row(pi, si, p) for pi, p in enumerate(pontos))

        r_mmed = _bump_ids(row_mmed_tpl, si, 0x2000)
        r_mmed = r_mmed.replace('>198 W<', f'>{round(m_med)} W<')

        r_hor = _bump_ids(row_hor_tpl, si, 0x3000)
        r_hor = r_hor.replace('>08:50 – 10:01<', f'>{_xe(horario)}<')
        r_hor = r_hor.replace('>Uniforme de Trabalho (0)<', f'>{_xe(vestimenta)}<')

        r_conc = _bump_ids(row_conc_tpl, si, 0x4000)
        r_conc = r_conc.replace(
            'O limite de tolerância para exposição ao calor, segundo o Quadro Nº 1, do Anexo Nº 3, na NR-09, para uma taxa de metabolismo média de 198 W é de 30,2 IBUTG.',
            _xe(c1))
        r_conc = r_conc.replace(
            'O IBUTG médio encontrado foi de 21,7 ºC, não ultrapassando o limite de tolerância.',
            _xe(c2))

        photo_rows = _photo_rows(si, pontos)

        tbl_xml = (tbl_prefix + row_hdr + data_rows + r_ibutg + act_rows +
                   r_mmed + r_hor + row_conc_hdr + r_conc +
                   ''.join(photo_rows) + '</w:tbl>')

        pb = ('<w:p><w:pPr><w:pStyle w:val="CORPODETEXTO"/></w:pPr>'
              '<w:r><w:rPr><w:noProof/></w:rPr><w:br w:type="page"/></w:r></w:p>') if si > 0 else ''
        sector_blocks.append(pb + cp + tbl_xml)

    # Quadros de referência (metodologia) antes dos setores; histograma depois (antes do certificado)
    quadros_xml    = _build_quadros_xml()
    histograma_xml = _build_histograma_xml(aval.get('histograma'), _add_image)
    xml = (xml[:sec_start] + quadros_xml + ''.join(sector_blocks)
           + histograma_xml + xml[tbl_end:])

    # ── ART section (conditional) ────────────────────────────────────
    art_numero = aval.get('artNumero', '').strip()
    if not art_numero:
        p1 = xml.find('RESPOSABILIDADE')
        p2 = xml.find('RESPOSABILIDADE', p1 + 1) if p1 != -1 else -1
        if p2 != -1:
            art_tbl = xml.rfind('<w:tbl>', 0, p2)
            sect_pr = xml.find('<w:sectPr', art_tbl)
            if art_tbl != -1 and sect_pr != -1:
                xml = xml[:art_tbl] + xml[sect_pr:]

    # Inject new image relationships
    if extra_rels:
        rels_xml = rels_xml.replace('</Relationships>',
                                    '\n'.join(extra_rels) + '</Relationships>')

    # Fix IDs duplicados
    _idc = [1]
    def _nwid(m):
        _idc[0] += 1
        return f'w:id="{_idc[0]}"'
    xml = re.sub(r'w:id="\d+"', _nwid, xml)
    _pseen = set()
    def _npar(m):
        import random
        v = m.group(1)
        while v in _pseen:
            v = '%08X' % random.randint(1, 0x7FFFFFFE)
        _pseen.add(v)
        return f'w14:paraId="{v}"'
    xml = re.sub(r'w14:paraId="([^"]+)"', _npar, xml)

    zout = io.BytesIO()
    with zipfile.ZipFile(zout,'w',zipfile.ZIP_DEFLATED) as zw:
        for item in zin.infolist():
            if item.filename == 'word/document.xml':
                zw.writestr(item, xml.encode('utf-8'))
            elif item.filename == 'word/_rels/document.xml.rels':
                zw.writestr(item, rels_xml.encode('utf-8'))
            elif item.filename == 'word/media/image1.png':
                zw.writestr(item, logo_bytes)
            else:
                zw.writestr(item, zin.read(item.filename))
        for path, data in extra_media.items():
            zw.writestr(path, data)
    zin.close()
    return zout.getvalue()

@app.route('/gerar_calor', methods=['POST'])
@login_required
def gerar_calor():
    data = request.json
    if not data or not data.get('empresa',{}).get('razaoSocial','').strip():
        return jsonify({'erro': 'Informe a Razão Social'}), 400
    if not data.get('setores'):
        return jsonify({'erro': 'Adicione pelo menos um setor'}), 400
    try:
        docx_bytes = gerar_calor_bytes(data)
        nome = data['empresa']['razaoSocial']
        nome_safe = re.sub(r'[/\\:*?"<>|]','_', nome)
        filename = f"Laudo de Calor - {nome_safe} - {mes_ano().replace(' / ','_')}.docx"
        usuario = current_user.nome if current_user.is_authenticated else 'anônimo'
        # ── Salvar no banco ────────────────────────────────────────────
        try:
            from controle.db import save_coleta_outros, upsert_empresa
            emp = data.get('empresa', {})
            cnpj = emp.get('cnpj', '') or ''
            empresa_nome = emp.get('razaoSocial', '') or nome
            empresa_id = None
            if empresa_nome:
                try:
                    eid = upsert_empresa(cnpj, empresa_nome)
                    empresa_id = eid
                except Exception:
                    pass
            # Achata setores/pontos do laudo na MESMA estrutura `ibutg_setores`
            # que o wizard de campo usa (1 item por ponto). Passar como chave
            # EXTRA — save_coleta_outros só persiste o que não é coluna fixa.
            # (Antes mandava 'dados_json' pré-serializado, que o save IGNORA →
            #  a coleta gravava dados_json NULO.)
            ibutg_setores = []
            for _s in (data.get('setores') or []):
                _nome = _s.get('nome') or ''
                for _p in (_s.get('pontos') or []):
                    try:
                        _tbn = float(_p.get('tbn') or 0)
                        _tg  = float(_p.get('tg') or 0)
                        _ibu = round(0.7 * _tbn + 0.3 * _tg, 1)
                    except Exception:
                        _ibu = ''
                    # TBS preenchido = céu aberto → IBUTG externo (0,7tbn+0,1tbs+0,2tg),
                    # mesma convenção da planilha de campo e do laudo.
                    try:
                        _tbs = float(_p.get('tbs') or 0)
                        _ibu_ext = round(0.7 * _tbn + 0.1 * _tbs + 0.2 * _tg, 1) if _tbs > 0 else ''
                    except Exception:
                        _ibu_ext = ''
                    ibutg_setores.append({
                        'setor':         _nome or (_p.get('local') or ''),
                        'duracao':       _p.get('tempo') or '',
                        'tbs':           _p.get('tbs'),
                        'tbn':           _p.get('tbn'),
                        'tg':            _p.get('tg'),
                        'ibutg_interno': _ibu,
                        'ibutg_externo': _ibu_ext,
                        'M':             _p.get('M'),
                        'regime':        _p.get('atividade') or '',
                    })
            save_coleta_outros({
                'tipo': 'calor',
                'empresa_id': empresa_id,
                'empresa_nome': empresa_nome,
                'avaliador': usuario,
                'data_coleta': data.get('data') or '',
                'cidade': emp.get('cidade', '') or '',
                'unidade': emp.get('unidade', '') or '',
                'observacao': data.get('observacao', '') or '',
                'status': 'concluida',
                # chaves EXTRA → save_coleta_outros serializa em dados_json
                'ibutg_setores': ibutg_setores,
                'setores': data.get('setores', []),
                'config': data.get('config', {}),
            })
        except Exception as _db_err:
            import traceback; traceback.print_exc()
            # Não impede geração do DOCX
        # ──────────────────────────────────────────────────────────────
        registrar_evento('laudo_calor_gerado', f'Laudo Calor: {nome}',
                         usuario=usuario, ip=request.remote_addr)
        return send_file(io.BytesIO(docx_bytes), as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500

def gerar_quimico_bytes(d):
    emp        = d.get('empresa', {})
    evals      = d.get('avaliacoes', [])
    conf       = d.get('config', {})
    try:
        with open(_GUIA_PATH, encoding='utf-8') as _gf:
            _guia = _json.load(_gf).get('by_cas', {})
    except Exception:
        _guia = {}
    pump       = conf.get('bomba', '')
    pump_sn    = conf.get('bombaSN', '')
    calibrad   = conf.get('calibrador', '')
    fotos_viii = conf.get('fotosVIII', [])   # [{img: b64, desc: str}]
    laudo_imgs = conf.get('laudoImgs', [])   # [b64]  — pages of lab result PDF
    planilha_imgs = conf.get('planilhaImgs', [])  # [b64] — páginas da planilha de campo (anexada)
    logo_b64   = emp.get('logo', '')

    img_ctr    = [60]
    extra_rels = []
    extra_media= {}

    tpl = os.path.join(BASE_DIR, 'template_quimico.docx')
    with open(tpl, 'rb') as f: raw = f.read()
    zin = zipfile.ZipFile(io.BytesIO(raw))
    xml      = zin.read('word/document.xml').decode('utf-8')
    rels_xml = zin.read('word/_rels/document.xml.rels').decode('utf-8')

    # ── Logo (replace image1.png — company logo on cover) ────────────
    BLANK_PNG = base64.b64decode(
        'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAAC0lEQVQI12NgAAIABQ'
        'AABjkB6QAAAABJRU5ErkJggg==')
    if logo_b64:
        hdr, dat = logo_b64.split(',', 1) if ',' in logo_b64 else ('', logo_b64)
        try:
            logo_bytes = base64.b64decode(dat)
        except Exception:
            logo_bytes = BLANK_PNG
    else:
        logo_bytes = BLANK_PNG

    # ── Company replacements ─────────────────────────────────────────
    # Ordem importa: substituir a string MAIS LONGA primeiro
    razao_full = _xe(emp.get('razaoSocial', ''))
    xml = xml.replace('UNIMED BELO HORIZONTE COOPERATIVA DE TRABALHO MEDICO', razao_full)
    tit = _xe(emp.get('titulo', emp.get('razaoSocial', '')))
    xml = xml.replace('CENTRO DE PROMOÇÃO DA SAÚDE UNIMED - UNIDADE BETIM', tit)
    xml = xml.replace('HOSPITAL UNIMED - UNIDADE CONTORNO', tit)
    # Cabeçalho que aparece na faixa superior das tabelas (sem COOPERATIVA)
    xml = xml.replace('UNIMED BELO HORIZONTE', tit)
    xml = xml.replace('Av. Gov. Valadares, ', _xe(emp.get('endereco', '')))
    xml = xml.replace('>619<', f'>{_xe(emp.get("numero",""))}<')
    xml = xml.replace('16.513.178/0036-04', _xe(emp.get('cnpj', '')))
    xml = xml.replace('30130-040', _xe(emp.get('cep', '')))
    xml = xml.replace('>Betim <', f'>{_xe(emp.get("cidade",""))}<')
    xml = xml.replace('>Centro<', f'>{_xe(emp.get("bairro",""))}<')
    xml = xml.replace('>MG<', f'>{_xe(emp.get("uf","MG"))}<')
    xml = xml.replace('65.50-2-00', _xe(emp.get('cnae', '')))
    xml = xml.replace('Planos de saúde', _xe(emp.get('descricaoCnae', '')))
    _grau = str(emp.get('grauRisco', '')).strip()
    if _grau:
        xml = _rp(xml, 'Grau de Risco', _grau)   # célula da capa (valor em parágrafo separado)

    # ── Pump/calibrator name substitution in methodology section ─────
    # A bomba é definida por avaliação — derivamos a(s) bomba(s) usada(s) das
    # avaliações (pump/pump_sn ficam só como fallback legado da config).
    _pump_pairs = []
    for _ev in evals:
        _bk  = _ev.get('bomba')   or pump
        _bsn = _ev.get('bombaSN') or pump_sn
        if _bk and _bsn:
            _nome = _ev.get('bombaLabel') or _PUMP_NAMES.get(_bk, _bk)
            _pair = f'{_nome} — S/N: {_bsn}'
            if _pair not in _pump_pairs:
                _pump_pairs.append(_pair)
    if _pump_pairs:
        xml = xml.replace('AIRLITE – A060502', _xe(' / '.join(_pump_pairs)))  # template placeholder
    if calibrad:
        xml = xml.replace('DEFENDER 510-M', _xe(_CALIB_NAMES.get(calibrad, calibrad)))

    # ── Section VI: eval blocks ──────────────────────────────────────
    pos1   = xml.find('DADOS DA AMOSTRAGEM')
    pos2   = xml.find('DADOS DA AMOSTRAGEM', pos1 + 1)
    tbl1_s = xml.rfind('<w:tbl>', 0, pos1)
    tbl2_s = xml.rfind('<w:tbl>', 0, pos2)
    tbl2_e = xml.find('</w:tbl>', tbl2_s) + 8
    pos8   = xml.find('VIII', tbl2_e)
    sec6_e = xml.rfind('</w:tbl>', tbl2_e, pos8) + 8

    eval_tpl = xml[tbl1_s:tbl2_s]
    OLD_CONC = ('concluímos que C &lt; LT, a concentração é menor que o LT.'
                ' Logo a situação é considerada regular em função da baixa concentração.')
    OLD_BRIEF = ('Como a concentração está abaixo do nível de ação e do limite de '
                 'tolerância, não há necessidade de cálculo para o BRIEF &amp; SCALA.')
    NEW_BRIEF = ('Como a concentração ultrapassou o limite de tolerância, recomenda-se '
                 'a adoção imediata de medidas de controle e nova avaliação após '
                 'implementação.')

    # Mesmo parser do quadro resumo IX (`_classificar_quimico`) — VI e IX precisam
    # classificar REGULAR/IRREGULAR do mesmo jeito, então dividem a função.
    _qf = _qf_num

    blocks = []
    _bs_usados = []          # (fator, rótulo da jornada) que a metodologia declara
    for i, ev in enumerate(evals):
        b = eval_tpl
        b = _rr(b, 'Cargo',                    ev.get('cargo', ''))
        b = _rr(b, 'Trabalhador',              ev.get('trabalhador', ''))
        b = _rr(b, 'Setor',                    ev.get('setor', ''))
        b = _rr(b, 'Jornada de trabalho',      ev.get('jornada', ''))
        b = _rr(b, 'Data da coleta',           ev.get('dataColeta', ''))
        b = _rr(b, 'Data da análise',          ev.get('dataAnalise', ''))
        b = _rr(b, 'Agentes Analisados (CAS)', ev.get('agente', ''))
        # #11 Bernardo: citar o ponto estacionário (amostragem de ponto fixo/ambiental).
        _fonte = ev.get('fonte', '')
        if ev.get('pontoEstacionario'):
            _fonte = (_fonte + ' — ' if _fonte else '') + 'Amostragem estacionária no ponto: ' + ev['pontoEstacionario']
        b = _rr(b, 'Fonte Geradora:',          _fonte)
        # Método de Coleta / Métodos Analíticos / descrição do amostrador —
        # pelo agente via guia_metodos (antes ficavam fixos no texto do template).
        # Método Analítico + descrição do Amostrador (Filtro): a FONTE DA VERDADE
        # é o RA do laboratório (ev['metodo'] / ev['amostradorDesc']); o
        # guia_metodos entra só como fallback quando o RA não trouxe.
        _ge = _guia_entry(ev.get('agente', ''), _guia) or {}
        _md    = str(_ge.get('metodoCod', '')).strip()
        _mdesc = str(_ge.get('metodoDesc', '')).strip()
        _metodo = (ev.get('metodo') or '').strip() or (
            f'{_md} – {_mdesc}' if (_md and _mdesc) else (_md or _mdesc))
        if _metodo:
            b = _rr(b, 'Métodos Analíticos', _metodo)
        _amos = (ev.get('amostradorDesc') or '').strip() or str(_ge.get('amostradorDesc', '')).strip()
        if _amos:
            b = _rr(b, 'Filtro', _amos)   # célula 'Filtro' (vem antes de 'Filtro Número')
        # passivo só quando o método diz "PASSIVO" OU não há vazão NEM bomba na avaliação.
        # (amostragem ativa = bomba + vazão; o RA traz a vazão em vazaoInicial/Final.)
        # Antes checava _ge.get('vazao') do GUIA de métodos — que nunca tem 'vazao' —,
        # então marcava "Amostrador passivo" sempre, mesmo com bomba e vazão.
        _tem_vazao = any(_qf(ev.get(k)) not in (None, 0) for k in ('vazaoInicial', 'vazaoFinal', 'vazao'))
        _tem_bomba = bool(ev.get('bomba') or pump)
        _passivo = ('PASSIVO' in (_metodo or _md).upper()) or (not _tem_vazao and not _tem_bomba)
        b = _rr(b, 'Método de Coleta',
                'Amostrador passivo' if _passivo else 'Bomba de amostragem – NHO 08')
        b = _rr(b, 'Filtro Número',            ev.get('filtroNumero', ''))
        b = _ri(b, 'Vazão Inicial (L/min): ',  ev.get('vazaoInicial', ''))
        b = _ri(b, 'Vazão Fina (L/min): ',     ev.get('vazaoFinal', ''))
        b = _ri(b, 'Tempo de Coleta (Min): ',  ev.get('tempoColeta', ''))
        b = _ri(b, 'Volume Amostrado (L):  ',  ev.get('volume', ''))
        b = _ri(b, 'Tempo de exposição ao agente durante a jornada de trabalho: ',
                ev.get('tempoExposicao', ''))
        b = _ri(b, 'Acessórios utilizados: ',  ev.get('acessorios', ''))
        # Vazão média / variação — só substitui se o payload tiver os dados
        _vi = _qf(ev.get('vazaoInicial'))
        _vf = _qf(ev.get('vazaoFinal'))
        _vm = _qf(ev.get('vazao'))
        if _vm is None and _vi is not None and _vf is not None:
            _vm = (_vi + _vf) / 2
        if _vm is not None:
            b = _rp(b, 'Vazão Média (L/min): ', f'{_vm:.5f}'.replace('.', ','))
            if _vi is not None and _vf is not None and _vm:
                _var = abs(_vi - _vf) / _vm * 100
                b = _ri(b, 'Variação da vazão (%): ', f'{_var:.1f}'.replace('.', ','))
        # Bomba da tabela de amostragem — usa a bomba da avaliação (ou da config)
        _bk = ev.get('bomba') or pump
        _bsn = ev.get('bombaSN') or pump_sn
        if _bk and _bsn:
            _bnome = ev.get('bombaLabel') or _PUMP_NAMES.get(_bk, _bk)
            b = _ri(b, 'Bomba Gravimétrica da marca ', f'{_bnome} {_bsn}')
        if ev.get('ltNR15'):  b = _rr(b, 'Limite de Tolerância ', ev['ltNR15'], nth=1)
        if ev.get('naNR15'):  b = _rr(b, 'Nível de Ação ',        ev['naNR15'], nth=1)
        # ACGIH-TWA: idem ao STEL — sem limite ACGIH-TWA, não exibir o valor-padrão do template.
        if ev.get('ltTWA'):   b = _rr(b, 'Limite de Tolerância ', ev['ltTWA'],  nth=2)
        else:                 b = _rr(b, 'Limite de Tolerância ', 'Não se aplica (agente sem limite ACGIH-TWA)', nth=2)
        if ev.get('naTWA'):   b = _rr(b, 'Nível de Ação ',        ev['naTWA'],  nth=2)
        else:                 b = _rr(b, 'Nível de Ação ',        'Não se aplica', nth=2)
        # STEL (#2 Bernardo): só vale com TLV-STEL E medição ≤15 min (TLV-STEL = média de
        # 15 min). Chumbo e afins não têm STEL → não exibir o valor fantasma do template.
        _durmin = _qf(ev.get('tempoColeta', ''))
        if ev.get('ltSTEL') and _durmin is not None and _durmin <= 15:
            b = _rr(b, 'Limite de Tolerância ', ev['ltSTEL'], nth=3)
        elif ev.get('ltSTEL'):
            b = _rr(b, 'Limite de Tolerância ', 'Não se aplica (TLV-STEL exige medição de até 15 min)', nth=3)
        else:
            b = _rr(b, 'Limite de Tolerância ', 'Não se aplica (agente sem TLV-STEL ACGIH)', nth=3)
        conc = ev.get('concentracao', '')
        if conc:
            # A unidade real vem junto do valor (ex.: "1,0127 mg/m³"). Coloca a
            # unidade no RÓTULO e deixa só o número no valor — o template fixava
            # "(PPM)", errado p/ metais/poeira (mg/m³). 'conc' fica intacto p/ a
            # lógica de conclusão (_qf).
            _um   = re.search(r'(mg/m³|mg/m3|µg/m³|μg/m³|ppm|µg|μg|mg|f/cc)', conc, re.IGNORECASE)
            _uni  = _um.group(1) if _um else ''
            _cval = conc.replace(_uni, '').strip() if _uni else conc
            # "<" no resultado = abaixo do limite de detecção → não detectado (#7 Bernardo)
            if '<' in _cval:
                _cval = '%s (não detectado pelo método de amostragem)' % _cval
            b = _rr(b, 'Concentração (PPM)', _cval, nth=1)
            b = _rr(b, 'Concentração (PPM)', _cval, nth=2)
            b = _rr(b, 'Concentração (PPM)', _cval, nth=3)
            # Unidade no rótulo; 'ppm' sempre MINÚSCULO (Bernardo) e metais/poeira
            # em mg/m³. Default ppm quando o RA não trouxe unidade.
            _uni_disp = (_uni or 'ppm').replace('PPM', 'ppm').replace('Ppm', 'ppm')
            b = b.replace('Concentração (PPM)', 'Concentração (%s)' % _uni_disp)
        _cv  = _qf(conc) if conc not in ('', 'N.D.') else None
        _ltv = _qf(ev.get('ltNR15', '') or ev.get('ltTWA', ''))
        # BRIEF & SCALA (#1 Bernardo): corrige o LT ACGIH–TWA para a jornada REAL da
        # avaliação — semanal ou diária, lida do campo `jornada` (pedido do Bernardo em
        # 26/08/2026). Antes era 0,88 fixo, a conta de 44 h/sem, aplicada a toda empresa:
        # numa escala 12x36 o modelo diário dá 0,50 e o laudo assinava REGULAR exposição
        # acima do limite corrigido. Jornada ilegível volta para 0,88, o comportamento
        # antigo. Calculado SEMPRE que há limite ACGIH-TWA (Bernardo).
        _fb  = lambda x: (f'{x:.4f}'.rstrip('0').rstrip('.')).replace('.', ',')
        # Fator e limite são número de laudo assinado: 2 casas no mínimo. Sem
        # isto saía "FR = 0,5" e "17,6136 ppm" no mesmo parágrafo.
        _ffat = lambda x: (f'{x:.4f}'.rstrip('0').rstrip('.')).replace('.', ',') \
            if len(f'{x:.4f}'.rstrip('0').rstrip('.').split('.')[-1]) > 2 else f'{x:.2f}'.replace('.', ',')
        _flt = lambda x: f'{x:.2f}'.replace('.', ',')
        # extrai o número de qualquer lugar do texto (o LT TWA pode vir "ACGIH – TWA: 200 ppm")
        _mw  = re.search(r'(\d+(?:[.,]\d+)?)', str(ev.get('ltTWA', '')))
        _ltw = float(_mw.group(1).replace(',', '.')) if _mw else None
        if _ltw is not None:
            _uma  = re.search(r'(mg/m³|mg/m3|µg/m³|μg/m³|ppm|mg|f/cc)', str(ev.get('ltTWA', '')), re.I)
            _unia = (_uma.group(1).lower() if _uma else 'ppm')
            _fr, _rot, _leu = _fator_brief_scala(ev.get('jornada', ''))
            _ltc  = _ltw * _fr
            _frtx = _ffat(_fr)
            # A metodologia (escrita uma vez, antes das avaliações) declara os
            # fatores que este documento realmente usou — avaliações de jornadas
            # diferentes no mesmo laudo geram fatores diferentes.
            if (_frtx, _rot) not in _bs_usados:
                _bs_usados.append((_frtx, _rot))
            if _cv is not None:
                _ok = _cv < _ltc
                _bs = ('Aplicando o fator de redução de Brief &amp; Scala (FR = %s) para a '
                       '%s, o Limite de Tolerância ACGIH–TWA corrigido '
                       'é %s %s. A concentração obtida (%s %s) está %s do LT corrigido; situação '
                       'considerada %s.' % (_frtx, _rot, _flt(_ltc), _unia, _fb(_cv), _unia,
                       'abaixo' if _ok else 'ACIMA',
                       'regular' if _ok else 'IRREGULAR, recomendando-se a adoção imediata de medidas de controle e reavaliação'))
            else:
                _bs = ('Limite de Tolerância ACGIH–TWA corrigido pela Brief &amp; Scala '
                       '(FR = %s, %s): %s %s.' % (_frtx, _rot, _flt(_ltc), _unia))
            b = b.replace(OLD_BRIEF, _bs)
        elif _cv is not None and _ltv is not None and _cv >= _ltv:
            b = b.replace(OLD_BRIEF, _xe(NEW_BRIEF))
        new_conc = ev.get('conclusao', '')
        if new_conc:
            b = b.replace(f'>{OLD_CONC}</', f'>{_xe(new_conc)}</')
        else:
            # #3 Bernardo: CONCLUSÃO por LIMITE (NR-15 + ACGIH corrigido + STEL), no lugar de
            # uma frase única só de NR-15; com a sinalização regular/IRREGULAR por limite (#8).
            # Interpretação de layout: mantém a seção CONCLUSÃO, uma frase por limite — o
            # Bernardo pode pedir inline em cada seção depois.
            _und = '<' in str(conc)
            _parts = []
            _ltn = _qf(ev.get('ltNR15', ''))
            # 1ª frase COMPLETA o lead-in do template ("...conforme a NR-15,"), sem repetir "Quanto à NR-15".
            if _ltn is not None:
                if _und:
                    _parts.append('o resultado foi não detectado pelo método de amostragem (abaixo do '
                                  'limite de detecção e, portanto, do Limite de Tolerância), caracterizando situação REGULAR.')
                elif _cv is not None:
                    _parts.append('verifica-se que a concentração obtida (%s) %s Limite de Tolerância (%s), caracterizando situação %s.'
                        % (conc, 'está abaixo do' if _cv < _ltn else 'ultrapassa o', ev.get('ltNR15', ''),
                           'REGULAR' if _cv < _ltn else 'IRREGULAR, recomendando-se a adoção imediata de medidas de controle e reavaliação'))
            if _ltw is not None and _cv is not None and not _und:
                _okt = _cv < _ltc
                _parts.append('Quanto à ACGIH (LT-TWA corrigido pela Brief & Scala para %s = %s %s): a concentração %s limite corrigido, situação %s.'
                    % (_rot, _flt(_ltc), _unia, 'está abaixo do' if _okt else 'ultrapassa o', 'REGULAR' if _okt else 'IRREGULAR'))
            _lts = _qf(ev.get('ltSTEL', ''))
            if _lts is not None and _durmin is not None and _durmin <= 15 and _cv is not None and not _und:
                _oks = _cv < _lts
                _parts.append('Quanto ao TLV-STEL (%s): a concentração %s limite de curta duração, situação %s.'
                    % (ev.get('ltSTEL', ''), 'está abaixo do' if _oks else 'ultrapassa o', 'REGULAR' if _oks else 'IRREGULAR'))
            if _parts:
                b = b.replace(f'>{OLD_CONC}</', '>' + _xe(' '.join(_parts)) + '</')
        b = re.sub(r'w14:paraId="([0-9A-Fa-f]{8})"',
                   lambda m, ii=i: f'w14:paraId="{(int(m.group(1),16)+(ii+1)*0x10000)&0xFFFFFFFE:08X}"',
                   b)
        blocks.append(b)

    # #1 Bernardo: fórmula do Brief & Scala na metodologia (antes das planilhas).
    # Desde 27/08/2026 o texto declara a jornada e o fator que ESTE laudo usou,
    # em vez de afirmar 44 h / 0,88 para toda empresa.
    if _bs_usados:
        _bs_lista = '; '.join('FR = %s para %s' % (f, r) for f, r in _bs_usados)
        _bs_aplic = ('Neste relatório: %s. O limite corrigido (LT ACGIH × FR) é o valor '
                     'comparado com a concentração obtida.' % _bs_lista)
    else:
        _bs_aplic = ('O limite corrigido (LT ACGIH × FR) é o valor comparado com a '
                     'concentração obtida.')
    _bs_meth = ('<w:p><w:pPr><w:spacing w:after="120"/></w:pPr>'
                '<w:r><w:rPr><w:b/><w:bCs/><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
                '<w:t xml:space="preserve">Metodologia de cálculo — Correção pela Brief &amp; Scala: </w:t></w:r>'
                '<w:r><w:rPr><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
                '<w:t xml:space="preserve">os limites da ACGIH são definidos para jornada padrão de 40 horas '
                'semanais (8 horas diárias). Quando a jornada real difere desse padrão, o Limite de Tolerância '
                'ACGIH é corrigido pelo fator de redução de Brief &amp; Scala, aplicado conforme a carga horária '
                'declarada em cada avaliação — modelo semanal: FR = (Hsp/Hsr) × ((168 − Hsr)/(168 − Hsp)), com '
                'Hsp = 40 h e 168 = horas/semana; modelo diário: FR = (Hdp/Hdr) × ((24 − Hdr)/(24 − Hdp)), com '
                'Hdp = 8 h e 24 = horas/dia. O fator é aplicado apenas a jornadas estendidas: em jornada igual ou '
                'inferior à padrão da ACGIH nenhuma redução é feita (FR = 1). ' + _xe(_bs_aplic) +
                '</w:t></w:r></w:p>')
    xml = xml[:tbl1_s] + _bs_meth + ''.join(blocks) + xml[sec6_e:]

    # ── Section bounds (after eval replacement) ──────────────────────
    viii_ts, viii_te = _q_sec_head(xml, 'MEMORIAL FOTOGR')
    ix_ts,   ix_te   = _q_sec_head(xml, 'QUADRO RESUMO')
    x_ts,    x_te    = _q_sec_head(xml, 'RESULTADOS DAS AN')
    xi_ts,   xi_te   = _q_sec_head(xml, 'CERTIFICADO DE CALIBRA')
    xii_ts,  _       = _q_sec_head(xml, 'RESPONSABILIDADE T')

    # ── Build section VIII: Memorial Fotográfico ─────────────────────
    if fotos_viii:
        viii_new = ''
        for fi, foto in enumerate(fotos_viii):
            img = foto.get('img', '')
            desc= foto.get('desc', '')
            if not img:
                continue
            rid, iid = _q_add_b64(img, extra_rels, extra_media, img_ctr)
            viii_new += _q_img_para(rid, iid, cx=5400040, cy=3628800)
            if desc:
                pid = f'{(fi * 0x1000 + 0x770000) & 0xFFFFFFFE:08X}'
                viii_new += (f'<w:p w14:paraId="{pid}" w14:textId="77777777">'
                             '<w:pPr><w:jc w:val="center"/></w:pPr>'
                             f'<w:r><w:t>{_xe(desc)}</w:t></w:r></w:p>')
    else:
        # Sem fotos: parágrafo vazio (evita legendas-fantasma do template)
        viii_new = ('<w:p w14:paraId="77000001" w14:textId="77777777">'
                    '<w:pPr><w:jc w:val="center"/></w:pPr>'
                    '<w:r><w:t></w:t></w:r></w:p>')

    # ── Build section IX: Quadro Resumo ──────────────────────────────
    ix_new = _build_ix_xml(evals) if evals else xml[ix_te:x_ts]

    # ── Build section X: Resultados laboratoriais ────────────────────
    if laudo_imgs:
        # Área útil A4 do template (EMU): largura 5.400.040, altura 7.811.770.
        # As páginas do laudo são retrato (A4) → preenche a largura e calcula a
        # altura pela proporção REAL (antes ia num box fixo paisagem e espremia).
        _TXT_W, _TXT_H = 5400040, 7811770
        x_new = ''
        for li, img in enumerate(laudo_imgs):
            if not img:
                continue
            rid, iid = _q_add_b64(img, extra_rels, extra_media, img_ctr)
            _raw = base64.b64decode(img.split(',', 1)[1] if ',' in img else img)
            _w, _h = _img_wh(_raw)
            if _w and _h:
                _cx = _TXT_W; _cy = int(_cx * _h / _w)
                if _cy > _TXT_H:
                    _cy = _TXT_H; _cx = int(_cy * _w / _h)
            else:
                _cx, _cy = _TXT_W, int(_TXT_W * 1.414)   # fallback A4 retrato
            x_new += _q_img_para(rid, iid, cx=_cx, cy=_cy)
    else:
        x_new = xml[x_te:xi_ts]

    # ── Build section XI: Certificado de Calibração ──────────────────
    xi_new = ''
    # Certificados das bombas — derivados das avaliações (bomba é por avaliação;
    # pump/pump_sn ficam só como fallback legado da config).
    _cert_pairs = []
    for _ev in evals:
        _bk  = _ev.get('bomba')   or pump
        _bsn = _ev.get('bombaSN') or pump_sn
        if _bk and _bsn and (_bk, _bsn) not in _cert_pairs:
            _cert_pairs.append((_bk, _bsn))
    for _bk, _bsn in _cert_pairs:
        if _bk in _PUMP_CERT_PAGES and _bsn in _PUMP_CERT_PAGES[_bk]:
            for pg in range(1, _PUMP_CERT_PAGES[_bk][_bsn] + 1):
                path = os.path.join(CERTS_DIR, f'cert_{_bk}_{_bsn}_p{pg}.jpg')
                if os.path.exists(path):
                    rid, iid = _q_add_file(path, extra_rels, extra_media, img_ctr)
                    with open(path, 'rb') as _cf: _cx, _cy = _q_fit_extent(_cf.read())
                    xi_new += _q_img_para(rid, iid, cx=_cx, cy=_cy)
    if calibrad and calibrad in _CALIB_CERT_PAGES:
        for pg in range(1, _CALIB_CERT_PAGES[calibrad] + 1):
            path = os.path.join(CERTS_DIR, f'cert_{calibrad}_p{pg}.jpg')
            if os.path.exists(path):
                rid, iid = _q_add_file(path, extra_rels, extra_media, img_ctr)
                with open(path, 'rb') as _cf: _cx, _cy = _q_fit_extent(_cf.read())
                xi_new += _q_img_para(rid, iid, cx=_cx, cy=_cy)
    if not xi_new:
        xi_new = xml[xi_te:xii_ts]

    # ── Seção PLANILHA DE CAMPO (anexo, opcional) — páginas embutidas como imagens ──
    # Reaproveita o mesmo embed da seção X (resultados do lab). Fonte das páginas:
    # planilha de campo de uma coleta finalizada, escolhida no módulo de laudo.
    pl_new = ''
    if planilha_imgs:
        pl_new = ('<w:p><w:pPr><w:spacing w:before="240" w:after="120"/><w:jc w:val="center"/></w:pPr>'
                  '<w:r><w:rPr><w:b/><w:bCs/><w:sz w:val="24"/><w:szCs w:val="24"/></w:rPr>'
                  '<w:t>PLANILHA DE CAMPO</w:t></w:r></w:p>')
        _TXT_W2, _TXT_H2 = 5400040, 7811770
        for img in planilha_imgs:
            if not img:
                continue
            rid, iid = _q_add_b64(img, extra_rels, extra_media, img_ctr)
            _raw = base64.b64decode(img.split(',', 1)[1] if ',' in img else img)
            _w, _h = _img_wh(_raw)
            if _w and _h:
                _cx = _TXT_W2; _cy = int(_cx * _h / _w)
                if _cy > _TXT_H2:
                    _cy = _TXT_H2; _cx = int(_cy * _w / _h)
            else:
                _cx, _cy = _TXT_W2, int(_TXT_W2 * 1.414)
            pl_new += _q_img_para(rid, iid, cx=_cx, cy=_cy)

    # ── Assemble final XML ────────────────────────────────────────────
    xml = (xml[:viii_te] + viii_new +
           xml[ix_ts:ix_te] + ix_new +
           pl_new +
           xml[x_ts:x_te]  + x_new  +
           xml[xi_ts:xi_te] + xi_new +
           xml[xii_ts:])

    # ── Relationships & output ────────────────────────────────────────
    if extra_rels:
        rels_xml = rels_xml.replace('</Relationships>',
                                    '\n'.join(extra_rels) + '</Relationships>')

    # ── Índice (TOC): troca as entradas "Técnica de Enfermagem" (cargo) do
    # template Unimed pelos CARGOS REAIS e remove as captions "AVALIAÇÃO
    # REALIZA" da seção VIII (essas saíram do corpo). As entradas de seção
    # (I–XII) ficam; o updateFields no settings.xml recalcula a numeração.
    def _toc_span(anchor):
        a = xml.find('w:anchor="%s"' % anchor)
        if a < 0:
            return None
        return (xml.rfind('<w:p ', 0, a), xml.find('</w:p>', a) + 6)

    _m146 = _toc_span('_Toc225948146')   # 1ª entrada de cargo (vira modelo)
    if _m146:
        _mp  = re.search(r'<w:pPr>.*?</w:pPr>', xml[_m146[0]:_m146[1]], re.S)
        _ppr = _mp.group(0) if _mp else '<w:pPr><w:pStyle w:val="Sumrio3"/></w:pPr>'
        _seen, _cargos = set(), []
        for _ev in evals:                 # cargos reais, distintos, na ordem
            _c = (_ev.get('cargo') or '').strip()
            if _c and _c.upper() not in _seen:
                _seen.add(_c.upper()); _cargos.append(_c)
        _novas = ''.join(
            '<w:p>%s<w:r><w:t xml:space="preserve">%s</w:t></w:r></w:p>' % (_ppr, _xe(_c))
            for _c in _cargos) or ('<w:p>%s</w:p>' % _ppr)
        xml = xml[:_m146[0]] + _novas + xml[_m146[1]:]
    for _anc in ('_Toc225948147', '_Toc225948149', '_Toc225948150'):
        _sp = _toc_span(_anc)             # 2ª entrada de cargo + 2 captions
        if _sp:
            xml = xml[:_sp[0]] + xml[_sp[1]:]

    # ── Bookmarks: renumera mantendo o PAR start/end (os únicos w:id do doc são
    # bookmarks). O renumber antigo numerava start e end separadamente → par
    # quebrado → PAGEREF do índice falhava ("Erro! Indicador não definido").
    _bid, _bstk = [1000], {}
    def _fix_bm(m):
        tag, oid = m.group(1), m.group(2)
        if tag == 'bookmarkStart':
            _bid[0] += 1
            _bstk.setdefault(oid, []).append(_bid[0])
            return '<w:bookmarkStart w:id="%d"' % _bid[0]
        _lst = _bstk.get(oid)
        return '<w:bookmarkEnd w:id="%d"' % (_lst.pop() if _lst else _bid[0] + 1)
    xml = re.sub(r'<w:(bookmarkStart|bookmarkEnd) w:id="(\d+)"', _fix_bm, xml)

    zout = io.BytesIO()
    with zipfile.ZipFile(zout, 'w', zipfile.ZIP_DEFLATED) as zw:
        for item in zin.infolist():
            if item.filename == 'word/document.xml':
                zw.writestr(item, xml.encode('utf-8'))
            elif item.filename == 'word/_rels/document.xml.rels':
                zw.writestr(item, rels_xml.encode('utf-8'))
            elif item.filename in ('word/media/image1.png', 'word/media/image3.png'):
                # image1 = logo do cliente na capa; image3 = a MESMA logo na faixa
                # do cabeçalho (slot do cliente). Ambas eram a logo Unimed do
                # template — trocadas pela logo do cliente (ou em branco se não houver).
                zw.writestr(item, logo_bytes)
            elif item.filename == 'word/settings.xml':
                # Força o Word a atualizar campos ao abrir → recalcula a NUMERAÇÃO
                # de página do índice (PAGEREF). Os nomes das entradas já são
                # corrigidos acima (cargos reais; captions removidas).
                _s = zin.read(item.filename).decode('utf-8')
                if 'w:updateFields' not in _s:
                    _s = re.sub(r'(<w:settings\b[^>]*>)',
                                r'\1<w:updateFields w:val="true"/>', _s, count=1)
                zw.writestr(item, _s.encode('utf-8'))
            else:
                zw.writestr(item, zin.read(item.filename))
        for path, data in extra_media.items():
            zw.writestr(path, data)
    zin.close()
    return zout.getvalue()


@app.route('/gerar_quimico', methods=['POST'])
@login_required
def gerar_quimico():
    data = request.json
    if not data or not data.get('empresa', {}).get('razaoSocial', '').strip():
        return jsonify({'erro': 'Informe a Razão Social'}), 400
    if not data.get('avaliacoes'):
        return jsonify({'erro': 'Adicione pelo menos uma avaliação'}), 400
    try:
        docx_bytes = gerar_quimico_bytes(data)
        nome = data['empresa']['razaoSocial']
        nome_safe = re.sub(r'[/\\:*?"<>|]', '_', nome)
        filename = f"Análise Química - {nome_safe} - {mes_ano().replace(' / ','_')}.docx"
        usuario = current_user.nome if current_user.is_authenticated else 'anônimo'
        # O que FOI PARA O PAPEL vira dado (fonte='digitado'). É o outro lado da
        # comparação com o laudo do lab: divergindo, abre divergência e nenhum
        # dos dois é sobrescrito.
        try:
            from controle.resultado_lab import gravar_muitos
            _res = gravar_muitos(data.get('avaliacoes') or [], 'digitado',
                                 f'laudo químico · {usuario}')
            for _d in _res.get('divergencias', []):
                registrar_evento('resultado_lab_divergente', _d,
                                 usuario=usuario, ip=request.remote_addr)
        except Exception as _e:
            import logging
            logging.getLogger(__name__).warning(
                '[resultado_lab] gravar do laudo falhou: %s', _e)
        registrar_evento('laudo_quimico_gerado', f'Laudo Químico: {nome}',
                         usuario=usuario, ip=request.remote_addr)
        return send_file(io.BytesIO(docx_bytes), as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'erro': f'Erro interno: {str(e)}'}), 500


# ── API: Usuários/Técnicos ────────────────────────────────────────────
@app.route('/api/tecnicos')
@login_required
def api_tecnicos():
    """Técnicos elaboradores do select dos documentos: usuários com login +
    catálogo de TSTs (quem executa/assina sem usar o sistema). Já vem com o
    registro MTE, para ninguém digitar registro à mão."""
    try:
        from controle.db import list_tecnicos_documento
        return jsonify(list_tecnicos_documento())
    except Exception as e:
        print(f'[api_tecnicos] {e}')
        try:
            with get_db() as conn:
                rows = conn.execute(
                    'SELECT id, nome, registro_mte FROM usuarios WHERE ativo=1 ORDER BY nome'
                ).fetchall()
            return jsonify([row_to_dict(r) for r in rows])
        except Exception:
            return jsonify([])


@app.route('/api/me', methods=['GET', 'PATCH'])
def api_me():
    """Retorna dados do usuário logado (ou null). PATCH atualiza nome e registro_mte."""
    if not current_user.is_authenticated:
        return jsonify(None)
    if request.method == 'PATCH':
        d = request.json or {}
        nome = (d.get('nome') or '').strip()
        mte  = (d.get('registro_mte') or '').strip()
        try:
            from controle.db import get_db
            with get_db() as conn:
                conn.execute(
                    'UPDATE usuarios SET nome=?, registro_mte=? WHERE id=?',
                    (nome or current_user.nome, mte or None, current_user.id)
                )
            # Atualizar objeto em memória
            current_user.nome = nome or current_user.nome
            current_user.registro_mte = mte or None
        except Exception as e:
            return jsonify({'erro': str(e)}), 500
        return jsonify({'ok': True})
    return jsonify({
        'id':           current_user.id,
        'nome':         current_user.nome,
        'email':        getattr(current_user, 'email', ''),
        'registro_mte': current_user.registro_mte,
        'role':         current_user.role,
    })


# ── API: Guia de Métodos lookup ──────────────────────────────────────
@app.route('/api/cargos')
def api_cargos():
    return jsonify(CARGOS_SUGESTOES)

@app.route('/api/agentes')
def api_agentes():
    result = []
    for ghe, agentes in GHE_AGENTES.items():
        for ag in agentes:
            tipo = ag[0]
            if tipo == 'ruido':
                _, db, nivel, acao_nr15, insalub = ag
                result.append({
                    'ghe': ghe,
                    'tipo': 'Ruído',
                    'agente': f'Ruído — {db}',
                    'nivel': nivel,
                    'risco': nivel,
                    'insalubridade': insalub,
                    'acao_nr15': acao_nr15,
                    'valor': db,
                })
            elif tipo == 'quant':
                _, nome, subtipo, lt, na, conc, risco, insalub = ag
                result.append({
                    'ghe': ghe,
                    'tipo': 'Químico',
                    'agente': nome,
                    'nivel': risco,
                    'risco': risco,
                    'insalubridade': insalub,
                    'lt': lt,
                    'na': na,
                    'concentracao': conc,
                })
            elif tipo == 'ergon':
                _, nome, risco = ag
                result.append({
                    'ghe': ghe,
                    'tipo': 'Ergonômico',
                    'agente': nome,
                    'nivel': risco,
                    'risco': risco,
                    'insalubridade': False,
                })
            elif tipo == 'acid':
                _, nome, risco = ag
                result.append({
                    'ghe': ghe,
                    'tipo': 'Acidente',
                    'agente': nome,
                    'nivel': risco,
                    'risco': risco,
                    'insalubridade': False,
                })
    return jsonify(result)

@app.route('/api/metodos')
def api_metodos():
    try:
        with open(_GUIA_PATH, encoding='utf-8') as f:
            return _json.dumps(_json.load(f), ensure_ascii=False), 200, {'Content-Type': 'application/json'}
    except Exception:
        return jsonify({'by_cas': {}, 'by_name': {}}), 200


# ── API: Pump serial numbers ─────────────────────────────────────────
@app.route('/api/pump_sns')
def api_pump_sns():
    pump = request.args.get('model', '')
    return jsonify(_PUMP_SN.get(pump, []))


# ── API: Convert lab result PDF to base64 JPG images + extract data ──
def ler_laudo_ra_pdf(raw):
    """Lê UM PDF de RA e devolve (paginas_jpeg_base64, dadosExtraidos).

    Extraído da rota `/api/convert_laudo` sem mudar uma linha da extração: o
    mesmo laudo precisa ser lido também quando o PDF vem da CAIXA do laboratório
    (o servidor já recebe esses anexos), e não só do upload manual do técnico.
    """
    import fitz
    import re as _re
    doc = fitz.open(stream=raw, filetype='pdf')
    imgs = []
    full_text = ''
    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
        imgs.append('data:image/jpeg;base64,' + base64.b64encode(pix.tobytes('jpeg')).decode())
        full_text += page.get_text() + '\n'
    doc.close()

    def _g(patterns):
        for pat in patterns:
            m = _re.search(pat, full_text, _re.IGNORECASE | _re.MULTILINE)
            if m:
                return m.group(1).strip().strip('—–-').strip()
        return ''

    # Trabalhador: linha ALL-CAPS imediatamente antes de "Função:"
    _CAPS = r'[A-ZÀ-ÖØ-Ý]'
    trabalhador = _g([r'(' + _CAPS + r'[A-ZÀ-ÖØ-Ý\s\.]+?)\s*\nFun'])

    # Cargo: depois de "Função:"
    cargo = _g([r'Fun[cç][aã]o:\s*([^\n]+)'])

    # Setor: linha ALL-CAPS seguida de outra linha ALL-CAPS (nome do responsável)
    setor = ''
    _ms = _re.search(r'\n([A-Z][A-Z\s]{2,29})\n[A-Z]+ [A-Z]+\n', full_text)
    if _ms:
        setor = _ms.group(1).strip()

    # Nº amostrador (ex: FL22335, EC81076A, TCP1671AV2) — token antes de "Nº do Branco de Campo"
    filtro = _g([
        r'\n([A-Z][A-Z0-9]{4,11})\s*\n+\s*N[^\n]*do Branco de Campo',
        r'\b([A-Z]{2}\d{5}[A-Z]?\d?)\b',
    ])

    # Data coleta: data na linha imediatamente anterior a "Tempo de Amostragem"
    data_col = _g([r'(\d{2}/\d{2}/\d{4})\s*\nTempo'])

    # Vazão: "2,006  L/Min" → captura o número com vírgula
    vazao_raw = _g([r'([\d]+,[\d]+)\s+L/[Mm]in'])
    vazao_fmt = ''
    if vazao_raw:
        try:
            vazao_fmt = '{:.4f}'.format(float(vazao_raw.replace(',', '.'))).replace('.', ',')
        except:
            vazao_fmt = vazao_raw

    # Volume: "Volume de Ar Amostrado: 0,0802  m³" → converte m³ → L
    volume_raw = _g([r'Volume de Ar Amostrado:\s*([\d,]+)\s*m'])
    volume_fmt = ''
    if volume_raw:
        try:
            v = float(volume_raw.replace(',', '.'))
            if v < 1:
                v *= 1000
            volume_fmt = '{:.3f}'.format(v).replace('.', ',')
        except:
            volume_fmt = volume_raw

    # Tempo de amostragem: "Tempo de Amostragem (H): 0:40:00"
    tempo_raw = _g([r'Tempo de Amostragem[^:]*:\s*([\d:]+)'])
    tempo_min = ''
    if tempo_raw:
        parts = tempo_raw.split(':')
        try:
            tempo_min = str(int(parts[0]) * 60 + int(parts[1]))
        except:
            tempo_min = tempo_raw

    # ── Tabela de RESULTADO do lab (UniScientific RA) ───────────────
    # Estrutura por linha na tabela de resultados:
    #   <agente> \n <unidade> \n <resultado> \n
    #   <NR15 MP 8h> \n <NR15 Teto> \n <ACGIH TWA> \n <ACGIH STEL> \n ...
    _UNI = r'ppm|mg/m³|mg/m3|mg|µg|μg|f/cc'
    # Poeira/metal traz a FRAÇÃO colada na unidade: 'mg/m³ (I)'. A unidade era
    # exigida pura, então o laudo inteiro saía sem agente e sem concentração — em
    # 05/08/2026, 3 dos 7 laudos do RA 81962595 (Cromo metálico, cassete IOM).
    _FRAC = r'(?:\s*\(\s*[IRT]\s*\))?'
    # Separar a fração da unidade é a MESMA regra do lado do backfill (que também
    # recebe 'mg/m³ (I)'): mora em controle.resultado_lab para não virar duas verdades.
    from controle.resultado_lab import separar_fracao as _separar_fracao
    agente = ''
    concentracao = ''
    fracao = ''
    lt_nr15 = lt_twa = lt_stel = ''
    _mr = _re.search(
        r'\n([^\n<>]+)\n((?:' + _UNI + r')' + _FRAC + r')\n([^\n]+)\n([^\n]+)\n([^\n]+)\n([^\n]+)\n([^\n]+)',
        full_text, _re.IGNORECASE)
    if _mr:
        agente    = _mr.group(1).strip().strip('—–-').strip()
        unidade, fracao = _separar_fracao(_mr.group(2))
        resultado = _mr.group(3).strip().replace(' ', '')
        concentracao = f'{resultado} {unidade}' if resultado else ''

        def _lim(s):
            s = (s or '').strip()
            return s if _re.match(r'^[<>]?\s*[\d.,]+$', s) else ''
        lt_nr15 = _lim(_mr.group(4))   # NR-15 MP 8h
        lt_twa  = _lim(_mr.group(6))   # ACGIH TWA
        lt_stel = _lim(_mr.group(7))   # ACGIH STEL
    else:
        agente = _g([r'\n([^\n<>]+)\n(?:' + _UNI + r')' + _FRAC + r'\n'])

    # Nível de Ação = metade do LT (NR-09/PGR p/ agentes químicos)
    def _na(lt):
        try:
            v = float(lt.replace(',', '.')) / 2
            return '{:.4f}'.format(v).rstrip('0').rstrip('.').replace('.', ',')
        except Exception:
            return ''
    na_nr15 = _na(lt_nr15)
    na_twa  = _na(lt_twa)

    # Nº do RA — "Relatório de Análise - Nº 81962595-1". Guarda só a BASE (antes do
    # '-'), que é o número pelo qual o laudo é chamado; o sufixo é a sequência do
    # tubo dentro do RA. Sem isto o resultado gravado ficava sem RA e não havia como
    # achar "o que veio no RA 81962595".
    _mra = _re.search(r'Relat[óo]rio de An[áa]lise\s*-?\s*N[ºo°]?\s*([\d-]+)',
                      full_text, _re.IGNORECASE)
    ra_num = _mra.group(1).split('-')[0].strip() if _mra else ''

    # Data da análise (processamento) = data isolada logo após a emissão.
    data_analise = _g([r'S[ãa]o Bernardo do Campo,\s*\d{2}/\d{2}/\d{4}\.\s*\n\s*(\d{2}/\d{2}/\d{4})'])

    # Método analítico + descrição do amostrador — DIRETO DO RA (fonte da
    # verdade). Antes vinham do guia_metodos genérico e saíam errados.
    _mm = _re.search(r'M[ÉE]TODO\s*\(?s?\)?\s*\n(.*?)\n\s*4\s*[-–]', full_text, _re.S | _re.I)
    metodo = _re.sub(r'\s+', ' ', _mm.group(1)).strip() if _mm else ''
    # A descrição (CASSETE/TUBO/…) vem entre "Nº do Branco de Campo:" e
    # "Informações da amostragem". No texto do fitz, o rótulo "Descrição do
    # Amostrador:" é seguido pelos OUTROS rótulos (layout 2 colunas), então
    # ancorar nele pegava lixo ("Data da Amostragem: Vazão…").
    _am = _re.search(
        r'Branco de Campo:[^\n]*\n(.*?)\n\s*Informa[çc][õo]es da amostragem',
        full_text, _re.S | _re.I)
    amostrador_desc = _re.sub(r'\s+', ' ', _am.group(1)).strip().rstrip('.') if _am else ''
    # guarda: se o layout vier diferente e capturar rótulos, descarta
    if _re.search(r'Data da Amostragem|Vaz[ãa]o M[ée]dia|Funcion|Respons[áa]vel|Descri[çc]',
                  amostrador_desc, _re.I):
        amostrador_desc = ''

    dados = {k: v for k, v in {
        'ra_num':       ra_num,
        'filtroNumero': filtro,
        'trabalhador':  trabalhador,
        'cargo':        cargo,
        'setor':        setor,
        'dataColeta':   data_col,
        'dataAnalise':  data_analise,
        'vazaoInicial': vazao_fmt,
        'vazaoFinal':   vazao_fmt,
        'volume':       volume_fmt,
        'tempoColeta':  tempo_min,
        'agente':       agente,
        'concentracao': concentracao,
        'fracao':       fracao,
        'metodo':         metodo,
        'amostradorDesc': amostrador_desc,
        'ltNR15':       lt_nr15,
        'naNR15':       na_nr15,
        'ltTWA':        lt_twa,
        'naTWA':        na_twa,
        'ltSTEL':       lt_stel,
    }.items() if v}

    return imgs, dados


def _guardar_resultados_lidos(lista, origem=''):
    """Todo laudo LIDO vira dado (`resultados_lab`, fonte='pdf').

    O valor já era extraído e descartado; guardar aqui é o que permite comparar
    depois com o que o técnico lançou. Nunca derruba a resposta da rota: se a
    gravação falhar, a leitura do PDF continua servindo a tela.
    """
    try:
        from controle.resultado_lab import gravar_muitos
        return gravar_muitos(lista, 'pdf', origem)
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('[resultado_lab] gravar do PDF falhou: %s', e)
        return {'gravados': 0, 'ignorados': len(lista or []), 'divergencias': []}


@app.route('/api/convert_laudo', methods=['POST'])
@login_required
def api_convert_laudo():
    """RA que o TÉCNICO sobe à mão. Mesma leitura de `ler_laudo_ra_pdf`."""
    try:
        import fitz  # noqa: F401  — só para dar erro claro se faltar a lib
    except ImportError:
        return jsonify({'erro': 'pymupdf nao instalado'}), 500
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'erro': 'Nenhum arquivo'}), 400
        imgs, dados = ler_laudo_ra_pdf(f.read())
        _guardar_resultados_lidos([dados], origem=(f.filename or 'upload'))
        return jsonify({'paginas': imgs, 'dadosExtraidos': dados})
    except Exception as e:
        import traceback
        traceback.print_exc(); return jsonify({'erro': str(e)}), 500


@app.route('/api/laudo_do_lab')
@login_required
def api_laudo_do_lab():
    """Puxa o(s) laudo(s) de RA direto da CAIXA do laboratório, sem upload.

    O servidor já recebe esses anexos (o backfill abre 384 PDFs por dia só para
    casar o amostrador), então o técnico baixar o anexo do Outlook e subir aqui é
    trabalho que a máquina já fez. Aceita `?ra=<numero>` ou `?amostrador=<codigo>`
    e devolve a MESMA estrutura do upload, um item por PDF — o front reaproveita
    o caminho de importação em massa sem mudar a lógica de cruzamento.
    """
    ra_alvo = re.sub(r'\D', '', request.args.get('ra', '') or '')
    cod_alvo = re.sub(r'\s+', '', (request.args.get('amostrador', '') or '')).upper()
    if not ra_alvo and not cod_alvo:
        return jsonify({'erro': 'Informe o nº do RA ou o código do amostrador'}), 400
    try:
        from controle.lab_inbox import (_mailboxes, _search_lab_emails, _ra_do_assunto,
                                        _codigo_do_anexo_ra, _norm)
        from controle.graph import graph_get
    except Exception as e:
        return jsonify({'erro': f'integração com a caixa do lab indisponível: {e}'}), 500

    MAX_PDFS = 8          # a resposta carrega as páginas em base64 — não devolver a caixa toda
    MAX_EMAILS = 60       # teto da varredura cega (por código desconhecido)

    # Buscar por CÓDIGO sem saber o RA custa uma chamada de anexos por e-mail de RA
    # da caixa — passava de 45 s. O banco já sabe em qual RA o tubo veio (gravado a
    # cada laudo lido, e pelo backfill noturno): com isso a busca por código cai no
    # mesmo caminho rápido da busca por RA.
    ras_conhecidos = set()
    if cod_alvo and not ra_alvo:
        try:
            from controle.resultado_lab import ras_do_amostrador
            ras_conhecidos = set(ras_do_amostrador(cod_alvo))
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning('[laudo_do_lab] RA por código: %s', e)

    laudos, vistos, truncado = [], set(), False
    varridos, varredura_limitada = 0, False
    try:
        for box in _mailboxes():
            for e in _search_lab_emails(box):
                ra_email = _ra_do_assunto(e['subject']) or ''
                if ra_alvo and ra_email != ra_alvo:
                    continue
                if ras_conhecidos and ra_email not in ras_conhecidos:
                    continue
                if not ra_alvo and not ras_conhecidos:
                    # varredura cega: sem teto, uma busca por código lia a caixa toda
                    if varridos >= MAX_EMAILS:
                        varredura_limitada = True
                        continue
                    varridos += 1
                try:
                    metas = graph_get(f"/users/{box}/messages/{e['id']}/attachments"
                                      f"?$select=id,name,contentType").get('value', [])
                except Exception:
                    continue
                for meta in metas:
                    nome = meta.get('name', '') or ''
                    if not nome.lower().endswith('.pdf'):
                        continue
                    cod_pdf = _norm(_codigo_do_anexo_ra(nome))
                    # Busca por amostrador: o código vem no nome do anexo. Sem esse
                    # filtro viria o laudo de todos os tubos daquele e-mail.
                    if cod_alvo and cod_alvo not in (cod_pdf or '') and cod_pdf not in cod_alvo:
                        continue
                    if nome in vistos:
                        continue
                    vistos.add(nome)
                    if len(laudos) >= MAX_PDFS:
                        truncado = True
                        continue
                    try:
                        full = graph_get(f"/users/{box}/messages/{e['id']}/attachments/{meta['id']}")
                        raw = base64.b64decode(full.get('contentBytes') or '')
                        imgs, dados = ler_laudo_ra_pdf(raw)
                    except Exception as ex:
                        laudos.append({'arquivo': nome, 'erro': str(ex)[:160]})
                        continue
                    laudos.append({'arquivo': nome, 'assunto': e['subject'],
                                   'data': e.get('data', ''), 'amostrador': cod_pdf,
                                   'paginas': imgs, 'dadosExtraidos': dados})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'erro': str(e)}), 500
    guardados = _guardar_resultados_lidos(
        [l['dadosExtraidos'] for l in laudos if l.get('dadosExtraidos')],
        origem=f'caixa do lab · RA {ra_alvo or cod_alvo}')
    return jsonify({'laudos': laudos, 'total': len(laudos), 'truncado': truncado,
                    'limite': MAX_PDFS, 'guardados': guardados,
                    # como a busca foi resolvida — 'banco' é o caminho rápido
                    'busca': ('ra' if ra_alvo else 'banco' if ras_conhecidos else 'varredura'),
                    'ras_conhecidos': sorted(ras_conhecidos),
                    'varredura_limitada': varredura_limitada, 'varridos': varridos})


# ── API: Parse chain of custody Excel (Uniscientific format) ─────────
@app.route('/api/parse_cadeia', methods=['POST'])
@login_required
def api_parse_cadeia():
    try:
        import openpyxl
        import unicodedata as _ud
    except ImportError:
        return jsonify({'erro': 'openpyxl nao instalado'}), 500
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'erro': 'Nenhum arquivo'}), 400
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)

        # Prefer "Dados Agentes" sheet (Uniscientific format)
        ws = wb['Dados Agentes'] if 'Dados Agentes' in wb.sheetnames else wb.active

        all_rows = list(ws.iter_rows(values_only=True))

        def _norm(s):
            """Remove acentos e converte para ASCII maiúsculo para comparação."""
            return _ud.normalize('NFD', str(s)).encode('ascii', 'ignore').decode('ascii').upper()

        # Find header row: normalized text contains FUNCIONARIO or FUNCAO
        header_idx = None
        header_row = []
        for i, row in enumerate(all_rows):
            row_txt = ' '.join(_norm(c) for c in row if c)
            if 'FUNCIONARIO' in row_txt or 'FUNCAO' in row_txt:
                header_idx = i
                header_row = [str(c).strip() if c else '' for c in row]
                break

        if header_idx is None:
            return jsonify({'erro': 'Cabecalho nao encontrado (nenhuma coluna FUNCIONARIO/FUNCAO)'}), 400

        # find_col: compara com texto normalizado (sem acentos)
        header_norm = [_norm(h) for h in header_row]

        def find_col(keywords):
            for j, hn in enumerate(header_norm):
                for kw in keywords:
                    if _norm(kw) in hn:
                        return j
            return None

        # Número do amostrador: a coluna "NÚMERO DO AMOSTRADOR" (código do lab, ex.
        # EC81053A — é o que casa com o RA), NÃO a "(CLIENTE)" (numeração interna do
        # cliente, que costuma vir vazia). Fallback p/ a (CLIENTE) se a do lab faltar.
        col_id = col_id_cli = None
        for j, hn in enumerate(header_norm):
            if 'NUMERO DO AMOSTRADOR' in hn:
                if 'CLIENTE' in hn:
                    col_id_cli = j
                elif col_id is None:
                    col_id = j
        if col_id is None:
            col_id = col_id_cli
        col_data    = find_col(['DATA AMOSTRAGEM', 'DATA DE AMOSTRAGEM'])
        col_nome    = find_col(['NOME DO FUNCIONARIO', 'FUNCIONARIO'])
        col_funcao  = find_col(['FUNCAO', 'CARGO'])
        col_setor   = find_col(['SETOR'])
        col_vazao   = find_col(['VAZAO MEDIA', 'VAZAO'])
        col_volume  = find_col(['VOLUME AMOSTRADO', 'VOLUME'])
        col_inicio  = find_col(['INICIO DA AMOSTRAGEM'])
        col_termino = find_col(['TERMINO DA AMOSTRAGEM'])
        agente_cols = [j for j, hn in enumerate(header_norm) if 'AGENTE' in hn]

        # col_funcao não pode ser igual a col_nome — se for, refinar busca
        if col_funcao is not None and col_funcao == col_nome:
            # Busca pela coluna que seja EXATAMENTE "FUNCAO" (sem "FUNCIONARIO" no meio)
            for j, hn in enumerate(header_norm):
                stripped = hn.replace('(*)', '').strip()
                if stripped == 'FUNCAO' or stripped == 'CARGO':
                    col_funcao = j
                    break

        def _cv(row, col):
            if col is None or col >= len(row):
                return None
            return row[col]

        def _str(v):
            if v is None:
                return ''
            if hasattr(v, 'strftime'):
                try:
                    return v.strftime('%d/%m/%Y')
                except:
                    return str(v)
            s = str(v).strip()
            return '' if s == 'None' else s

        def _time_str(v):
            if v is None:
                return ''
            if hasattr(v, 'hour'):
                return '{:02d}:{:02d}'.format(v.hour, v.minute)
            parts = str(v).strip().split(':')
            return '{}:{}'.format(parts[0], parts[1]) if len(parts) >= 2 else str(v)

        def _fmt_float(v, dec=4):
            if v is None:
                return ''
            try:
                return ('{:.' + str(dec) + 'f}').format(float(str(v).replace(',', '.'))).replace('.', ',')
            except:
                return str(v).strip()

        def _tempo_min(ini, fim):
            try:
                if hasattr(ini, 'hour') and hasattr(fim, 'hour'):
                    from datetime import datetime as _dt2, date
                    d = date.today()
                    diff = _dt2.combine(d, fim) - _dt2.combine(d, ini)
                    return str(int(diff.total_seconds() / 60))
                parts_i = str(ini).split(':')
                parts_f = str(fim).split(':')
                mins = (int(parts_f[0]) * 60 + int(parts_f[1])) - (int(parts_i[0]) * 60 + int(parts_i[1]))
                return str(mins)
            except:
                return ''

        avaliacoes = []
        for row in all_rows[header_idx + 1:]:
            if not any(row):
                continue
            nome = _str(_cv(row, col_nome))
            if not nome:
                continue

            agentes = [str(_cv(row, ac)).strip() for ac in agente_cols
                       if _cv(row, ac) and str(_cv(row, ac)).strip() not in ('', 'None')]

            ini_v  = _cv(row, col_inicio)
            fim_v  = _cv(row, col_termino)
            vaz_v  = _cv(row, col_vazao)
            vol_v  = _cv(row, col_volume)
            vaz_fmt = _fmt_float(vaz_v, 4)

            # Volume in Litros (sheet already in L)
            vol_fmt = _fmt_float(vol_v, 3)

            # Date
            data_v = _cv(row, col_data)
            data_s = data_v.strftime('%d/%m/%Y') if hasattr(data_v, 'strftime') else _str(data_v)

            ev = {
                'filtroNumero':  _str(_cv(row, col_id)) or (_str(_cv(row, col_id_cli)) if col_id_cli is not None else ''),
                'dataColeta':    data_s,
                'trabalhador':   nome,
                'cargo':         _str(_cv(row, col_funcao)),
                'setor':         _str(_cv(row, col_setor)),
                'vazaoInicial':  vaz_fmt,
                'vazaoFinal':    vaz_fmt,
                'volume':        vol_fmt,
                'tempoColeta':   _tempo_min(ini_v, fim_v),
                'inicioColeta':  _time_str(ini_v),
                'terminoColeta': _time_str(fim_v),
                'agente':        agentes[0] if agentes else '',
            }
            avaliacoes.append(ev)

        return jsonify({'avaliacoes': avaliacoes})
    except Exception as e:
        import traceback
        traceback.print_exc(); return jsonify({'erro': str(e)}), 500


# ════════════════════════════════════════════════════════════════════
# GERADOR DE LAUDO DE RUÍDO
# ════════════════════════════════════════════════════════════════════

# Fallback legado — técnicos agora vêm do DB via /api/tecnicos
_TECNICO_FALLBACK = {'nome': 'TÉCNICO RESPONSÁVEL', 'mte': ''}

TPL_RUIDO = os.path.join(BASE_DIR, 'template_ruido.docx')


def _r_esc(s):
    """Escapa XML básico para inserção no documento."""
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')


# Corpo do laudo de ruído é Verdana (383 usos no template); runs gerados sem
# rFonts herdariam Calibri (fonte minor do tema) e sairiam com fonte diferente
# do resto do documento. Força Verdana em todo run injetado.
_RF_RUIDO = '<w:rFonts w:ascii="Verdana" w:hAnsi="Verdana"/>'


def _r_p(text, bold=False, size=18, color=None, center=False, fill=None):
    """Parágrafo simples de uma corrida."""
    ppr = '<w:pPr>'
    if center:
        ppr += '<w:jc w:val="center"/>'
    ppr += '</w:pPr>' if ppr != '<w:pPr>' else ''
    if ppr == '<w:pPr></w:pPr>':
        ppr = '<w:pPr/>'
    rpr = f'<w:rPr>{_RF_RUIDO}{"<w:b/>" if bold else ""}<w:sz w:val="{size}"/><w:szCs w:val="{size}"/>'
    if color:
        rpr += f'<w:color w:val="{color}"/>'
    rpr += '</w:rPr>'
    return f'<w:p>{ppr}<w:r>{rpr}<w:t xml:space="preserve">{_r_esc(text)}</w:t></w:r></w:p>'


def _r_row2(label, value, lw=3500, vw=6967, lbold=True, size=18):
    """Linha 2-colunas: label | valor."""
    borders = ('<w:tcBorders>'
               '<w:top w:val="single" w:sz="4" w:color="AAAAAA"/>'
               '<w:bottom w:val="single" w:sz="4" w:color="AAAAAA"/>'
               '<w:left w:val="single" w:sz="4" w:color="AAAAAA"/>'
               '<w:right w:val="single" w:sz="4" w:color="AAAAAA"/>'
               '</w:tcBorders>')
    lpr = f'<w:tcPr><w:tcW w:w="{lw}" w:type="dxa"/><w:shd w:val="clear" w:fill="EEEEEE" w:color="auto"/>{borders}</w:tcPr>'
    vpr = f'<w:tcPr><w:tcW w:w="{vw}" w:type="dxa"/><w:shd w:val="clear" w:fill="FFFFFF" w:color="auto"/>{borders}</w:tcPr>'
    lp  = f'<w:p><w:r><w:rPr>{_RF_RUIDO}{"<w:b/>" if lbold else ""}<w:sz w:val="{size}"/><w:szCs w:val="{size}"/></w:rPr><w:t>{_r_esc(label)}</w:t></w:r></w:p>'
    vp  = f'<w:p><w:r><w:rPr>{_RF_RUIDO}<w:sz w:val="{size}"/><w:szCs w:val="{size}"/></w:rPr><w:t xml:space="preserve">{_r_esc(value)}</w:t></w:r></w:p>'
    return f'<w:tr><w:tc>{lpr}{lp}</w:tc><w:tc>{vpr}{vp}</w:tc></w:tr>'


def _r_row_header(text, fill='1F497D', color='FFFFFF', span=2):
    """Linha de cabeçalho que ocupa toda a largura."""
    borders = ('<w:tcBorders>'
               '<w:top w:val="single" w:sz="6" w:color="000000"/>'
               '<w:bottom w:val="single" w:sz="6" w:color="000000"/>'
               '</w:tcBorders>')
    tcp = f'<w:tcPr><w:gridSpan w:val="{span}"/><w:shd w:val="clear" w:fill="{fill}" w:color="auto"/>{borders}</w:tcPr>'
    rpr = f'<w:rPr>{_RF_RUIDO}<w:b/><w:color w:val="{color}"/><w:sz w:val="20"/><w:szCs w:val="20"/></w:rPr>'
    return f'<w:tr><w:tc>{tcp}<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r>{rpr}<w:t>{_r_esc(text)}</w:t></w:r></w:p></w:tc></w:tr>'


def _r_row_sub(text, span=2, fill='D9D9D9'):
    """Sub-cabeçalho de seção dentro da tabela (Q3, Q5, etc.)."""
    tcp = f'<w:tcPr><w:gridSpan w:val="{span}"/><w:shd w:val="clear" w:fill="{fill}" w:color="auto"/></w:tcPr>'
    rpr = f'<w:rPr>{_RF_RUIDO}<w:b/><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
    return f'<w:tr><w:tc>{tcp}<w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r>{rpr}<w:t>{_r_esc(text)}</w:t></w:r></w:p></w:tc></w:tr>'


def _r_img_para(rid, iid, cx, cy):
    """Parágrafo com imagem inline centralizada."""
    return (
        '<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
        f'<w:r><w:rPr/><w:drawing>'
        f'<wp:inline xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing">'
        f'<wp:extent cx="{cx}" cy="{cy}"/>'
        f'<wp:docPr id="{iid}" name="img{iid}"/>'
        f'<a:graphic xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f'<a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:pic xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture">'
        f'<pic:nvPicPr><pic:cNvPr id="{iid}" name="img{iid}"/><pic:cNvPicPr/></pic:nvPicPr>'
        f'<pic:blipFill><a:blip r:embed="{rid}" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/>'
        f'<a:stretch><a:fillRect/></a:stretch></pic:blipFill>'
        f'<pic:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></pic:spPr>'
        f'</pic:pic></a:graphicData></a:graphic>'
        f'</wp:inline></w:drawing></w:r></w:p>'
    )


def _r_page_break():
    return '<w:p><w:r><w:br w:type="page"/></w:r></w:p>'


def _build_ruido_aval(av, idx, img_rids):
    """
    Constrói o bloco XML de uma avaliação (RESULTADOS + QUADRO RESUMO).
    img_rids = {'tabela': rid, 'histograma': rid}
    """
    n = idx + 1
    cargo = av.get('cargo','')
    setor = av.get('setor','')
    trabalhador = av.get('trabalhador','')

    # ── Título do bloco de RESULTADOS ────────────────────────────────
    tbl_tbl = (
        '<w:tbl>'
        '<w:tblPr>'
        '<w:tblW w:w="10467" w:type="dxa"/>'
        '<w:shd w:val="clear" w:fill="1F497D" w:color="auto"/>'
        '</w:tblPr>'
        '<w:tblGrid><w:gridCol w:w="10467"/></w:tblGrid>'
        '<w:tr><w:tc>'
        '<w:tcPr><w:tcW w:w="10467" w:type="dxa"/>'
        '<w:shd w:val="clear" w:fill="1F497D" w:color="auto"/></w:tcPr>'
        '<w:p><w:pPr><w:pStyle w:val="hil1"/></w:pPr>'
        f'<w:r><w:t>RESULTADOS – AVALIAÇÃO {n:02d}</w:t></w:r>'
        '</w:p></w:tc></w:tr></w:tbl>'
    )

    # ── Imagens do dosímetro ─────────────────────────────────────────
    W = 5486400   # ~15.2cm em EMU
    H_tab = 4200000
    H_hist = 5400000

    imgs_xml = ''
    if img_rids.get('tabela'):
        imgs_xml += _r_img_para(img_rids['tabela'], n*10+1, W, H_tab)
    if img_rids.get('histograma'):
        imgs_xml += _r_img_para(img_rids['histograma'], n*10+2, W, H_hist)

    # ── Tabela de dados ──────────────────────────────────────────────
    # OBS: o título "QUADRO RESUMO" é a 1ª linha DESTA tabela (não uma tabela
    # separada). Duas <w:tbl> adjacentes sem parágrafo entre elas se FUNDEM no
    # Word — e com grades diferentes (1 col do título vs 2 col dos dados) a
    # tabela saía toda desalinhada. Dobrar o título na linha-cabeçalho elimina
    # a fusão sem depender de parágrafo-espaçador (que o Word pode colapsar).
    TOTAL = 10467
    LW, VW = 3500, 6967

    tbl_grid = f'<w:tblGrid><w:gridCol w:w="{LW}"/><w:gridCol w:w="{VW}"/></w:tblGrid>'
    tbl_pr   = (f'<w:tblPr>'
                f'<w:tblStyle w:val="Tabelacomgrade"/>'
                f'<w:tblW w:w="{TOTAL}" w:type="dxa"/>'
                f'<w:tblLayout w:type="fixed"/>'
                f'</w:tblPr>')

    rows = _r_row_header(f'QUADRO RESUMO – AVALIAÇÃO {n:02d} — {cargo.upper()} / {setor.upper()}')
    rows += _r_row2('Setor',                   setor)
    rows += _r_row2('Cargo',                   cargo)
    rows += _r_row2('Funcionário(a)',           trabalhador)
    rows += _r_row2('Data da Avaliação',        av.get('dataColeta',''))
    rows += _r_row2('Horário Início / Fim',     f"{av.get('horaInicio','')} / {av.get('horaFim','')}")
    rows += _r_row2('Jornada de Trabalho',      av.get('jornada',''))
    rows += _r_row2('N° Série Dosímetro',       av.get('serie',''))
    rows += _r_row2('Fonte(s) Geradora(s)',     av.get('fontes',''))
    rows += _r_row2('Descrição das Atividades', av.get('atividades',''))
    rows += _r_row2('Medidas de Controle Col.', av.get('controleColetivo', 'N.A.'))
    rows += _r_row2('EPI Utilizado',            av.get('epi', 'Protetor Auditivo'))
    # O bloco Q = 3 dB (NHO-01 / Fundacentro) NÃO entra no documento — decisão do
    # Bernardo em 27/08/2026: o laudo mostra só Trabalhista (Q=5) e Previdenciária
    # (Q=5* / NEN INSS). O dado Q3 continua sendo lido do dosímetro, mas não sai.
    rows += _r_row_sub('Q = 5 dB / Dosimetria NR-15 — Legislação Trabalhista')
    rows += _r_row2('TWA',                     f"{av.get('twaQ5','')} dB(A)")
    rows += _r_row2('LAVG',                    f"{av.get('lavgQ5','')} dB(A)")
    rows += _r_row2('DOSE',                    f"{av.get('doseQ5','')} %")
    rows += _r_row_sub('Q = 5* dB / NEN INSS — Legislação Previdenciária')
    rows += _r_row2('NE',                      f"{av.get('neQ5','')} dB")
    rows += _r_row2('NEN',                     f"{av.get('nenQ5','')} dB")

    # Conclusão automática — DUPLA: uma para a Legislação Trabalhista (NR-15, Q=5)
    # e outra para a Previdenciária (NEN INSS, Q=5*). Sem valor válido a conclusão
    # daquela legislação é omitida — nunca conclui "não ultrapassou" sem dado
    # (antes o default 0 dava laudo verde).
    _CORES = {3: 'FFD0D0', 2: 'FFFACC', 1: 'D0FFD0', 0: 'FFFFFF'}

    def _num_ruido(v):
        try:
            return float(str(v).replace(',', '.').replace(' dB(A)', '').strip())
        except (TypeError, ValueError):
            return None

    def _conclui_ruido(valor, prefixo):
        if valor is None:
            return None
        if valor >= 85:
            return (f"{prefixo}: a exposição de {cargo} ULTRAPASSOU o limite de tolerância "
                    f"de 85,0 dB(A) — necessário adotar medidas de controle coletivas e/ou "
                    f"individuais.", 3)
        if valor >= 80:
            return (f"{prefixo}: a exposição de {cargo} está acima do NÍVEL DE AÇÃO de "
                    f"80,0 dB(A) — necessário adotar medidas de controle para reduzir a "
                    f"exposição.", 2)
        return (f"{prefixo}: a exposição de {cargo} NÃO ultrapassou o limite de tolerância "
                f"de 85,0 dB(A).", 1)

    # Trabalhista usa Q=5 — TWA, com fallback p/ LAVG Q5.
    # Previdenciária usa o NEN do bloco INSS (Q=5*) — nunca o Q=3.
    _v_nr15 = _num_ruido(av.get('twaQ5'))
    if _v_nr15 is None:
        _v_nr15 = _num_ruido(av.get('lavgQ5'))
    conclusoes = [c for c in (
        _conclui_ruido(_v_nr15,
                       'Conforme a Legislação Trabalhista — NR-15, Anexo 1 (Q=5 dB)'),
        _conclui_ruido(_num_ruido(av.get('nenQ5')),
                       'Conforme a Legislação Previdenciária — NEN (Q=5* dB)'),
    ) if c]
    if not conclusoes:
        manual = av.get('conclusao', '')
        conclusoes = [(manual, 0)] if manual else []

    header_rank = max((rk for _t, rk in conclusoes), default=0)
    rows += _r_row_header('CONCLUSÃO', fill=_CORES[header_rank], color='000000')
    for texto, _rk in conclusoes:
        rows += (f'<w:tr><w:tc>'
                 f'<w:tcPr><w:gridSpan w:val="2"/>'
                 f'<w:shd w:val="clear" w:fill="FFFFFF" w:color="auto"/></w:tcPr>'
                 f'<w:p><w:r><w:rPr>{_RF_RUIDO}<w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
                 f'<w:t xml:space="preserve">{_r_esc(texto)}</w:t>'
                 f'</w:r></w:p></w:tc></w:tr>')

    tbl_dados = f'<w:tbl>{tbl_pr}{tbl_grid}{rows}</w:tbl>'

    return tbl_tbl + imgs_xml + _r_page_break() + tbl_dados + _r_page_break()


def _find_section_tbl(xml, heading_text):
    """Encontra tabela que contém hil1 + heading_text, retorna (start, end)."""
    import re as _re
    for m in _re.finditer(r'<w:tbl[ >]', xml):
        ts = m.start()
        te_m = xml.find('</w:tbl>', ts)
        if te_m < 0:
            continue
        te = te_m + len('</w:tbl>')
        chunk = xml[ts:te]
        if 'hil1' in chunk and heading_text in chunk:
            return ts, te
    return None, None


def gerar_ruido_bytes(d):
    emp       = d.get('empresa', {})
    avals     = d.get('avaliacoes', [])
    tecnico   = d.get('tecnico', 'kelly')
    data_laud = d.get('dataLaudo', datetime.now().strftime('%d/%m/%Y'))

    # Aceita {nome, mte} direto (usuário do sistema) ou chave antiga
    if isinstance(tecnico, dict):
        tec = {'nome': tecnico.get('nome', '').upper(), 'mte': tecnico.get('mte', '')}
    else:
        tec = _TECNICO_FALLBACK

    with open(TPL_RUIDO, 'rb') as f:
        tpl_bytes = f.read()

    zin = zipfile.ZipFile(io.BytesIO(tpl_bytes))
    zout_buf = io.BytesIO()
    zout = zipfile.ZipFile(zout_buf, 'w', zipfile.ZIP_DEFLATED)

    extra_media = {}
    extra_rels  = {}
    next_rid    = [30]

    def _new_rid():
        r = f'rId{next_rid[0]}'
        next_rid[0] += 1
        return r

    def _add_img_b64(b64str):
        rid = _new_rid()
        if ',' in b64str:
            b64str = b64str.split(',', 1)[1]
        data = base64.b64decode(b64str)
        name = f'image_ruido_{next_rid[0]-1}.jpeg'
        extra_media[f'word/media/{name}'] = data
        extra_rels[rid] = f'../media/{name}'
        return rid

    # ── Logo (substitui image1.png pelo logo da empresa, ou em branco) ─
    BLANK_PNG = base64.b64decode(
        'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAAC0lEQVQI12NgAAIABQ'
        'AABjkB6QAAAABJRU5ErkJggg==')
    logo_b64 = emp.get('logo', '')
    if logo_b64:
        _hdr, _dat = logo_b64.split(',', 1) if ',' in logo_b64 else ('', logo_b64)
        try:
            logo_bytes_ruido = base64.b64decode(_dat)
        except Exception:
            logo_bytes_ruido = BLANK_PNG
    else:
        logo_bytes_ruido = BLANK_PNG

    # Copia arquivos originais do template (substituindo image1.png pelo logo)
    doc_xml  = None
    rels_xml = None
    ct_xml   = None

    for item in zin.namelist():
        data = zin.read(item)
        if item == 'word/document.xml':
            doc_xml = data.decode('utf-8')
        elif item == 'word/_rels/document.xml.rels':
            rels_xml = data.decode('utf-8')
        elif item == '[Content_Types].xml':
            ct_xml = data.decode('utf-8')
        elif item == 'word/media/image1.png':
            zout.writestr(item, logo_bytes_ruido)
        else:
            zout.writestr(item, data)

    # ── Substitui dados da empresa ────────────────────────────────────
    razao = emp.get('razaoSocial', '')
    # Remove sufixos comuns para casar com runs quebrados pelo Word
    razao_curta = re.sub(r'\s+(LTDA|S\.?A\.?|EIRELI|ME|EPP|S/A)\s*$', '', razao, flags=re.I).strip()
    razao_upper = razao.upper()
    razao_curta_upper = razao_curta.upper()
    replacements = {
        # Variações com LTDA (texto completo)
        'Helisul Taxi Aéreo LTDA':       razao,
        'HELISUL TAXI AERIO LTDA':       razao_upper,
        # Variações sem LTDA (texto quebrado em runs pelo Word)
        # OBS: no template, o run seguinte é " LTDA" (com espaço à esquerda) —
        # não acrescentar espaço aqui, senão a razão social sai com espaço duplo
        'Helisul Taxi Aéreo ':           razao_curta,
        'HELISUL TAXI AERIO ':           razao_curta_upper + ' ',
        'HELISUL TAXI AERIO':            razao_curta_upper,
        # Alt-text de imagens (descr=)
        'Helisul Taxi Aereo':            razao_curta,
        # Carta de apresentação (placeholder literal do template)
        'NOME DA EMPRESA':               (emp.get('razaoSocial') or emp.get('nomeFantasia') or '').upper(),
        # Outros campos
        'Rua Gardênia N.º 165':          emp.get('endereco', ''),
        '11.483.174/0004-11':            emp.get('cnpj', ''),
        '32150-190':                     emp.get('cep', ''),
        'Contagem':                      emp.get('cidade', ''),
        'Chácara Boa Vista':             emp.get('bairro', ''),
        '56.11-2':                       emp.get('cnae', ''),
        'Restaurantes e outros estabelecimentos de serviços de alimentação e bebida': emp.get('descricaoCnae', ''),
        'Wilde José Silva de Abreu':     emp.get('responsavel', ''),
        '31 3213-3089':                  emp.get('telefone', ''),
        'bionatural@ymail.com':          emp.get('email', ''),
    }
    if emp.get('grauRisco'):
        doc_xml = doc_xml.replace('>2<', f'>{emp["grauRisco"]}<', 1)
    for old, new in replacements.items():
        if new:
            doc_xml = doc_xml.replace(_r_esc(old), _r_esc(new))

    # ── Substitui responsável técnica ─────────────────────────────────
    # Nome completo do template PRIMEIRO (senão "Sued Iagor" vira o nome novo
    # e o resto do sobrenome do template fica colado: "Matheus Costa de
    # Mimrop Rodrigues da Silva")
    _tec_t, _tec_u = tec['nome'].title(), tec['nome'].upper()
    for _full in ('Sued Iagor de Mimrop Rodrigues da Silva',
                  'Sued Iagor Rodrigues da Silva'):
        doc_xml = doc_xml.replace(_full, _tec_t)
        doc_xml = doc_xml.replace(_full.upper(), _tec_u)
    doc_xml = doc_xml.replace('Sued Iagor', _tec_t)
    doc_xml = doc_xml.replace('SUED IAGOR', _tec_u)
    # Limpa resto de sobrenome do template que tenha sobrado colado ao nome
    for _sobra in (' de Mimrop Rodrigues da Silva', ' Rodrigues da Silva',
                   ' DE MIMROP RODRIGUES DA SILVA', ' RODRIGUES DA SILVA'):
        doc_xml = doc_xml.replace(_tec_t + _sobra, _tec_t)
        doc_xml = doc_xml.replace(_tec_u + _sobra, _tec_u)
    # MTE: no XML o template pode quebrar "0065338-MG" em runs ("0065338- MG");
    # substitui o número e a UF separadamente para sobreviver às quebras.
    _mte = (tec.get('mte') or '').strip()
    if _mte:
        _mte_num = _mte.split('-')[0].strip()
        _mte_uf  = _mte.split('-')[1].strip() if '-' in _mte else 'MG'
        doc_xml = doc_xml.replace('0065338-MG', _mte)
        doc_xml = doc_xml.replace('0065338- MG', f'{_mte_num}- {_mte_uf}')
        doc_xml = doc_xml.replace('0065338', _mte_num)

    # ── Monta seção de avaliações ─────────────────────────────────────
    aval_xml = ''
    for i, av in enumerate(avals):
        img_rids = {}
        if av.get('tabelaImg'):
            img_rids['tabela'] = _add_img_b64(av['tabelaImg'])
        if av.get('histogramaImg'):
            img_rids['histograma'] = _add_img_b64(av['histogramaImg'])
        aval_xml += _build_ruido_aval(av, i, img_rids)

    # ── Seção de certificados ─────────────────────────────────────────
    cert_imgs_xml = ''
    for av in avals:
        for ci, c in enumerate(av.get('certImgs', [])):
            crid = _add_img_b64(c)
            cert_imgs_xml += _r_img_para(crid, next_rid[0], 4800000, 6900000)
            cert_imgs_xml += _r_page_break()

    # ── Seção de assinaturas ──────────────────────────────────────────
    sig_avaliado = d.get('sig_avaliado')
    sig_empresa  = d.get('sig_empresa')
    sig_xml = ''
    if sig_avaliado or sig_empresa:
        def _sig_cell(b64, label, tec_nome=''):
            inner = ''
            if b64:
                rid = _add_img_b64(b64)
                inner += _r_img_para(rid, next_rid[0], 2400000, 675000)
            # linha separadora + nome
            inner += ('<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                      f'<w:r><w:rPr><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
                      f'<w:t>{"_" * 36}</w:t></w:r></w:p>')
            if tec_nome:
                inner += ('<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                          f'<w:r><w:rPr><w:b/><w:sz w:val="18"/><w:szCs w:val="18"/></w:rPr>'
                          f'<w:t xml:space="preserve">{_r_esc(tec_nome)}</w:t></w:r></w:p>')
            inner += ('<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                      f'<w:r><w:rPr><w:sz w:val="16"/><w:szCs w:val="16"/></w:rPr>'
                      f'<w:t xml:space="preserve">{_r_esc(label)}</w:t></w:r></w:p>')
            tcp = ('<w:tcPr><w:tcW w:w="4676" w:type="dxa"/>'
                   '<w:tcBorders><w:top w:val="none"/><w:left w:val="none"/>'
                   '<w:bottom w:val="none"/><w:right w:val="none"/></w:tcBorders>'
                   '<w:vAlign w:val="bottom"/></w:tcPr>')
            return f'<w:tc>{tcp}{inner}</w:tc>'

        tec_nome = tec['nome'].title() if tec.get('nome') else ''
        left_cell  = _sig_cell(sig_avaliado, 'Avaliado')
        right_cell = _sig_cell(sig_empresa, 'Responsável Técnico', tec_nome)

        # Parágrafo de título
        title_p = ('<w:p><w:pPr><w:jc w:val="center"/></w:pPr>'
                   '<w:r><w:rPr><w:b/><w:sz w:val="20"/><w:szCs w:val="20"/></w:rPr>'
                   '<w:t>ASSINATURAS</w:t></w:r></w:p>')
        sig_tbl = (
            '<w:tbl>'
            '<w:tblPr><w:tblW w:w="9353" w:type="dxa"/>'
            '<w:tblBorders><w:top w:val="none"/><w:left w:val="none"/>'
            '<w:bottom w:val="none"/><w:right w:val="none"/><w:insideH w:val="none"/>'
            '<w:insideV w:val="none"/></w:tblBorders></w:tblPr>'
            '<w:tblGrid><w:gridCol w:w="4676"/><w:gridCol w:w="4677"/></w:tblGrid>'
            f'<w:tr>{left_cell}{right_cell}</w:tr>'
            '</w:tbl>'
        )
        sig_xml = _r_page_break() + title_p + sig_tbl

    # ── Localiza seções no template XML e substitui ────────────────────
    res_ts,  res_te  = _find_section_tbl(doc_xml, 'RESULTADOS')
    cert_ts, cert_te = _find_section_tbl(doc_xml, 'CERTIFICADO DE CALIBRA')

    if res_ts is not None and cert_ts is not None:
        # Tudo entre fim da seção RESULTADOS (incluso) e início de CERTIFICADO
        # é substituído pelas avaliações geradas
        doc_xml = (doc_xml[:res_ts] +
                   aval_xml +
                   doc_xml[cert_ts:cert_te] +
                   cert_imgs_xml +
                   sig_xml +
                   doc_xml[cert_te:])
    else:
        # Fallback: appende antes do </w:body>
        doc_xml = doc_xml.replace('</w:body>', aval_xml + cert_imgs_xml + sig_xml + '</w:body>')

    # ── QUADRO RESUMO DAS AVALIAÇÕES (geral) ──────────────────────────
    # O template tem uma linha-exemplo ("Coordenador (a) de base" / 80,5 dB)
    # que servia só de molde. Substitui por uma linha por avaliação.
    _qi = doc_xml.find('Coordenador (a) de base')
    if _qi != -1:
        _qtrs = max(doc_xml.rfind('<w:tr ', 0, _qi), doc_xml.rfind('<w:tr>', 0, _qi))
        _qtre = doc_xml.find('</w:tr>', _qi) + len('</w:tr>')

        def _qg_color(v):
            try:
                f = float(str(v).replace(',', '.'))
            except Exception:
                return '000000'
            if f >= 85: return 'FF0000'
            if f >= 80: return 'FFC000'
            return '00B050'

        def _qg_cell(w, text, bullet=None):
            runs = ''
            if bullet:
                runs += ('<w:r><w:rPr><w:rFonts w:ascii="Verdana" w:hAnsi="Verdana"/>'
                         f'<w:color w:val="{bullet}"/><w:sz w:val="16"/><w:szCs w:val="16"/></w:rPr>'
                         '<w:t xml:space="preserve">● </w:t></w:r>')
            runs += ('<w:r><w:rPr><w:rFonts w:ascii="Verdana" w:hAnsi="Verdana"/><w:b/>'
                     '<w:sz w:val="16"/><w:szCs w:val="16"/></w:rPr>'
                     f'<w:t xml:space="preserve">{_r_esc(text)}</w:t></w:r>')
            return (f'<w:tc><w:tcPr><w:tcW w:w="{w}" w:type="dxa"/>'
                    '<w:shd w:val="clear" w:color="auto" w:fill="FFFFFF"/>'
                    '<w:vAlign w:val="center"/></w:tcPr>'
                    f'<w:p><w:pPr><w:jc w:val="center"/></w:pPr>{runs}</w:p></w:tc>')

        _qrows = ''
        for av in avals:
            # Sem fallback para Q3: o valor de q=3 é de outro critério e sairia
            # rotulado como Trabalhista/Previdenciária, que é justamente o que o
            # Bernardo mandou tirar. Sem o dado de Q=5 a célula fica '-'.
            _lavg = str(av.get('lavgQ5') or '').strip()
            _nen  = str(av.get('nenQ5') or '').strip()
            _qrows += ('<w:tr><w:trPr><w:trHeight w:val="362"/></w:trPr>'
                       + _qg_cell(2195, av.get('setor', '') or '-')
                       + _qg_cell(2195, av.get('cargo', ''))
                       + _qg_cell(2023, f'{_lavg} dB(A)' if _lavg else '-',
                                  _qg_color(_lavg) if _lavg else None)
                       + _qg_cell(2024, f'{_nen} dB(A)' if _nen else '-',
                                  _qg_color(_nen) if _nen else None)
                       + '</w:tr>')
        # Mesmo sem avaliações, remove a linha-fantasma do template
        doc_xml = doc_xml[:_qtrs] + _qrows + doc_xml[_qtre:]

    # ── Adiciona relacionamentos das imagens geradas ───────────────────
    for rid, target in extra_rels.items():
        rels_xml = rels_xml.replace(
            '</Relationships>',
            f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{target}"/></Relationships>'
        )

    # Adiciona Content-Type para JPEG se não existir
    if 'image/jpeg' not in ct_xml:
        ct_xml = ct_xml.replace(
            '</Types>',
            '<Default Extension="jpeg" ContentType="image/jpeg"/></Types>'
        )

    # Fix IDs duplicados
    _idc2 = [1]
    def _nwid2(m):
        _idc2[0] += 1
        return f'w:id="{_idc2[0]}"'
    doc_xml = re.sub(r'w:id="\d+"', _nwid2, doc_xml)

    zout.writestr('word/document.xml', doc_xml.encode('utf-8'))
    zout.writestr('word/_rels/document.xml.rels', rels_xml.encode('utf-8'))
    zout.writestr('[Content_Types].xml', ct_xml.encode('utf-8'))
    for path, data in extra_media.items():
        zout.writestr(path, data)

    zout.close()
    return zout_buf.getvalue()


# ── API: Parse dosimeter PDF ──────────────────────────────────────────
@app.route('/api/parse_dosimetro', methods=['POST'])
@login_required
def api_parse_dosimetro():
    try:
        import fitz
    except ImportError:
        return jsonify({'erro': 'pymupdf nao instalado'}), 500
    try:
        import re as _re
        f = request.files.get('file')
        if not f:
            return jsonify({'erro': 'Nenhum arquivo'}), 400
        raw = f.read()
        doc = fitz.open(stream=raw, filetype='pdf')

        page0 = doc[0]
        text  = page0.get_text()

        # Tabela: página 0 em resolução media
        pix0 = page0.get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        tabela_img = 'data:image/jpeg;base64,' + base64.b64encode(pix0.tobytes('jpeg')).decode()

        # Histograma: página 1
        histograma_img = ''
        if doc.page_count > 1:
            pix1 = doc[1].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
            histograma_img = 'data:image/jpeg;base64,' + base64.b64encode(pix1.tobytes('jpeg')).decode()

        doc.close()

        def _g(patterns):
            for pat in patterns:
                m = _re.search(pat, text, _re.IGNORECASE | _re.MULTILINE)
                if m:
                    return m.group(1).strip()
            return ''

        serie      = _g([r'nº de série[:\s]+(\w+)'])
        avaliado   = _g([r'Avaliado\(a\):\n([^\n]+)'])
        funcao     = _g([r'Função\s*:\n([^\n]+)'])
        depto      = _g([r'Departamento:\n([^\n]+)'])
        data_med   = _g([r'Data da Medição:\n(\d{2}/\d{2}/\d{4})'])
        hora_ini   = _g([r'Início:\s*([\d:]+)'])
        hora_fim   = _g([r'Final:\s*([\d:]+)'])
        # Jornada aparece depois do primeiro "Slow"
        jornada    = _g([r'Slow\n(\d{1,2}:\d{2})\n'])
        calib_data = _g([r'Calibração do audiodosímetro: Data:\s*(\d{2}/\d{2}/\d{4})'])
        empresa_cli = _g([r'Dados da Avaliada\nEmpresa:\nEndereço:\n([^\n]+)'])
        cnpj_cli   = _g([r'CNPJ:\s*([\d\.\-/]+)'])
        endereco_cli = _g([r'Dados da Avaliada\nEmpresa:\nEndereço:\n[^\n]+\n([^\n]+)'])
        # Tempo de amostragem: após "Final: HH:MM:SS\n"
        tempo_amos = _g([r'Final:\s*[\d:]+\n([\d:]+)\n'])

        # Extrair valores numéricos da tabela
        # Localiza o trecho de valores (antes das labels LAVG, LEQ, etc.)
        col_lbl_pos = text.find('LAVG\n')
        val_section = text[:col_lbl_pos] if col_lbl_pos > 0 else text

        # Todos os pares (número, %) na seção de valores
        all_vals = _re.findall(r'([\d]+,[\d]+)\s*(%?)', val_section[-600:])
        plain = [v for v, p in all_vals if not p]  # dB values
        pcts  = [v for v, p in all_vals if p]       # dose values (%)

        # Mapeamento por posição (ordem do display do dosímetro):
        # [0] LAVG Dos01, [1] LAVG Dos02
        # [2] LEQ Dos01,  [3] LEQ Dos02
        # ...mais adiante: [8] NEN Dos01, [9] NEN Dos02, [10] TWA Dos01, [11] TWA Dos02
        lavg_q3 = plain[0]  if len(plain) > 0  else ''
        lavg_q5 = plain[1]  if len(plain) > 1  else ''
        dose_q3 = pcts[0]   if len(pcts) > 0   else ''
        dose_q5 = pcts[1]   if len(pcts) > 1   else ''
        nen_q3  = plain[8]  if len(plain) > 8  else ''
        twa_q5  = plain[9]  if len(plain) > 9  else ''

        dados = {k: v for k, v in {
            'serie':       serie,
            'trabalhador': avaliado,
            'cargo':       funcao,
            'setor':       depto,
            'dataColeta':  data_med,
            'horaInicio':  hora_ini,
            'horaFim':     hora_fim,
            'jornada':     jornada,
            'calibData':   calib_data,
            'empresaCli':  empresa_cli,
            'cnpjCli':     cnpj_cli,
            'enderecoCli': endereco_cli,
            'tempoAmos':   tempo_amos,
            'lavgQ3':      lavg_q3,
            'nenQ3':       nen_q3,
            'doseQ3':      dose_q3,
            'twaQ5':       twa_q5,
            'lavgQ5':      lavg_q5,
            'doseQ5':      dose_q5,
        }.items() if v}

        return jsonify({'dados': dados, 'tabela': tabela_img, 'histograma': histograma_img})
    except Exception as e:
        import traceback
        traceback.print_exc(); return jsonify({'erro': str(e)}), 500


# ── Rota: Gerar laudo de ruído ────────────────────────────────────────
@app.route('/gerar-ruido', methods=['POST'])
@login_required
def gerar_ruido():
    try:
        d = request.get_json(force=True)
        docx_bytes = gerar_ruido_bytes(d)
        emp   = d.get('empresa', {})
        nome  = emp.get('nomeFantasia') or emp.get('razaoSocial') or 'Empresa'
        fname = re.sub(r'[^\w\s-]', '', nome)[:40].strip().replace(' ', '_')
        fname = f'Laudo_Ruido_{fname}.docx'
        usuario = current_user.nome if current_user.is_authenticated else 'anônimo'
        registrar_evento('laudo_ruido_gerado', f'Laudo Ruído: {nome}',
                         usuario=usuario, ip=request.remote_addr)
        return send_file(
            io.BytesIO(docx_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=fname
        )
    except Exception as e:
        import traceback
        traceback.print_exc(); return jsonify({'erro': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
