# -*- coding: utf-8 -*-
"""Gerador da Cadeia de Custódia (formulário IT02-M do laboratório).

Nasceu do cruzamento entre a cadeia da Destak preenchida à mão (30/07/2026) e o
banco: amostrador, funcionário, função, setor, vazão e horários já batiam — o
técnico redigitava o que o sistema sabe. O que NÃO batia é o que estes testes
travam: volume e agente nunca chegavam ao banco, e o técnico/data divergiam.
"""
import warnings

import pytest

from controle.cadeia_custodia import (
    _fmt_data, _fmt_hora, _minutos, _volume, _agentes_da_linha,
    coletar_dados, gerar_xlsx, nome_arquivo, marcar_despacho, LINHA_1,
)
from controle.db import get_db, init_db

warnings.filterwarnings('ignore')


# ── volume: a conta que a planilha fazia à mão ────────────────────────

def test_volume_sai_da_vazao_e_das_horas():
    # Caso real EC98029A: 2,0202 L/min de 14:45 a 15:45 = 121,212 L.
    assert _volume(2.0202, None, 0, '14:45', '15:45') == 121.212


def test_volume_usa_o_gravado_quando_existe():
    assert _volume(2.0, 60, 999.0, '14:45', '15:45') == 999.0


def test_volume_sem_dado_nao_inventa():
    assert _volume(None, None, None, None, None) is None
    assert _volume(2.0, 0, 0, None, None) is None


def test_minutos_atravessando_meia_noite():
    assert _minutos('23:30', '00:30') == 60
    assert _minutos('08:31', '12:31') == 240


# ── agente: NÃO pode ser chutado ──────────────────────────────────────

def test_agente_so_vem_do_que_foi_gravado():
    assert _agentes_da_linha('Tolueno') == ['Tolueno']
    assert _agentes_da_linha('Tolueno; Xileno') == ['Tolueno', 'Xileno']
    assert _agentes_da_linha('') == []
    assert _agentes_da_linha(None) == []


def test_agente_nunca_e_preenchido_com_a_lista_da_os(monkeypatch):
    """A OS diz o que a empresa contratou no total, não o que cada tubo tem.
    Preencher todos em todas as linhas mandaria o lab analisar N substâncias
    por amostra — custo multiplicado."""
    import controle.cadeia_custodia as cc
    monkeypatch.setattr(cc, '_agentes_sugeridos',
                        lambda conn, did: ['Tolueno', 'Xileno', 'Benzeno'])
    init_db()
    with get_db() as conn:
        conn.execute("DELETE FROM amostradores WHERE id=-100")
        conn.execute(
            "INSERT INTO amostradores (id, codigo, tipo, status, data_medicao, arquivado)"
            " VALUES (-100,'CCTEST01','TCP','laboratorio','2026-07-29',0)")
    try:
        d = coletar_dados([-100], demanda_id=1)
        assert d['linhas'][0]['agentes'] == []            # nada foi chutado
        assert d['agentes_sugeridos'] == ['Tolueno', 'Xileno', 'Benzeno']
        assert any('AGENTE não informado' in a for a in d['avisos'])
    finally:
        with get_db() as conn:
            conn.execute("DELETE FROM amostradores WHERE id=-100")


def test_escolha_do_tecnico_vence():
    init_db()
    with get_db() as conn:
        conn.execute("DELETE FROM amostradores WHERE id=-101")
        conn.execute(
            "INSERT INTO amostradores (id, codigo, tipo, status, data_medicao, arquivado)"
            " VALUES (-101,'CCTEST02','TCP','laboratorio','2026-07-29',0)")
    try:
        d = coletar_dados([-101], agentes_por_codigo={'CCTEST02': ['Acetato de etila']})
        assert d['linhas'][0]['agentes'] == ['Acetato de etila']
        assert not any('AGENTE não informado' in a for a in d['avisos'])
    finally:
        with get_db() as conn:
            conn.execute("DELETE FROM amostradores WHERE id=-101")


# ── despacho: o gancho que faltava para data_envio_lab ────────────────

def test_gerar_carimba_data_envio_lab():
    init_db()
    with get_db() as conn:
        conn.execute("DELETE FROM amostradores WHERE id=-102")
        conn.execute(
            "INSERT INTO amostradores (id, codigo, tipo, status, data_medicao, arquivado)"
            " VALUES (-102,'CCTEST03','TCP','laboratorio','2026-07-29',0)")
    try:
        marcar_despacho([-102], '2026-07-30')
        with get_db() as conn:
            r = conn.execute(
                "SELECT data_envio_lab, dias_validade FROM amostradores WHERE id=-102").fetchone()
        assert str(r['data_envio_lab'])[:10] == '2026-07-30'
        assert r['dias_validade'] == 45
    finally:
        with get_db() as conn:
            conn.execute("DELETE FROM amostradores WHERE id=-102")


# ── xlsx: o layout é do laboratório, não pode ser recriado ────────────

def test_xlsx_preenche_o_template_do_laboratorio():
    import openpyxl, io
    dados = {
        'empresa': {'nome': 'Empresa Teste Ltda', 'cnpj': '00.000.000/0001-00',
                    'cidade': 'Belo Horizonte', 'uf': 'MG'},
        'linhas': [{
            'codigo': 'TCP9999', 'data': _fmt_data('2026-07-29'),
            'funcionario': 'Fulano de Tal', 'funcao': 'Montador', 'setor': 'Produção',
            'tecnico': 'Helbert', 'vazao': 0.201, 'volume': 3.618,
            'hora_ini': _fmt_hora('11:45'), 'hora_fim': _fmt_hora('12:03'),
            'intervalos': '', 'obs': '', 'agentes': ['Tolueno', 'Xileno'],
        }],
    }
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(dados)))
    # As 4 abas do formulário do lab continuam de pé
    assert wb.sheetnames == ['Dados Empresa', 'Dados Agentes',
                             'Informações Adicionais', 'Plan1']
    wa = wb['Dados Agentes']
    assert wa.cell(LINHA_1, 4).value == 'TCP9999'          # D amostrador
    assert wa.cell(LINHA_1, 6).value == 'Fulano de Tal'    # F funcionário
    assert wa.cell(LINHA_1, 8).value == 'Produção'         # H setor
    assert wa.cell(LINHA_1, 10).value == 0.201             # J vazão
    assert wa.cell(LINHA_1, 11).value == 3.618             # K volume
    assert wa.cell(LINHA_1, 17).value == 'Tolueno'         # Q agente 1
    assert wa.cell(LINHA_1, 18).value == 'Xileno'          # R agente 2
    assert wb['Dados Empresa']['D19'].value == 'Empresa Teste Ltda'
    # o catálogo de 468 substâncias do lab tem que sobreviver ao preenchimento
    cat = [wa.cell(r, 1).value for r in range(11, wa.max_row + 1) if wa.cell(r, 1).value]
    assert len(cat) > 400


def test_hora_vai_como_hora_nao_como_texto():
    # O lab importa eletronicamente: hora em texto quebra a importação.
    import openpyxl, io, datetime as dt
    dados = {'empresa': {'nome': 'X'}, 'linhas': [{
        'codigo': 'A1', 'data': _fmt_data('2026-07-29'), 'funcionario': '', 'funcao': '',
        'setor': '', 'tecnico': '', 'vazao': None, 'volume': None,
        'hora_ini': _fmt_hora('08:31'), 'hora_fim': _fmt_hora('12:31'),
        'intervalos': '', 'obs': '', 'agentes': [],
    }]}
    wa = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(dados)))['Dados Agentes']
    assert isinstance(wa.cell(LINHA_1, 12).value, dt.time)
    assert isinstance(wa.cell(LINHA_1, 3).value, (dt.date, dt.datetime))


def test_nome_do_arquivo_segue_o_padrao_da_pasta():
    n = nome_arquivo({'empresa': {'nome': 'Destak Design Soluções em Móveis Ltda'}},
                     '2026-07-30')
    assert n == 'Cadeia de Custodia - Destak Design Soluções em Móveis Ltda - 30.07.2026.xlsx'


def test_nome_do_arquivo_sem_caractere_invalido():
    n = nome_arquivo({'empresa': {'nome': 'A/B: C*D?'}}, '2026-07-30')
    assert not any(c in n for c in '\\/:*?"<>|'.replace('/', ''))


def test_sem_amostrador_nao_estoura():
    d = coletar_dados([])
    assert d['linhas'] == [] and d['avisos']
