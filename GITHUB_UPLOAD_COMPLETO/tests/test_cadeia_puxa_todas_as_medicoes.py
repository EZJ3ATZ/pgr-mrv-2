# -*- coding: utf-8 -*-
"""A cadeia de custódia tem de puxar TODAS as medições do tubo — e o CNPJ.

Wesley, 11/08/2026: "A cadeia de custódia não puxa todas as medições realizadas
na planilha gerada e também não vem com o CNPJ da empresa preenchido."

Três furos provados no Postgres de produção:

1. Casamento do código era EXATO (`cqa.id_amostrador IN (codigo)`). O estoque
   guarda o código partido em tipo+codigo (id 191: tipo 'MTS', codigo '3091') e
   o técnico digita junto na planilha ('MTS3091') — a coleta não era achada e a
   linha saía quase vazia, só com o número do amostrador.

2. `setdefault` mantinha só a PRIMEIRA coleta de cada tubo. No EC98029A (Destak,
   29/07) há 2 coletas — Manganês e Ferro no mesmo filtro. O formulário do lab
   tem AGENTE 1..10 exatamente para isso; ia só o Manganês e o Ferro nunca era
   analisado.

3. CNPJ saía vazio porque só lia `empresas.cnpj`, que existe em 3 de 520
   empresas. O dado chega no texto da task do Planner (196 tasks cruas).
"""
import io

import openpyxl
import pytest

from controle.cadeia_custodia import (
    LINHA_1, cnpj_do_texto, coletar_dados, gerar_xlsx,
)
from controle.db import get_db, init_db


@pytest.fixture
def limpar():
    ids = []
    yield ids
    with get_db() as conn:
        for i in ids:
            conn.execute('DELETE FROM amostradores WHERE id=?', (i,))


def _tubo(conn, aid, codigo, tipo):
    conn.execute('DELETE FROM amostradores WHERE id=?', (aid,))
    conn.execute(
        "INSERT INTO amostradores (id, codigo, tipo, status, data_medicao, arquivado) "
        "VALUES (?,?,?,'laboratorio','2026-07-29',0)", (aid, codigo, tipo))


def _demanda(conn, did, titulo='OS teste', descricao='', empresa_id=None):
    """coletas_quimico.demanda_id tem FK — a OS precisa existir (e ela exige
    empresa_id NOT NULL)."""
    conn.execute('DELETE FROM demandas WHERE id=?', (did,))
    if empresa_id is None:
        empresa_id = conn.execute(
            "INSERT INTO empresas (nome) VALUES (?)", (f'EMP {did}',)).lastrowid
    conn.execute(
        "INSERT INTO demandas (id, titulo, descricao, status, empresa_id) "
        "VALUES (?,?,?,'aberta',?)", (did, titulo, descricao, empresa_id))
    return did


def _coleta(conn, demanda_id, func, subst, cod, vazao=0.2, hi='08:00', hf='16:00'):
    cur = conn.execute(
        "INSERT INTO coletas_quimico (demanda_id, data_coleta, substancias, "
        "nome_funcionario, funcao, setor, responsavel_coleta, status) "
        "VALUES (?,'2026-07-29',?,?,'Operador','Produção','Helbert','concluida')",
        (demanda_id, subst, func))
    cid = cur.lastrowid
    conn.execute(
        "INSERT INTO coletas_quimico_amostr (coleta_id, seq, id_amostrador, "
        "tipo_amostrador, substancia, vazao_media, hora_inicio, hora_final) "
        "VALUES (?,1,?,?,?,?,?,?)", (cid, cod, 'EC', subst, vazao, hi, hf))
    return cid


# ── 1. código partido no estoque × digitado junto na planilha ─────────

def test_codigo_partido_em_tipo_e_numero_ainda_acha_a_coleta(limpar):
    init_db()
    with get_db() as conn:
        _tubo(conn, -200, '3091', 'MTS')       # estoque: tipo + número
        limpar.append(-200)
        _demanda(conn, 9001)
        _coleta(conn, 9001, 'Josilene Lina', 'Ácido Peracético', 'MTS3091')
    d = coletar_dados([-200])
    ln = d['linhas'][0]
    assert ln['funcionario'] == 'Josilene Lina', \
        'a coleta digitada como MTS3091 não casou com o estoque MTS + 3091'
    assert ln['agentes'] == ['Ácido Peracético']
    assert ln['vazao'] == 0.2


def test_codigo_com_espaco_e_caixa_diferente_tambem_casa(limpar):
    init_db()
    with get_db() as conn:
        _tubo(conn, -201, 'PVC90U96', 'PVC')
        limpar.append(-201)
        _demanda(conn, 9002)
        _coleta(conn, 9002, 'Noé Costa', 'Poeira Total', ' pvc90u96 ')
    assert coletar_dados([-201])['linhas'][0]['funcionario'] == 'Noé Costa'


# ── 2. um tubo, vários agentes (o caso EC98029A) ──────────────────────

def test_duas_medicoes_no_mesmo_tubo_viram_dois_agentes_na_mesma_linha(limpar):
    init_db()
    with get_db() as conn:
        _tubo(conn, -202, 'EC98029A', 'EC')
        limpar.append(-202)
        _demanda(conn, 9003)
        _coleta(conn, 9003, 'José Francisco', 'Manganês', 'EC98029A')
        # nome sem vírgula de propósito: _agentes_da_linha trata ',' como
        # separador de agentes ("Tolueno, Xileno"), então "Ferro, Óxido (Fe2O3)"
        # entraria partido em dois. É comportamento antigo e assunto separado.
        _coleta(conn, 9003, 'José Francisco', 'Ferro', 'EC98029A')
    d = coletar_dados([-202])
    assert len(d['linhas']) == 1, 'é UM tubo físico → uma linha na cadeia'
    ln = d['linhas'][0]
    assert ln['coletas'] == 2
    assert ln['agentes'] == ['Manganês', 'Ferro'], \
        'o 2º agente do mesmo filtro não chegou à cadeia — o lab não analisaria'


def test_os_dois_agentes_saem_nas_colunas_do_formulario(limpar):
    init_db()
    with get_db() as conn:
        _tubo(conn, -203, 'EC97917A', 'EC')
        limpar.append(-203)
        _demanda(conn, 9004)
        _coleta(conn, 9004, 'José Francisco', 'Manganês', 'EC97917A')
        _coleta(conn, 9004, 'José Francisco', 'Ferro', 'EC97917A')
    d = coletar_dados([-203])
    wa = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(d)))['Dados Agentes']
    assert wa.cell(LINHA_1, 4).value == 'EC97917A'
    assert wa.cell(LINHA_1, 17).value == 'Manganês'                  # Q AGENTE 1
    assert wa.cell(LINHA_1, 18).value == 'Ferro'                     # R AGENTE 2


def test_escolha_do_tecnico_continua_vencendo_o_que_esta_no_banco(limpar):
    init_db()
    with get_db() as conn:
        _tubo(conn, -204, 'EC90001A', 'EC')
        limpar.append(-204)
        _demanda(conn, 9005)
        _coleta(conn, 9005, 'Alguém', 'Manganês', 'EC90001A')
        _coleta(conn, 9005, 'Alguém', 'Ferro', 'EC90001A')
    d = coletar_dados([-204], agentes_por_codigo={'EC90001A': ['Chumbo']})
    assert d['linhas'][0]['agentes'] == ['Chumbo']


# ── 3. CNPJ ───────────────────────────────────────────────────────────

@pytest.mark.parametrize('texto,esperado', [
    ('6425245 SGS GEOSOL Laboratórios Ltda 02.417.115/0001-01', '02.417.115/0001-01'),
    ('CNPJ 59116081000132 da empresa', '59.116.081/0001-32'),
    ('OS 6536301 MS Revestimento', ''),
    (None, ''),
])
def test_cnpj_pescado_do_texto_da_os(texto, esperado):
    assert cnpj_do_texto(texto) == esperado


def test_cnpj_vem_da_descricao_da_os_quando_o_cadastro_esta_vazio(limpar):
    init_db()
    with get_db() as conn:
        conn.execute("DELETE FROM demandas WHERE id=-300")
        cur = conn.execute("INSERT INTO empresas (nome) VALUES ('EMP SEM CNPJ')")
        emp = cur.lastrowid
        conn.execute(
            "INSERT INTO demandas (id, titulo, descricao, status, empresa_id) "
            "VALUES (-300,'6561748 Ferrosider','Medição — CNPJ 59.116.081/0001-32','aberta',?)",
            (emp,))
        _tubo(conn, -205, 'EC90002A', 'EC')
        limpar.append(-205)
        conn.execute('UPDATE amostradores SET empresa_id=? WHERE id=-205', (emp,))
    try:
        d = coletar_dados([-205], demanda_id=-300)
        assert d['empresa'].get('cnpj') == '59.116.081/0001-32'
        assert d['empresa'].get('cnpj_da_os') is True
        assert not any('CNPJ' in a for a in d['avisos'])
    finally:
        with get_db() as conn:
            conn.execute("DELETE FROM demandas WHERE id=-300")


def test_sem_cnpj_em_lugar_nenhum_a_cadeia_cobra(limpar):
    init_db()
    with get_db() as conn:
        cur = conn.execute("INSERT INTO empresas (nome) VALUES ('EMP SEM CNPJ 2')")
        emp = cur.lastrowid
        _tubo(conn, -206, 'EC90003A', 'EC')
        limpar.append(-206)
        conn.execute('UPDATE amostradores SET empresa_id=? WHERE id=-206', (emp,))
    d = coletar_dados([-206])
    assert any('CNPJ' in a for a in d['avisos']), \
        'CNPJ é (*) no formulário do lab — vazio tem de aparecer nos avisos'


def test_cnpj_do_cadastro_nunca_e_sobrescrito(limpar):
    init_db()
    with get_db() as conn:
        conn.execute("DELETE FROM demandas WHERE id=-301")
        cur = conn.execute(
            "INSERT INTO empresas (nome, cnpj) VALUES ('EMP COM CNPJ','11.111.111/0001-11')")
        emp = cur.lastrowid
        conn.execute(
            "INSERT INTO demandas (id, titulo, descricao, status, empresa_id) "
            "VALUES (-301,'OS','outro 59.116.081/0001-32','aberta',?)", (emp,))
        _tubo(conn, -207, 'EC90004A', 'EC')
        limpar.append(-207)
        conn.execute('UPDATE amostradores SET empresa_id=? WHERE id=-207', (emp,))
    try:
        d = coletar_dados([-207], demanda_id=-301)
        assert d['empresa']['cnpj'] == '11.111.111/0001-11'
        assert not d['empresa'].get('cnpj_da_os')
    finally:
        with get_db() as conn:
            conn.execute("DELETE FROM demandas WHERE id=-301")
