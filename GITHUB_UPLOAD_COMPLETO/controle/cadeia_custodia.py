# -*- coding: utf-8 -*-
"""Gerador da Cadeia de Custódia — formulário IT02-M do laboratório.

O laboratório (UniScientific) importa a cadeia ELETRONICAMENTE, então o
arquivo tem que sair no layout exato do formulário deles: preenchemos o
template `tpl/cadeia_custodia.xlsx`, não geramos um xlsx do zero.

Por que aqui e não numa tela de "gerar documento": comparando a cadeia da
Destak (preenchida à mão em 30/07/2026) com o banco, amostrador, funcionário,
função, setor, vazão e horários já batiam — o técnico estava redigitando o que
o sistema sabe. E o que NÃO batia expõe o motivo real:

  - `substancia` e `volume_l` estão vazios em 100% dos registros: nascem só na
    planilha, então o sistema nunca soube o que cada amostrador coleta.
  - técnico divergia (sistema: Helbert; cadeia: Rafael) e a data também
    (29/07 × 28/07) — duas versões do mesmo fato.

Gerar daqui faz o dado nascer na origem e carimba `data_envio_lab`, que é o
que o alerta de atraso precisa (ver `lab_inbox._alertar_nunca_despachados`).

O template original pesava 2,79 MB, dos quais ~95% eram `xl/externalLinks/`
— links quebrados para planilhas de outras máquinas, que se propagavam a cada
cópia (79 cadeias/ano × 2,79 MB). A versão limpa tem 147 KB.
"""
import io
import os
import re
import logging
from datetime import datetime

from .db import get_db, row_to_dict, agora_brt

log = logging.getLogger(__name__)

TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        'tpl', 'cadeia_custodia.xlsx')

ABA_EMPRESA = 'Dados Empresa'
ABA_AGENTES = 'Dados Agentes'

# Primeira linha de dados da grade de amostras (a 10 é o cabeçalho).
LINHA_1 = 11
MAX_AGENTES = 10          # colunas Q..Z = AGENTE 1..10

# Aba "Dados Empresa" — a coluna do rótulo e a do valor são fixas no formulário.
CEL_DATA_ENVIO = 'E7'
CEL_AVALIADA = {
    'razao_social': 'D19', 'cnpj': 'D20',   'cidade': 'L20',
    'endereco':     'D21', 'numero': 'C22', 'bairro': 'F22', 'uf': 'L22',
    'contato':      'D23', 'cep':    'L23',
    'fone':         'D24', 'email':  'L24',
}


def _fmt_data(v):
    """Qualquer coisa que pareça data → date (o Excel formata pelo template)."""
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v)[:10]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fmt_hora(v):
    """'14:45' / '14:45:00' → time. O lab rejeita hora como texto."""
    if not v:
        return None
    if hasattr(v, 'hour') and not isinstance(v, str):
        return v
    m = re.match(r'^\s*(\d{1,2})[:h](\d{2})', str(v))
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if 0 <= h <= 23 and 0 <= mi <= 59:
        from datetime import time
        return time(h, mi)
    return None


def _minutos(hora_ini, hora_fim):
    """Duração em minutos entre duas horas. Vira o dia se o fim for menor."""
    a, b = _fmt_hora(hora_ini), _fmt_hora(hora_fim)
    if not a or not b:
        return None
    ini, fim = a.hour * 60 + a.minute, b.hour * 60 + b.minute
    dur = fim - ini
    if dur < 0:
        dur += 24 * 60          # amostragem que atravessa a meia-noite
    return dur or None


def _volume(vazao, tempo_min, gravado, hora_ini=None, hora_fim=None):
    """Volume amostrado em litros = vazão (L/min) × tempo (min).

    `volume_l` E `tempo_min` estão zerados em 100% dos registros, mas as horas
    são gravadas — então o tempo sai delas. A conta confere com as cadeias
    preenchidas à mão: 2,0202 L/min × 60 min (14:45→15:45) = 121,212 L.
    """
    try:
        if gravado and float(gravado) > 0:
            return round(float(gravado), 4)
    except (TypeError, ValueError):
        pass
    try:
        t = float(tempo_min or 0)
    except (TypeError, ValueError):
        t = 0
    if t <= 0:
        t = _minutos(hora_ini, hora_fim) or 0
    try:
        v = float(vazao or 0)
        return round(v * t, 4) if v > 0 and t > 0 else None
    except (TypeError, ValueError):
        return None


def _norm_cod(v):
    """Código de amostrador comparável: sem espaço, maiúsculo."""
    return re.sub(r'\s+', '', str(v or '')).upper()


CNPJ_RX = re.compile(r'\b(\d{2})[.\s]?(\d{3})[.\s]?(\d{3})[/\s]?(\d{4})[-\s]?(\d{2})\b')


def cnpj_do_texto(*textos):
    """Primeiro CNPJ que aparecer nos textos, formatado 00.000.000/0000-00.

    A empresa vem do Planner e o CNPJ costuma estar no título ou na descrição da
    task (196 das tasks cruas em 11/08/2026) — o cadastro tem em 3 de 520. O lab
    marca o campo com (*), então vale pescar do texto em vez de pedir ao técnico
    o que o sistema já recebeu.
    """
    for t in textos:
        if not t:
            continue
        m = CNPJ_RX.search(str(t))
        if m:
            return '{}.{}.{}/{}-{}'.format(*m.groups())
    return ''


def _norm_agente(v):
    """Nome de agente comparável com o guia: sem quebra de linha, maiúsculo."""
    return re.sub(r'\s+', ' ', str(v or '').replace('\n', ' ')).strip().upper().rstrip('.')


_GUIA_NOMES = None


def _nomes_do_guia():
    """Nomes de agente do guia de métodos (408), normalizados. Cache em memória.

    Serve para saber se uma vírgula SEPARA dois agentes ou faz parte do nome de
    um só. Guia ausente → conjunto vazio, e aí nada é separado por vírgula: é o
    lado seguro (agente a mais na cadeia é análise a mais cobrada).
    """
    global _GUIA_NOMES
    if _GUIA_NOMES is None:
        try:
            import json as _j
            caminho = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                'guia_metodos.json')
            with open(caminho, encoding='utf-8') as f:
                guia = _j.load(f)
            _GUIA_NOMES = {_norm_agente(k) for k in (guia.get('by_name') or {})}
        except Exception as e:
            log.warning('[cadeia] guia de métodos indisponível: %s', e)
            _GUIA_NOMES = set()
    return _GUIA_NOMES


# ';' e '/' o técnico usa para LISTAR agentes; vírgula e ' e ' aparecem dentro
# do nome canônico (98 dos 408 agentes do guia têm vírgula) e só separam quando
# os dois lados são, eles próprios, agentes do guia.
_SEP_LISTA = re.compile(r'[;/]')
_SEP_DUVIDOSO = re.compile(r',|\s+e\s+', re.IGNORECASE)


def _agentes_da_linha(sub_gravada):
    """Agentes DESTA amostra — só o que foi efetivamente gravado na coleta.

    Deliberadamente NÃO cai para os agentes da OS: a OS lista o que a empresa
    contratou no total, não o que cada amostrador carrega. Preencher todos em
    todas as linhas mandaria o laboratório analisar N substâncias por amostra e
    multiplicaria o custo da análise. Quando falta, o campo sai vazio, entra em
    `avisos` e quem decide é o técnico — que é o único que sabe o que pôs em
    cada tubo.

    A vírgula era separador fixo, então "Ferro, óxido (Fe2O3)" ia ao laboratório
    como DOIS agentes e "Manganês elementar e compostos inorgânicos, como Mn"
    como TRÊS — cada um vira uma análise cobrada. Agora quem decide é o guia de
    métodos: só separa quando cada pedaço é um agente de verdade.
    """
    if not sub_gravada:
        return []
    catalogo = _nomes_do_guia()
    achados = []
    for bruto in _SEP_LISTA.split(str(sub_gravada)):
        parte = re.sub(r'\s+', ' ', bruto).strip()
        if not parte:
            continue
        pedacos = [p.strip() for p in _SEP_DUVIDOSO.split(parte) if p.strip()]
        if (len(pedacos) > 1
                and _norm_agente(parte) not in catalogo
                and all(_norm_agente(p) in catalogo for p in pedacos)):
            achados.extend(pedacos)          # "Tolueno, Xileno" → dois agentes
        else:
            achados.append(parte)            # nome com vírgula continua inteiro
    vistos, final = set(), []
    for a in achados:
        k = _norm_agente(a)
        if k and k not in vistos:
            vistos.add(k)
            final.append(a)
    return final[:MAX_AGENTES]


def coletar_dados(amostrador_ids, demanda_id=None, agentes_por_codigo=None):
    """Junta o que a cadeia precisa a partir das coletas já registradas.

    `agentes_por_codigo` = {'TCP2912AV3': ['Tolueno'], ...} — a escolha do
    técnico na tela, que vence tudo. Sem ela, só entra o que a coleta gravou.

    Retorna {'empresa', 'linhas', 'avisos', 'agentes_sugeridos'}. Os sugeridos
    são os agentes químicos da OS: servem para MONTAR a lista na tela, nunca
    para preencher sozinho (ver `_agentes_da_linha`).
    """
    if not amostrador_ids:
        return {'empresa': {}, 'linhas': [], 'avisos': ['Nenhum amostrador selecionado']}

    ph = ','.join(['?'] * len(amostrador_ids))
    with get_db() as conn:
        amostradores = [row_to_dict(r) for r in conn.execute(
            f"SELECT id, codigo, tipo, empresa_id, data_medicao, avaliador "
            f"FROM amostradores WHERE id IN ({ph}) ORDER BY codigo", amostrador_ids).fetchall()]
        if not amostradores:
            return {'empresa': {}, 'linhas': [], 'avisos': ['Amostradores não encontrados']}

        emp_id = demanda_id and conn.execute(
            'SELECT empresa_id FROM demandas WHERE id=?', (demanda_id,)).fetchone()
        emp_id = (row_to_dict(emp_id).get('empresa_id') if emp_id else None) \
            or next((a['empresa_id'] for a in amostradores if a.get('empresa_id')), None)
        empresa = {}
        if emp_id:
            row = conn.execute('SELECT * FROM empresas WHERE id=?', (emp_id,)).fetchone()
            if row:
                empresa = row_to_dict(row)

        # Coleta química por CÓDIGO do amostrador (o vínculo é textual).
        # Casamento NORMALIZADO e por duas formas: o estoque às vezes guarda o
        # código partido em tipo+codigo ('MTS' + '3091') e o técnico digita junto
        # na planilha ('MTS3091'). Comparação exata perdia a coleta e a linha da
        # cadeia saía vazia (Wesley 11/08: "não puxa todas as medições").
        chave_de = {}                      # chave normalizada -> código do estoque
        for a in amostradores:
            cod = str(a.get('codigo') or '').strip()
            if not cod:
                continue
            for k in {_norm_cod(cod), _norm_cod(str(a.get('tipo') or '') + cod)}:
                if k:
                    chave_de.setdefault(k, cod)
        coletas = {}                       # código do estoque -> [coletas]
        if chave_de:
            chaves = list(chave_de.keys())
            ph2 = ','.join(['?'] * len(chaves))
            for r in conn.execute(f"""
                SELECT cqa.id_amostrador, cqa.substancia, cqa.vazao_media, cqa.volume_l,
                       cqa.hora_inicio, cqa.hora_final, cqa.tempo_min, cqa.intervalos,
                       cq.nome_funcionario, cq.funcao, cq.setor, cq.responsavel_coleta,
                       cq.data_coleta, cq.observacao
                FROM coletas_quimico_amostr cqa
                JOIN coletas_quimico cq ON cq.id = cqa.coleta_id
                WHERE REPLACE(UPPER(TRIM(cqa.id_amostrador)), ' ', '') IN ({ph2})
                ORDER BY cqa.id""", chaves).fetchall():
                d = row_to_dict(r)
                dono = chave_de.get(_norm_cod(d.get('id_amostrador')))
                if dono:
                    coletas.setdefault(dono, []).append(d)

        sugeridos = _agentes_sugeridos(conn, demanda_id)
        cnpj_texto = _cnpj_da_demanda(conn, demanda_id) if not empresa.get('cnpj') else ''

    escolhidos = {_norm_cod(k): v for k, v in (agentes_por_codigo or {}).items()}
    linhas, avisos = [], []
    for a in amostradores:
        cod = str(a.get('codigo') or '').strip()
        cs = coletas.get(cod) or []
        c = cs[0] if cs else {}

        def _prim(campo):
            """Primeiro valor não vazio entre as coletas do MESMO tubo."""
            for x in cs:
                v = x.get(campo)
                if v not in (None, '', 0):
                    return v
            return c.get(campo)

        # Um tubo pode carregar MAIS DE UM agente: o mesmo filtro EC vai ao lab
        # para Manganês E Ferro (Destak, 29/07 — 2 coletas no EC98029A). O
        # formulário tem AGENTE 1..10 justamente para isso; antes só a primeira
        # coleta era lida e o lab não analisava o resto.
        subs = []
        for x in cs:
            for s in _agentes_da_linha(x.get('substancia')):
                if s not in subs:
                    subs.append(s)
        agentes = ([x for x in escolhidos.get(_norm_cod(cod), []) if str(x).strip()][:MAX_AGENTES]
                   or subs[:MAX_AGENTES])
        linha = {
            'codigo':      cod,
            'data':        _fmt_data(_prim('data_coleta') or a.get('data_medicao')),
            'funcionario': (_prim('nome_funcionario') or '').strip(),
            'funcao':      (_prim('funcao') or '').strip(),
            'setor':       (_prim('setor') or '').strip(),
            'tecnico':     (_prim('responsavel_coleta') or a.get('avaliador') or '').strip(),
            'vazao':       _prim('vazao_media'),
            'volume':      _volume(_prim('vazao_media'), _prim('tempo_min'), _prim('volume_l'),
                                   _prim('hora_inicio'), _prim('hora_final')),
            'hora_ini':    _fmt_hora(_prim('hora_inicio')),
            'hora_fim':    _fmt_hora(_prim('hora_final')),
            'intervalos':  (_prim('intervalos') or '').strip(),
            'obs':         (_prim('observacao') or '').strip(),
            'agentes':     agentes,
            'coletas':     len(cs),
        }
        if not cs:
            avisos.append(f'{cod}: sem coleta química registrada — linha quase vazia')
        if not agentes:
            avisos.append(f'{cod}: AGENTE não informado — escolha antes de enviar '
                          f'(obrigatório no formulário do laboratório)')
        if not linha['data']:
            avisos.append(f'{cod}: DATA DA AMOSTRAGEM vazia (obrigatório)')
        linhas.append(linha)

    if not empresa.get('nome'):
        avisos.append('Empresa avaliada sem razão social')
    # CNPJ é (*) no formulário do laboratório. Se o cadastro não tem, tenta o
    # texto da OS (o Planner costuma trazer) — e se nem lá, cobra do técnico.
    if not str(empresa.get('cnpj') or '').strip() and cnpj_texto:
        empresa = {**empresa, 'cnpj': cnpj_texto, 'cnpj_da_os': True}
    if not str(empresa.get('cnpj') or '').strip():
        avisos.append('CNPJ da empresa vazio — obrigatório (*) no formulário do '
                      'laboratório. Preencha aqui: fica salvo no cadastro.')
    return {'empresa': empresa, 'linhas': linhas, 'avisos': avisos,
            'agentes_sugeridos': sugeridos}


def _cnpj_da_demanda(conn, demanda_id):
    """CNPJ pescado do texto da OS (título/descrição/checklist da task)."""
    if not demanda_id:
        return ''
    try:
        row = conn.execute(
            'SELECT titulo, descricao, checklist, cnpj FROM demandas WHERE id=?',
            (demanda_id,)).fetchone()
        if not row:
            return ''
        d = row_to_dict(row)
        return (str(d.get('cnpj') or '').strip()
                or cnpj_do_texto(d.get('titulo'), d.get('descricao'), d.get('checklist')))
    except Exception as e:
        log.warning('[cadeia] cnpj da demanda falhou: %s', e)
        return ''


def _agentes_sugeridos(conn, demanda_id):
    """Agentes da OS pelo motor de extração — usados quando a coleta não gravou
    a substância (hoje, 100% dos casos)."""
    if not demanda_id:
        return []
    try:
        row = conn.execute(
            'SELECT titulo, descricao, checklist, planner_bucket FROM demandas WHERE id=?',
            (demanda_id,)).fetchone()
        if not row:
            return []
        d = row_to_dict(row)
        import json as _j
        try:
            cl = _j.loads(d.get('checklist') or '[]')
        except Exception:
            cl = []
        from .inteligencia_demandas import extrair_agentes_multifonte
        ags = extrair_agentes_multifonte(
            titulo=d.get('titulo') or '', descricao=d.get('descricao') or '',
            checklist=cl, bucket=d.get('planner_bucket') or '')
        # Só químico entra na cadeia: ruído/calor/vibração não vão ao laboratório.
        return [a.canonical for a in ags if a.confianca >= 0.55 and a.tipo == 'quimico']
    except Exception as e:
        log.warning('[cadeia] agentes sugeridos falharam: %s', e)
        return []


def gerar_xlsx(dados, data_envio=None):
    """Preenche o template e devolve os bytes do arquivo."""
    import openpyxl
    if not os.path.exists(TEMPLATE):
        raise FileNotFoundError(f'template não encontrado: {TEMPLATE}')
    wb = openpyxl.load_workbook(TEMPLATE)

    emp = dados.get('empresa') or {}
    we = wb[ABA_EMPRESA]
    we[CEL_DATA_ENVIO] = _fmt_data(data_envio) or datetime.now().date()
    for campo, cel in CEL_AVALIADA.items():
        valor = {
            'razao_social': emp.get('nome'), 'cnpj': emp.get('cnpj'),
            'cidade': emp.get('cidade'), 'endereco': emp.get('endereco'),
            'numero': emp.get('numero'), 'bairro': emp.get('bairro'),
            'uf': emp.get('uf'), 'contato': emp.get('contato'),
            'cep': emp.get('cep'), 'fone': emp.get('telefone'),
            'email': emp.get('email'),
        }.get(campo)
        if valor:
            we[cel] = valor

    wa = wb[ABA_AGENTES]
    for i, ln in enumerate(dados.get('linhas') or []):
        r = LINHA_1 + i
        wa.cell(r, 3, ln.get('data'))            # C  DATA AMOSTRAGEM
        wa.cell(r, 4, ln.get('codigo'))          # D  NÚMERO DO AMOSTRADOR
        wa.cell(r, 6, ln.get('funcionario'))     # F  NOME DO FUNCIONÁRIO
        wa.cell(r, 7, ln.get('funcao'))          # G  FUNÇÃO
        wa.cell(r, 8, ln.get('setor'))           # H  SETOR
        wa.cell(r, 9, ln.get('tecnico'))         # I  TÉCNICO RESPONSÁVEL
        if ln.get('vazao') is not None:
            wa.cell(r, 10, ln['vazao'])          # J  VAZÃO MÉDIA (L/min)
        if ln.get('volume') is not None:
            wa.cell(r, 11, ln['volume'])         # K  VOLUME AMOSTRADO (L)
        wa.cell(r, 12, ln.get('hora_ini'))       # L  INÍCIO
        wa.cell(r, 13, ln.get('hora_fim'))       # M  TÉRMINO
        if ln.get('obs'):
            wa.cell(r, 16, ln['obs'])            # P  OBSERVAÇÕES
        for j, ag in enumerate(ln.get('agentes') or []):
            wa.cell(r, 17 + j, ag)               # Q..Z  AGENTE 1..10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nome_arquivo(dados, data_envio=None):
    """Mesmo padrão da pasta de rede: 'Cadeia de Custodia - <EMPRESA> - <data>.xlsx'."""
    nome = ((dados.get('empresa') or {}).get('nome') or 'Empresa').strip()
    nome = re.sub(r'[\\/:*?"<>|]', '', nome)[:60].strip()
    d = _fmt_data(data_envio) or datetime.now().date()
    return f'Cadeia de Custodia - {nome} - {d.strftime("%d.%m.%Y")}.xlsx'


_CAMPOS_EMPRESA = ('cnpj', 'cidade', 'uf', 'endereco', 'numero', 'bairro',
                   'cep', 'telefone', 'email', 'contato')


def completar_empresa(empresa_id, dados):
    """Grava no cadastro o que o técnico digitou na tela da cadeia.

    O formulário do lab marca CNPJ com (*), mas 99% das empresas estão sem ele
    no banco (515 de 518) — e a cadeia manual também só preenchia razão social
    e CNPJ, o resto ficava vazio lá igual. Em vez de pedir o dado toda vez, o
    que for digitado uma vez fica no cadastro. Só PREENCHE o que está vazio:
    nunca sobrescreve dado existente.
    """
    if not empresa_id or not dados:
        return []
    sets, vals, tocados = [], [], []
    with get_db() as conn:
        atual = conn.execute('SELECT * FROM empresas WHERE id=?', (empresa_id,)).fetchone()
        if not atual:
            return []
        atual = row_to_dict(atual)
        for campo in _CAMPOS_EMPRESA:
            novo = str(dados.get(campo) or '').strip()
            if novo and not str(atual.get(campo) or '').strip():
                sets.append(f'{campo}=?')
                vals.append(novo)
                tocados.append(campo)
        if sets:
            conn.execute(f"UPDATE empresas SET {', '.join(sets)} WHERE id=?",
                         vals + [empresa_id])
    return tocados


def marcar_despacho(amostrador_ids, data_envio=None, dias_validade=45, lote=''):
    """Carimba `data_envio_lab` nos amostradores da cadeia.

    Gerar a cadeia É o ato de despachar — este é o gancho que faltava para a
    data existir. Sem ela nem o alerta de atraso nem o gargalo do dashboard
    enxergam o amostrador (eram 7 dos 8 em 30/07/2026).
    """
    if not amostrador_ids:
        return 0
    data = _fmt_data(data_envio) or agora_brt().date()
    ph = ','.join(['?'] * len(amostrador_ids))
    with get_db() as conn:
        cur = conn.execute(
            f"UPDATE amostradores SET data_envio_lab=?, dias_validade=?, lote=?,"
            f" status=CASE WHEN status IN ('disponivel','reservado') THEN 'laboratorio' ELSE status END,"
            f" atualizado_em=CURRENT_TIMESTAMP WHERE id IN ({ph})",
            [data.isoformat(), int(dias_validade or 45), lote or ''] + list(amostrador_ids))
        return getattr(cur, 'rowcount', 0) or len(amostrador_ids)
