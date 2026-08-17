# -*- coding: utf-8 -*-
"""Importadores das planilhas atuais para o banco."""
import io
import re
import unicodedata
from datetime import datetime, date

import openpyxl

from .db import (get_db, upsert_empresa, normalizar_status_amostrador,
                 status_amostrador_conhecido)


def _norm(s):
    if s is None: return ''
    return unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode('ascii').upper().strip()


def _str(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)):
        try: return v.strftime('%Y-%m-%d')
        except: return str(v)
    return str(v).strip()


def _to_int(v, default=0):
    try:
        s = _str(v).split(',')[0].split('.')[0]
        if not s: return default
        # FALTA 1, FALTA 2 -> 0 (sem dado de quantidade)
        if not s.lstrip('-').isdigit(): return default
        return int(s)
    except Exception:
        return default


def importar_amostradores(file_bytes):
    """Importa planilha CONTROLE AMOSTRADORES 2026.
    Espera abas '2026', '2025' ou similar com colunas:
    Status | TIPO DE AMOSTRADOR | CODIGO | DATA DE ENTRADA | Empresa | Avaliador | Data da medicao
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    inserted = 0
    updated = 0
    sheets_ok = 0   # abas com cabeçalho reconhecido (0 = arquivo errado)
    errors = []

    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) in ('DADOS',):  # aba de referencia
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue

        # Detectar cabecalho (primeira linha com STATUS)
        header_idx = None
        for i, row in enumerate(rows):
            txt = ' '.join(_norm(c) for c in row if c)
            if 'STATUS' in txt and 'AMOSTRADOR' in txt:
                header_idx = i
                break
        if header_idx is None:
            continue
        sheets_ok += 1

        header_norm = [_norm(c) for c in rows[header_idx]]

        def col_idx(*kws):
            for j, h in enumerate(header_norm):
                for kw in kws:
                    if kw in h:
                        return j
            return None

        c_status   = col_idx('STATUS')
        c_tipo     = col_idx('TIPO DE AMOSTRADOR', 'TIPO')
        c_codigo   = col_idx('CODIGO')
        c_entrada  = col_idx('DATA DE ENTRADA')
        c_empresa  = col_idx('EMPRESA')
        c_avaliad  = col_idx('AVALIADOR')
        c_medicao  = col_idx('DATA DA MEDICAO')

        for row in rows[header_idx + 1:]:
            codigo = _str(row[c_codigo]) if c_codigo is not None else ''
            tipo   = _str(row[c_tipo])   if c_tipo   is not None else ''
            if not codigo or not tipo:
                continue
            # Normaliza o status para o valor canonico ANTES de gravar — senao o
            # valor cru da planilha ("Estoque", "UTILIZADO?", nome de empresa...)
            # nao casa nenhum filtro e o amostrador some das telas ate o proximo boot.
            status_raw = _str(row[c_status]) if c_status is not None else ''
            status   = normalizar_status_amostrador(status_raw or 'Estoque')
            # ... mas so sobrescrevemos o status de um amostrador JA EXISTENTE
            # quando a planilha traz um valor reconhecido. Texto solto na coluna
            # Status virava 'laboratorio' no normalizador e ressuscitava
            # amostrador 'concluido' na fila do laboratorio.
            status_confiavel = status_amostrador_conhecido(status_raw)
            entrada  = _str(row[c_entrada])  if c_entrada  is not None else ''
            empresa  = _str(row[c_empresa])  if c_empresa  is not None else ''
            avaliad  = _str(row[c_avaliad])  if c_avaliad  is not None else ''
            medicao  = _str(row[c_medicao])  if c_medicao  is not None else ''
            empresa_id = upsert_empresa('', empresa) if empresa and empresa != '-' else None

            with get_db() as conn:
                existing = conn.execute(
                    'SELECT id FROM amostradores WHERE codigo = ? AND tipo = ?',
                    (codigo, tipo)
                ).fetchone()
                if existing:
                    # Celula vazia NAO apaga dado do banco: a planilha e sugestao de
                    # campo, o estado real vem do app (reserva, envio ao lab, RA).
                    sets = ['atualizado_em=CURRENT_TIMESTAMP']
                    vals = []
                    if status_confiavel:
                        sets.append('status=?');       vals.append(status)
                    if entrada:
                        sets.append('data_entrada=?'); vals.append(entrada)
                    if empresa_id is not None:
                        sets.append('empresa_id=?');   vals.append(empresa_id)
                    if avaliad:
                        sets.append('avaliador=?');    vals.append(avaliad)
                    if medicao:
                        sets.append('data_medicao=?'); vals.append(medicao)
                    vals.append(existing['id'])
                    conn.execute(
                        'UPDATE amostradores SET ' + ', '.join(sets) + ' WHERE id=?',
                        tuple(vals))
                    updated += 1
                else:
                    try:
                        conn.execute("""
                            INSERT INTO amostradores
                                (codigo, tipo, status, data_entrada, empresa_id, avaliador, data_medicao)
                            VALUES (?, ?, ?, ?, ?, ?, ?)""",
                            (codigo, tipo, status, entrada, empresa_id, avaliad, medicao))
                        inserted += 1
                    except Exception as e:
                        errors.append(f'Linha codigo={codigo}: {e}')

    return {'inserted': inserted, 'updated': updated,
            'sheets_reconhecidas': sheets_ok, 'errors': errors}


def _extrair_nome_empresa(nome_tarefa):
    """Extrai nome da empresa de 'MEDIÇÕES - OS - Empresa LTDA' ou 'OS - Empresa'."""
    if not nome_tarefa: return ''
    s = nome_tarefa.strip()
    # Remover prefixos comuns
    for prefix in ['MEDIÇÕES - ', 'MEDICOES - ', 'MEDIÇÃO - ']:
        if s.upper().startswith(prefix.upper()):
            s = s[len(prefix):]
    # Tirar OS no inicio (numero + hifen ou espaco)
    s = re.sub(r'^\d{4,8}\s*[-–]\s*', '', s)
    s = re.sub(r'^\d{4,8}\s+', '', s)
    # Tirar restos
    s = re.sub(r'\s+\-\s+MEDI[CÇÃ]', ' ', s, flags=re.I)
    return s.strip(' -,;')


def importar_demandas_planner(file_bytes):
    """Importa formato 'Demandas_Medicoes_Completo' (16 colunas) do Microsoft Planner.
    Aceita tambem formato antigo (4 colunas) por retrocompat.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    inseridas = 0
    atualizadas = 0
    sheets_ok = 0   # abas com cabeçalho reconhecido (0 = arquivo errado)
    erros = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue

        # Achar cabecalho
        header_idx = None
        for i, row in enumerate(rows):
            txt = ' '.join(_norm(c) for c in row if c)
            if ('NOME' in txt or 'EMPRESA' in txt) and ('CRIACAO' in txt or 'PRAZO' in txt or 'STATUS' in txt):
                header_idx = i; break
        if header_idx is None: continue
        sheets_ok += 1

        header_norm = [_norm(c) for c in rows[header_idx]]
        def col_idx(*kws):
            for j, h in enumerate(header_norm):
                for kw in kws:
                    if kw in h:
                        return j
            return None

        c_nome    = col_idx('NOME DA TAREFA', 'NOME DA EMPRESA', 'NOME')
        c_os      = col_idx('OS', 'NUMERO OS')
        c_cri     = col_idx('DATA DE CRIACAO', 'CRIACAO')
        c_prazo   = col_idx('DATA DE PRAZO', 'PRAZO', 'PREVISAO')
        c_concl   = col_idx('DATA DE CONCLUSAO', 'CONCLUSAO')
        c_resp    = col_idx('RESPONSAVEL', 'RESPONSAVEIS')
        c_status  = col_idx('STATUS')
        c_prog    = col_idx('PROGRESSO')
        c_check   = col_idx('CHECKLIST', 'CHECKLIST PROGRESSO')
        c_checkp  = None
        for j, h in enumerate(header_norm):
            if 'CHECKLIST' in h and 'PROGRESSO' in h:
                c_checkp = j; break
        c_bucket  = col_idx('BUCKET', 'GRUPO')
        c_etiq    = col_idx('ETIQUETAS')
        c_desc    = col_idx('DESCRICAO')
        c_cnpj    = col_idx('CNPJ')
        c_com     = col_idx('COMENTARIOS', 'TEM COMENTARIOS')

        for row in rows[header_idx + 1:]:
            if not row or (c_nome is not None and not row[c_nome]): continue
            nome_tarefa = _str(row[c_nome]) if c_nome is not None else ''
            os_num      = _str(row[c_os])      if c_os      is not None else ''
            data_cri    = _str(row[c_cri])     if c_cri     is not None else ''
            prazo       = _str(row[c_prazo])   if c_prazo   is not None else ''
            data_concl  = _str(row[c_concl])   if c_concl   is not None else ''
            resp        = _str(row[c_resp])    if c_resp    is not None else ''
            status_p    = _str(row[c_status])  if c_status  is not None else ''
            try:
                prog    = int(float(_str(row[c_prog]) or '0')) if c_prog is not None else 0
            except: prog = 0
            checklist   = _str(row[c_check])   if c_check   is not None else ''
            check_prog  = _str(row[c_checkp])  if c_checkp  is not None else ''
            bucket      = _str(row[c_bucket])  if c_bucket  is not None else ''
            etiq        = _str(row[c_etiq])    if c_etiq    is not None else ''
            desc        = _str(row[c_desc])    if c_desc    is not None else ''
            cnpj_val    = _str(row[c_cnpj])    if c_cnpj    is not None else ''
            tem_com     = _str(row[c_com])     if c_com     is not None else ''
            tem_com_int = 1 if tem_com and tem_com.lower() not in ('nao', 'no', '-', '0', '') else 0

            if cnpj_val == '-': cnpj_val = ''

            # Limpar
            for x in ['nome_tarefa','os_num','data_cri','prazo','data_concl','resp','status_p','checklist','check_prog','bucket','etiq','desc','cnpj_val']:
                pass

            # Extrair nome empresa do nome da tarefa
            nome_empresa = _extrair_nome_empresa(nome_tarefa)
            if not nome_empresa: continue

            # Status interno (concluido vs pendente)
            status_interno = 'concluida' if (status_p.lower().startswith('conclu') or prog >= 100) else 'pendente'

            empresa_id = upsert_empresa(cnpj_val, nome_empresa, contato=resp)
            if not empresa_id: continue

            with get_db() as conn:
                # Atualiza CNPJ na empresa se vier do Planner
                if cnpj_val:
                    conn.execute("UPDATE empresas SET cnpj=? WHERE id=? AND (cnpj IS NULL OR cnpj='')",
                                 (cnpj_val, empresa_id))

                # Match por OS
                existing = None
                if os_num and os_num not in ('-', ''):
                    existing = conn.execute('SELECT id FROM demandas WHERE numero_os=?', (os_num,)).fetchone()
                if not existing and nome_tarefa:
                    existing = conn.execute('SELECT id FROM demandas WHERE nome_tarefa=?', (nome_tarefa,)).fetchone()

                params = (
                    nome_tarefa, os_num or '', empresa_id, prazo, data_concl, resp, status_p,
                    prog, checklist, check_prog, bucket, etiq, desc, cnpj_val, tem_com_int,
                    status_interno, _parse_data(data_cri) or None,
                )
                if existing:
                    conn.execute("""
                        UPDATE demandas SET
                          nome_tarefa=?, numero_os=?, empresa_id=?, prazo=?,
                          data_conclusao=?, responsavel=?, status_planner=?,
                          progresso=?, checklist=?, checklist_prog=?, bucket=?,
                          etiquetas=?, descricao=?, cnpj=?, tem_comentarios=?,
                          status=?, criado_em=COALESCE(NULLIF(?, ''), criado_em)
                        WHERE id=?""",
                        params + (existing['id'],))
                    atualizadas += 1
                else:
                    conn.execute("""
                        INSERT INTO demandas
                          (nome_tarefa, numero_os, empresa_id, prazo, data_conclusao,
                           responsavel, status_planner, progresso, checklist,
                           checklist_prog, bucket, etiquetas, descricao, cnpj,
                           tem_comentarios, status, criado_em)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        params)
                    inseridas += 1

    return {'demandas_inseridas': inseridas, 'demandas_atualizadas': atualizadas,
            'sheets_reconhecidas': sheets_ok, 'errors': erros}


def _parse_data(s):
    """Converte 'dd/mm/aaaa' para 'aaaa-mm-dd HH:MM:SS' (ISO SQLite). Mantem se ja for ISO."""
    if not s: return ''
    s = s.strip()
    if not s: return ''
    # dd/mm/aaaa
    if len(s) == 10 and s[2] == '/' and s[5] == '/':
        try:
            d, m, y = s.split('/')
            return f'{y}-{m.zfill(2)}-{d.zfill(2)} 12:00:00'
        except: return ''
    return s


def _is_finalizada(cell):
    """Detecta se a celula tem fundo VERDE (= demanda finalizada).
    Considera verde:
      - Theme 9 (accent6) qualquer tint - padrao do template "Verde - Accent 6"
      - Theme 6 (accent3) qualquer tint - verde escuro
      - RGB direto com componente G dominante (G alto, R e B baixos)
    """
    try:
        fill = cell.fill
        if not fill or fill.patternType != 'solid':
            return False
        fg = fill.fgColor
        if fg.type == 'theme' and fg.theme in (6, 9):
            return True
        if fg.type == 'rgb' and fg.rgb:
            rgb = str(fg.rgb)
            if len(rgb) == 8: rgb = rgb[2:]  # remove alpha
            if len(rgb) == 6:
                r = int(rgb[0:2], 16); g = int(rgb[2:4], 16); b = int(rgb[4:6], 16)
                # Verde: G alto, R e B menores
                if g > 150 and g > r + 30 and g > b + 30:
                    return True
    except Exception:
        pass
    return False


# Ordem de avanco do status de medicao — usado no re-import para nunca REGREDIR
# uma medicao que o app ja finalizou (check-out do tecnico, RA do laboratorio).
_RANK_STATUS_MED = {'pendente': 0, 'parcial': 1, 'realizado': 2}


def importar_medicoes(file_bytes):
    """Importa planilha Controle de Medicoes - Helbert e Wesley.
    Cada linha eh um agente de uma OS. Agrupa por (OS, empresa) em demandas
    e cria 1 medicao por agente.

    DETECCAO DE COR: Se a celula da UNIDADE estiver pintada de VERDE,
    a medicao eh marcada como 'realizado' (demanda ja foi finalizada).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    inserted_demandas = 0
    inserted_medicoes = 0
    updated_medicoes = 0
    sheets_ok = 0   # abas com cabeçalho reconhecido (0 = arquivo errado)
    finalizadas_por_cor = 0
    demandas_tocadas = set()   # so estas entram no recalculo de status no fim
    errors = []

    for sheet_name in wb.sheetnames:
        if _norm(sheet_name) in ('DADOS',):
            continue
        ws = wb[sheet_name]
        # Pegar rows como CELULAS (nao values_only) para ler formato
        all_cells = list(ws.iter_rows(values_only=False))
        if not all_cells: continue

        # Detectar cabecalho
        header_idx = None
        for i, cells in enumerate(all_cells):
            txt = ' '.join(_norm(c.value) for c in cells if c.value is not None)
            if 'UNIDADE' in txt and 'AGENTE' in txt:
                header_idx = i
                break
        if header_idx is None:
            continue
        sheets_ok += 1

        header_norm = [_norm(c.value) for c in all_cells[header_idx]]
        def col_idx(*kws):
            for j, h in enumerate(header_norm):
                for kw in kws:
                    if kw in h:
                        return j
            return None

        c_unid    = col_idx('UNIDADE')
        c_os      = col_idx('NUMERO DA O.S', 'NUMERO DA OS', 'NUMERO O.S', 'NUMERO OS')
        c_resp    = col_idx('RESPONSAVEL')
        c_agente  = col_idx('AGENTE')
        c_amostr  = col_idx('AMOSTRADOR')
        c_pontos  = col_idx('QUANTIDADE DE PONTOS', 'QTD PONTOS')
        c_aval    = col_idx('PONTOS AVALIADOS')
        c_laudar  = col_idx('LAUDAR')
        c_obs     = col_idx('OBSERVAC')

        # Cache de demandas por (empresa_id, os)
        demanda_cache = {}

        for cells in all_cells[header_idx + 1:]:
            def cv(idx):
                if idx is None or idx >= len(cells): return ''
                return _str(cells[idx].value)
            unidade = cv(c_unid)
            os_num  = cv(c_os)
            agente  = cv(c_agente)
            if not (unidade or os_num) or not agente:
                continue
            resp    = cv(c_resp)
            amostr  = cv(c_amostr)
            pontos_raw = cells[c_pontos].value if c_pontos is not None and c_pontos < len(cells) else None
            pontos  = _to_int(pontos_raw, 1)
            aval    = _to_int(cells[c_aval].value if c_aval is not None and c_aval < len(cells) else None, 0)
            laudar  = cv(c_laudar)
            obs     = cv(c_obs)

            # DETECCAO DE COR VERDE = finalizada
            # Checa varias celulas da linha (unidade, agente, amostrador, pontos)
            verde = False
            for ci in (c_unid, c_agente, c_amostr, c_pontos):
                if ci is not None and ci < len(cells) and _is_finalizada(cells[ci]):
                    verde = True
                    break

            # Parse responsavel: nome (tel) email
            contato = resp.split('(')[0].strip() if resp else ''
            tel = ''
            tm = re.search(r'\((\d{2})\)\s*([\d\-\s]+)', resp)
            if tm: tel = f'({tm.group(1)}) {tm.group(2).strip()}'
            email = ''
            em = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', resp)
            if em: email = em.group()

            empresa_id = upsert_empresa('', unidade,
                contato=contato, telefone=tel, email=email)
            if not empresa_id:
                continue

            key = (empresa_id, os_num)
            if key not in demanda_cache:
                with get_db() as conn:
                    existing = conn.execute(
                        'SELECT id FROM demandas WHERE empresa_id=? AND numero_os=?',
                        (empresa_id, os_num)).fetchone()
                    if existing:
                        demanda_cache[key] = existing['id']
                    else:
                        cur = conn.execute(
                            'INSERT INTO demandas (numero_os, empresa_id, status, observacao) '
                            'VALUES (?, ?, ?, ?)',
                            (os_num, empresa_id, 'pendente', obs))
                        demanda_cache[key] = cur.lastrowid
                        inserted_demandas += 1
            demandas_tocadas.add(demanda_cache[key])

            # Insere medicao
            # Status: verde (planilha) > aval==pontos > parcial > pendente
            if verde:
                status_med = 'realizado'
                qtd_feita = pontos if pontos > 0 else 1
                finalizadas_por_cor += 1
            elif aval >= pontos and pontos > 0:
                status_med = 'realizado'
                qtd_feita = aval
            elif aval > 0:
                status_med = 'parcial'
                qtd_feita = aval
            else:
                status_med = 'pendente'
                qtd_feita = 0

            tipo_amostr = ''
            am = re.search(r'\b([A-Z]{2,5})\b', amostr.upper()) if amostr else None
            if am: tipo_amostr = am.group(1)

            with get_db() as conn:
                # Dedup por (demanda_id, agente, tipo_amostrador): re-import da mesma
                # planilha ATUALIZA em vez de duplicar (antes dobrava as medições e
                # inflava os contadores de todas as telas).
                ja = conn.execute(
                    "SELECT id, status, qtd_pontos_feita FROM medicoes "
                    "WHERE demanda_id=? AND agente=? "
                    "AND COALESCE(tipo_amostrador,'')=COALESCE(?,'')",
                    (demanda_cache[key], agente, tipo_amostr)).fetchone()
                if ja:
                    # A planilha e SUGESTAO de campo; o estado real da medicao vem do
                    # app (check-out do tecnico, RA do laboratorio, planner_sync).
                    # Por isso o re-import so AVANCA: nunca reverte medicao ja
                    # finalizada, nao zera pontos feitos e nao apaga observacao
                    # escrita no sistema com celula vazia da planilha.
                    sets, vals = [], []
                    if pontos_raw not in (None, ''):
                        sets.append('qtd_pontos_prevista=?'); vals.append(pontos)
                    if _RANK_STATUS_MED.get(status_med, 0) > \
                       _RANK_STATUS_MED.get(ja['status'] or 'pendente', 0):
                        sets.append('status=?'); vals.append(status_med)
                    if qtd_feita > (ja['qtd_pontos_feita'] or 0):
                        sets.append('qtd_pontos_feita=?'); vals.append(qtd_feita)
                    if laudar:
                        sets.append('necessita_laudo=?'); vals.append(laudar[:1])
                    if obs:
                        sets.append('observacao=?'); vals.append(obs)
                    if sets:
                        vals.append(ja['id'])
                        conn.execute(
                            'UPDATE medicoes SET ' + ', '.join(sets) + ' WHERE id=?',
                            tuple(vals))
                    updated_medicoes += 1
                else:
                    conn.execute("""
                        INSERT INTO medicoes
                            (demanda_id, agente, tipo_amostrador, qtd_pontos_prevista,
                             qtd_pontos_feita, necessita_laudo, status, observacao)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (demanda_cache[key], agente, tipo_amostr, pontos, qtd_feita,
                         laudar[:1] if laudar else '', status_med, obs))
                    inserted_medicoes += 1

    # Atualizar status das demandas TOCADAS por este import: se TODAS medicoes
    # realizadas -> concluida.
    # Antes este UPDATE era GLOBAL e rodava fora do loop de abas, ou seja em TODA
    # chamada: subir o arquivo errado (0 abas reconhecidas, rota devolvendo 400)
    # recalculava o status de toda a base. Agora so mexe nas demandas que
    # apareceram na planilha, e so quando alguma aba foi reconhecida.
    if sheets_ok and demandas_tocadas:
        ids = sorted(demandas_tocadas)
        marcadores = ','.join('?' * len(ids))
        with get_db() as conn:
            conn.execute(f"""
                UPDATE demandas SET status='concluida'
                WHERE id IN (
                    SELECT d.id FROM demandas d
                    WHERE d.id IN ({marcadores})
                    AND NOT EXISTS (
                        SELECT 1 FROM medicoes m
                        WHERE m.demanda_id = d.id AND m.status != 'realizado'
                    )
                    AND EXISTS (SELECT 1 FROM medicoes m WHERE m.demanda_id = d.id)
                )
            """, tuple(ids))

    return {
        'demandas_inseridas': inserted_demandas,
        'medicoes_inseridas': inserted_medicoes,
        'medicoes_atualizadas': updated_medicoes,
        'sheets_reconhecidas': sheets_ok,
        'finalizadas_por_cor_verde': finalizadas_por_cor,
        'errors': errors
    }
